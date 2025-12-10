# -*- coding: utf-8 -*-
"""
БЫСТРЫЙ ПАРСЕР ЦЕН WILDBERRIES - ТОЛЬКО API
Получает ВСЕ данные через API за секунды (без браузера!)
"""

import os
import json
import requests
from datetime import datetime
from openpyxl import load_workbook
from dotenv import load_dotenv
import time

# === КОНФИГУРАЦИЯ ===
EXCEL_FILE = "Парсер цен.xlsx"
SHEET_INPUT_WB = "Данные для парсера ВБ"
SHEET_OUTPUT_WB = "Парсер ВБ"

# API ENDPOINTS
WB_PRICES_API_URL = "https://discounts-prices-api.wildberries.ru/api/v2/list/goods/filter"
WB_CONTENT_API_URL = "https://content-api.wildberries.ru/content/v2/get/cards/list"

# Названия кабинетов (для .env файла)
CABINET_NAMES = ["COSMO", "MMA", "MAB", "MAU", "DREAMLAB", "BEAUTYLAB"]

# === ФУНКЦИИ ===

def load_api_keys_from_env():
    """
    Загружает API ключи WB из .env файла
    Формат в .env:
    COSMO=eyJhbGc...
    MMA=eyJhbGc...
    и т.д.
    """
    load_dotenv()  # Загружаем .env файл
    
    api_keys = []
    cabinet_info = []
    
    for cabinet_name in CABINET_NAMES:
        api_key = os.getenv(cabinet_name, "").strip()
        if api_key:
            api_keys.append(api_key)
            cabinet_info.append(cabinet_name)
    
    print(f"\n[API] Загружено из .env файла:")
    print(f"    Кабинетов: {len(api_keys)}")
    for name in cabinet_info:
        print(f"      ✓ {name}")
    
    return api_keys, cabinet_info


def get_product_info(articles, api_keys_list, cabinet_names=None):
    """
    Получает информацию о товарах через Content API
    Возвращает словарь {артикул: {название, nmID, vendorCode, cabinet}}
    """
    print("\n[API] Загрузка информации о товарах (названия, ID)...")
    
    if not api_keys_list:
        print("[!] API ключи не найдены!")
        return {}
    
    product_info = {}
    cabinet_map = {}  # nmID -> cabinet
    
    # Пробуем каждый API ключ
    for idx, api_key in enumerate(api_keys_list, 1):
        cabinet_name = cabinet_names[idx-1] if cabinet_names and idx-1 < len(cabinet_names) else f"Кабинет {idx}"
        print(f"\n[API] {cabinet_name} ({idx}/{len(api_keys_list)})...")
        
        try:
            headers = {
                "Authorization": api_key,
                "Content-Type": "application/json"
            }
            
            # Конвертируем артикулы в set для поиска
            articles_set = {str(art).strip() for art in articles}
            
            # Content API: получаем список карточек с пагинацией (максимум 100 за раз)
            cursor_updatedAt = ""
            cursor_nmID = 0
            total_found_this_cabinet = 0
            page = 0
            
            while True:
                page += 1
                
                payload = {
                    "settings": {
                        "cursor": {
                            "limit": 100
                        },
                        "filter": {
                            "withPhoto": -1
                        }
                    }
                }
                
                # Добавляем курсор для пагинации (если не первая страница)
                if cursor_updatedAt and cursor_nmID:
                    payload["settings"]["cursor"]["updatedAt"] = cursor_updatedAt
                    payload["settings"]["cursor"]["nmID"] = cursor_nmID
                
                response = requests.post(WB_CONTENT_API_URL, headers=headers, json=payload, timeout=30)
                
                if response.status_code == 200:
                    data = response.json()
                    
                    cards = data.get("cards", [])
                    if not cards and "data" in data:
                        cards = data.get("data", {}).get("cards", [])
                    
                    if not cards:
                        # Нет больше карточек
                        break
                    
                    # Обрабатываем карточки
                    for card in cards:
                        nm_id = str(card.get("nmID", ""))
                        vendor_code = str(card.get("vendorCode", ""))
                        
                        # Проверяем совпадение по nmID или vendorCode
                        if nm_id in articles_set or vendor_code in articles_set:
                            # Берем название (может быть в разных полях)
                            title = card.get("title") or card.get("object") or f"Товар {nm_id}"
                            
                            # Используем nmID как ключ
                            if nm_id:
                                product_info[nm_id] = {
                                    "title": title,
                                    "nmID": nm_id,
                                    "vendorCode": vendor_code,
                                    "cabinet": cabinet_name
                                }
                                total_found_this_cabinet += 1
                    
                    # Получаем курсор для следующей страницы
                    cursor_data = data.get("cursor", {})
                    cursor_updatedAt = cursor_data.get("updatedAt", "")
                    cursor_nmID = cursor_data.get("nmID", 0)
                    
                    # Если курсор пустой - больше страниц нет
                    if not cursor_updatedAt or not cursor_nmID:
                        break
                    
                    # Если нашли все нужные товары - можно остановиться
                    if len([x for x in product_info if str(x) in articles_set]) >= len(articles_set):
                        break
                    
                    time.sleep(0.2)  # Пауза между запросами пагинации
                
                else:
                    print(f"[!] Ошибка Content API: {response.status_code}")
                    print(f"    {response.text[:200]}")
                    break
            
            print(f"    Обработано страниц: {page}, найдено товаров: {total_found_this_cabinet}")
            time.sleep(0.3)
        
        except Exception as e:
            print(f"[!] Ошибка при запросе Content API (кабинет {idx}): {e}")
    
    print(f"\n[API] Итого загружено информации о {len(product_info)} товарах")
    return product_info


def get_prices_full_info(articles, api_keys_list, cabinet_names=None):
    """
    Получает ВСЕ цены через Prices API - ДО и ПОСЛЕ СПП!
    Возвращает словарь {артикул: {price_original, price_before_spp, price_after_spp, discount, spp, stocks}}
    
    Структура цен WB API:
    - price: базовая цена (без скидок)
    - discountedPrice: цена после обычных скидок (ДО СПП)
    - clubDiscountedPrice: финальная цена (ПОСЛЕ СПП и скидок кошелька)
    - stocks: остатки товара
    """
    print("\n[API] Загрузка цен ДО и ПОСЛЕ СПП через API...")
    
    if not api_keys_list:
        print("[!] API ключи не найдены!")
        return {}
    
    prices_info = {}
    
    for idx, api_key in enumerate(api_keys_list, 1):
        cabinet_name = cabinet_names[idx-1] if cabinet_names and idx-1 < len(cabinet_names) else f"Кабинет {idx}"
        print(f"\n[API] {cabinet_name} ({idx}/{len(api_keys_list)})...")
        
        try:
            headers = {
                "Authorization": api_key,
                "Content-Type": "application/json"
            }
            
            # Обрабатываем батчами по 1000
            batch_size = 1000
            
            for i in range(0, len(articles), batch_size):
                batch = articles[i:i + batch_size]
                nm_ids = [int(art) for art in batch if str(art).isdigit()]
                
                if not nm_ids:
                    continue
                
                # Правильный формат для Prices API
                payload = {
                    "limit": 1000,
                    "offset": 0,
                    "nmList": nm_ids  # ВАЖНО: nmList а не filterNmID!
                }
                
                response = requests.post(WB_PRICES_API_URL, headers=headers, json=payload, timeout=30)
                
                if response.status_code == 200:
                    data = response.json()
                    
                    # Парсим товары
                    goods_list = []
                    if "data" in data and "listGoods" in data["data"]:
                        goods_list = data["data"]["listGoods"]
                    elif "listGoods" in data:
                        goods_list = data["listGoods"]
                    
                    for item in goods_list:
                        nm_id = str(item.get("nmID", ""))
                        
                        # Берем данные из первого размера
                        sizes = item.get("sizes", [])
                        if sizes and len(sizes) > 0:
                            size_data = sizes[0]
                            
                            # Цены в рублях (целые числа)
                            price_original = size_data.get("price", 0)  # Базовая цена
                            price_discounted = size_data.get("discountedPrice", 0)  # После обычных скидок (ДО СПП)
                            price_club = size_data.get("clubDiscountedPrice", 0)  # После СПП (финальная)
                            
                            # Остатки - пробуем разные поля
                            stocks = size_data.get("stocks", 0) or size_data.get("wh", 0) or item.get("stocks", 0)
                            
                            # Проценты
                            discount_percent = item.get("discount", 0)  # Обычная скидка
                            club_discount_percent = item.get("clubDiscount", 0)  # СПП
                            
                            # Если нет цены после скидок, используем базовую
                            if not price_discounted and price_original:
                                price_discounted = price_original
                            
                            # Если нет клубной цены, используем цену после скидок
                            if not price_club and price_discounted:
                                price_club = price_discounted
                            
                            if nm_id:
                                prices_info[nm_id] = {
                                    "price_original": float(price_original) if price_original else 0,
                                    "price_before_spp": float(price_discounted) if price_discounted else 0,
                                    "price_after_spp": float(price_club) if price_club else 0,
                                    "discount": float(discount_percent) if discount_percent else 0,
                                    "spp": float(club_discount_percent) if club_discount_percent else 0,
                                    "stocks": int(stocks) if stocks else 0
                                }
                    
                    print(f"    Найдено {len(goods_list)} товаров в этом кабинете")
                
                else:
                    print(f"[!] Ошибка Prices API: {response.status_code}")
                    if response.status_code != 404:
                        print(f"    {response.text[:200]}")
                
                time.sleep(0.3)
        
        except Exception as e:
            print(f"[!] Ошибка при запросе Prices API (кабинет {idx}): {e}")
            import traceback
            traceback.print_exc()
    
    print(f"\n[API] Итого загружено цен для {len(prices_info)} товаров")
    return prices_info


def parse_wb_fast_api(wb, api_keys, cabinet_names=None):
    """
    БЫСТРЫЙ парсинг WB - ТОЛЬКО через API!
    Получает: название, nmID, цену до СПП, цену после СПП
    """
    print("\n" + "="*80)
    print("БЫСТРЫЙ ПАРСИНГ WB - ТОЛЬКО API (БЕЗ БРАУЗЕРА!)")
    print("="*80)
    
    # Загрузка артикулов
    ws_in = wb[SHEET_INPUT_WB]
    ws_out = wb[SHEET_OUTPUT_WB]
    
    articles = []
    for row in ws_in.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            article = str(row[0]).strip()
            articles.append(article)
    
    total = len(articles)
    print(f"\n[1/4] Найдено артикулов: {total}")
    
    if total == 0:
        print("[!] Нет артикулов для обработки!")
        return
    
    start_time = time.time()
    
    # Шаг 1: Получаем информацию о товарах (названия)
    print("\n[2/4] Получение информации о товарах через Content API...")
    product_info_dict = get_product_info(articles, api_keys, cabinet_names)
    
    print(f"\n[DEBUG] Загружено товаров с инфо: {len(product_info_dict)}")
    if len(product_info_dict) > 0:
        first_key = list(product_info_dict.keys())[0]
        print(f"[DEBUG] Пример товара: {first_key} = {product_info_dict[first_key]}")
    
    # Шаг 2: Получаем цены (до и после СПП)
    print("\n[3/4] Получение цен через Prices API...")
    prices_dict = get_prices_full_info(articles, api_keys, cabinet_names)
    
    print(f"\n[DEBUG] Загружено цен: {len(prices_dict)}")
    if len(prices_dict) > 0:
        first_key = list(prices_dict.keys())[0]
        print(f"[DEBUG] Пример цены: {first_key} = {prices_dict[first_key]}")
    
    # Шаг 3: Очищаем старые данные
    print(f"\n[4/5] Очистка старых записей...")
    
    # Удаляем все строки кроме заголовка
    if ws_out.max_row > 1:
        ws_out.delete_rows(2, ws_out.max_row)
        print(f"    ✓ Удалено старых записей: {ws_out.max_row - 1}")
    
    # Создаем заголовки
    ws_out.append(["Дата", "Кабинет", "Артикул", "Название", "Цена До СПП", "Наличие", "Цена После СПП", "СПП %", "Скидка %"])
    
    # Шаг 4: Объединяем данные и сохраняем
    print(f"\n[5/5] Сохранение результатов...")
    print("="*80)
    
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    success = 0
    failed = 0
    
    for i, article in enumerate(articles, 1):
        # Получаем данные
        info = product_info_dict.get(article, {})
        prices = prices_dict.get(article, {})
        
        title = info.get("title", "Не найдено")
        nm_id = info.get("nmID", article)
        cabinet = info.get("cabinet", "Неизвестно")
        
        price_original = prices.get("price_original", 0)
        price_before = prices.get("price_before_spp", 0)
        price_after = prices.get("price_after_spp", 0)
        discount = prices.get("discount", 0)
        spp = prices.get("spp", 0)
        stocks = prices.get("stocks", 0)
        
        # Считаем процент скидки СПП (от цены до СПП к цене после)
        spp_percent_calc = None
        if price_before and price_after and price_before > 0:
            spp_percent_calc = ((price_before - price_after) / price_before) * 100
        
        print(f"[{i}/{total}] [{cabinet}] {nm_id} | {title[:40]}")
        print(f"         До СПП: {price_before}₽ → После СПП: {price_after}₽ (СПП: {spp}%) | Остаток: {stocks} шт")
        
        # Статус наличия (столбец F)
        if stocks > 0:
            stock_status = f"{stocks} шт"
        else:
            stock_status = "Нет в наличии"
        
        if price_before or price_after:
            # Сохраняем: Дата | Кабинет | Артикул | Название | Цена До СПП | Наличие | Цена После СПП | СПП % | Скидка %
            new_row = [
                timestamp,
                cabinet,
                nm_id,
                title,
                price_before if price_before else None,
                stock_status,  # Столбец F - наличие
                price_after if price_after else None,
                spp_percent_calc if spp_percent_calc else spp,
                discount if discount else None
            ]
            ws_out.append(new_row)
            success += 1
        else:
            print(f"         [!] Цены не найдены")
            # Сохраняем и без цен
            new_row = [
                timestamp,
                cabinet,
                nm_id,
                title,
                None,
                stock_status,  # Столбец F - наличие
                None,
                None,
                None
            ]
            ws_out.append(new_row)
            failed += 1
    
    # Итоги
    elapsed = time.time() - start_time
    print(f"\n{'='*80}")
    print("ГОТОВО!")
    print(f"{'='*80}")
    print(f"Всего артикулов: {total}")
    print(f"Успешно обработано: {success}")
    print(f"Не найдено: {failed}")
    print(f"Время выполнения: {elapsed:.1f} сек ({elapsed/60:.2f} мин)")
    print(f"Скорость: {total/elapsed:.1f} артикулов/сек")
    print(f"{'='*80}")
    
    wb.save(EXCEL_FILE)
    print(f"\n[SAVE] ✓ Результаты сохранены в '{EXCEL_FILE}'")


def main():
    print("\n" + "!"*80)
    print("БЫСТРЫЙ ПАРСЕР WB - ТОЛЬКО API")
    print("!"*80)
    print("\nОсобенности:")
    print("  ✓ Работает БЕЗ браузера - только API запросы")
    print("  ✓ В 50-100 раз быстрее обычного парсинга")
    print("  ✓ Получает: название, ID, цену ДО СПП, цену ПОСЛЕ СПП")
    print("  ✓ Обрабатывает все 6 магазинов одновременно")
    print("\nТребования:")
    print("  1. ЗАКРОЙТЕ Excel файл перед запуском")
    print("  2. API ключи WB в файле .env (6 кабинетов)")
    print("  3. Артикулы в листе 'Данные для парсера ВБ' (столбец A)")
    print("!"*80)
    
    input("\n💡 Нажмите Enter чтобы начать...")
    
    # Загружаем API ключи из .env
    api_keys, cabinet_names = load_api_keys_from_env()
    
    if not api_keys:
        print("\n[!] ОШИБКА: Не найдено ни одного API ключа в .env файле!")
        print("\n📝 Создайте файл .env в той же папке со скриптом:")
        print("    COSMO=ваш_api_ключ_1")
        print("    MMA=ваш_api_ключ_2")
        print("    MAB=ваш_api_ключ_3")
        print("    MAU=ваш_api_ключ_4")
        print("    DREAMLAB=ваш_api_ключ_5")
        print("    BEAUTYLAB=ваш_api_ключ_6")
        return
    
    # Загружаем Excel
    try:
        wb = load_workbook(EXCEL_FILE)
    except Exception as e:
        print(f"\n[!] Ошибка открытия файла '{EXCEL_FILE}': {e}")
        print("    Убедитесь что файл существует и закрыт!")
        return
    
    try:
        # Быстрый парсинг через API
        parse_wb_fast_api(wb, api_keys, cabinet_names)
        
        print("\n" + "="*80)
        print("✓ ВСЕ ЗАДАЧИ ВЫПОЛНЕНЫ УСПЕШНО!")
        print("="*80)
        
    except Exception as e:
        print(f"\n[!] ОШИБКА: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        wb.close()
        print("\n[DONE] Завершено!")


if __name__ == "__main__":
    main()

