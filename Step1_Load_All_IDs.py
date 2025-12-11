# -*- coding: utf-8 -*-
"""
ШАГ 1: ЗАГРУЗКА ВСЕХ АРТИКУЛОВ ИЗ ВСЕХ КАБИНЕТОВ
Загружает все nmID из 6 кабинетов и записывает в Excel
"""

import os
import requests
from openpyxl import load_workbook, Workbook
from dotenv import load_dotenv
import time

# === КОНФИГУРАЦИЯ ===
EXCEL_FILE = "Парсер цен.xlsx"
SHEET_INPUT_WB = "Данные для парсера ВБ"

WB_CONTENT_API_URL = "https://content-api.wildberries.ru/content/v2/get/cards/list"
CABINET_NAMES = ["COSMO", "MMA", "MAB", "MAU", "DREAMLAB", "BEAUTYLAB"]

# === ФУНКЦИИ ===

def load_api_keys_from_env():
    """Загружает API ключи из .env"""
    load_dotenv()
    
    api_keys = []
    cabinet_info = []
    
    for cabinet_name in CABINET_NAMES:
        api_key = os.getenv(cabinet_name, "").strip()
        if api_key:
            api_keys.append(api_key)
            cabinet_info.append(cabinet_name)
    
    print(f"\n[API] Загружено из .env файла:")
    print(f"    Кабинетов: {len(api_keys)}")
    for name in cabinet_info:
        print(f"      ✓ {name}")
    
    return api_keys, cabinet_info


def get_all_nmids_from_cabinet(api_key, cabinet_name):
    """Получает ВСЕ nmID из одного кабинета"""
    print(f"\n[{cabinet_name}] Загрузка артикулов...")
    
    nm_ids = []
    headers = {
        "Authorization": api_key,
        "Content-Type": "application/json"
    }
    
    cursor_updatedAt = ""
    cursor_nmID = 0
    page = 0
    
    try:
        while True:
            page += 1
            
            payload = {
                "settings": {
                    "cursor": {
                        "limit": 100
                    },
                    "filter": {
                        "withPhoto": -1
                    }
                }
            }
            
            if cursor_updatedAt and cursor_nmID:
                payload["settings"]["cursor"]["updatedAt"] = cursor_updatedAt
                payload["settings"]["cursor"]["nmID"] = cursor_nmID
            
            response = requests.post(WB_CONTENT_API_URL, headers=headers, json=payload, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                
                cards = data.get("cards", [])
                if not cards and "data" in data:
                    cards = data.get("data", {}).get("cards", [])
                
                if not cards:
                    break
                
                # Собираем nmID
                for card in cards:
                    nm_id = str(card.get("nmID", ""))
                    if nm_id and nm_id.isdigit():
                        nm_ids.append(nm_id)
                
                print(f"    Страница {page}: +{len(cards)} товаров (всего: {len(nm_ids)})")
                
                # Курсор
                cursor_data = data.get("cursor", {})
                cursor_updatedAt = cursor_data.get("updatedAt", "")
                cursor_nmID = cursor_data.get("nmID", 0)
                
                if not cursor_updatedAt or not cursor_nmID:
                    break
                
                time.sleep(0.2)
            
            elif response.status_code == 401:
                print(f"    [!] Ошибка 401: Неверный API ключ")
                break
            else:
                print(f"    [!] Ошибка {response.status_code}")
                break
        
        print(f"    ✓ Загружено {len(nm_ids)} артикулов из {cabinet_name}")
        
    except Exception as e:
        print(f"    [!] Ошибка: {e}")
    
    return nm_ids


def main():
    print("\n" + "="*80)
    print("ШАГ 1: ЗАГРУЗКА ВСЕХ АРТИКУЛОВ ИЗ КАБИНЕТОВ")
    print("="*80)
    print("\n📋 Этот скрипт:")
    print("  1. Загрузит ВСЕ артикулы (nmID) из всех 6 кабинетов")
    print("  2. Запишет их в лист 'Данные для парсера ВБ'")
    print("  3. Удалит дубликаты")
    print("\n💡 После этого можете запустить Parser_WB_API_FAST.py")
    print("="*80)
    
    input("\n💡 Нажмите Enter чтобы начать...")
    
    # Загружаем API ключи
    api_keys, cabinet_names = load_api_keys_from_env()
    
    if not api_keys:
        print("\n[!] ОШИБКА: Не найдено API ключей!")
        return
    
    start_time = time.time()
    
    # Загружаем артикулы из всех кабинетов
    print("\n" + "="*80)
    print("ЗАГРУЗКА АРТИКУЛОВ")
    print("="*80)
    
    all_nm_ids = []
    
    for api_key, cabinet_name in zip(api_keys, cabinet_names):
        nm_ids = get_all_nmids_from_cabinet(api_key, cabinet_name)
        all_nm_ids.extend(nm_ids)
    
    print(f"\n✓ Всего загружено артикулов: {len(all_nm_ids)}")
    
    # Удаляем дубликаты
    unique_nm_ids = list(set(all_nm_ids))
    unique_nm_ids.sort()
    
    print(f"✓ Уникальных артикулов: {len(unique_nm_ids)}")
    print(f"  (удалено дубликатов: {len(all_nm_ids) - len(unique_nm_ids)})")
    
    if not unique_nm_ids:
        print("\n[!] Не найдено ни одного артикула!")
        return
    
    # Записываем в Excel
    print("\n" + "="*80)
    print("СОХРАНЕНИЕ В EXCEL")
    print("="*80)
    
    try:
        # Открываем или создаем Excel
        try:
            wb = load_workbook(EXCEL_FILE)
        except FileNotFoundError:
            wb = Workbook()
        
        # Создаем или очищаем лист
        if SHEET_INPUT_WB in wb.sheetnames:
            # Удаляем старый лист
            del wb[SHEET_INPUT_WB]
        
        # Создаем новый лист
        ws = wb.create_sheet(SHEET_INPUT_WB, 0)  # Вставляем первым
        
        # Заголовок
        ws.append(["Артикулы WB (nmID)"])
        
        # Записываем артикулы
        for i, nm_id in enumerate(unique_nm_ids, 1):
            ws.append([nm_id])
            
            if i % 100 == 0:
                print(f"    Записано: {i}/{len(unique_nm_ids)}")
        
        wb.save(EXCEL_FILE)
        print(f"\n✓ Сохранено {len(unique_nm_ids)} артикулов в '{EXCEL_FILE}'")
        print(f"  Лист: '{SHEET_INPUT_WB}'")
        
    except Exception as e:
        print(f"\n[!] Ошибка при сохранении: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Итоги
    elapsed = time.time() - start_time
    
    print("\n" + "="*80)
    print("ГОТОВО!")
    print("="*80)
    print(f"Загружено артикулов: {len(unique_nm_ids)}")
    print(f"Время выполнения: {elapsed:.1f} сек ({elapsed/60:.1f} мин)")
    print("="*80)
    
    print("\n📊 Статистика по кабинетам:")
    for cabinet_name in cabinet_names:
        count = len([x for x in all_nm_ids if x])
        print(f"  {cabinet_name}: загружено артикулов")
    
    print("\n" + "="*80)
    print("🎯 СЛЕДУЮЩИЙ ШАГ:")
    print("="*80)
    print("\n1. Откройте файл 'Парсер цен.xlsx'")
    print(f"2. Проверьте лист '{SHEET_INPUT_WB}' - там {len(unique_nm_ids)} артикулов")
    print("3. При желании отредактируйте список (удалите ненужные)")
    print("4. ЗАКРОЙТЕ Excel файл")
    print("5. Запустите: python Parser_WB_API_FAST.py")
    print("\n   ↓ Parser_WB_API_FAST.py получит цены для всех артикулов!")
    print("="*80)
    
    print("\n[DONE] Завершено!")


if __name__ == "__main__":
    main()



