# -*- coding: utf-8 -*-
"""
ПАРСЕР ЦЕН WILDBERRIES - ПОДКЛЮЧЕНИЕ К ОТКРЫТОМУ CHROME
Подключается к уже запущенному Chrome через remote debugging
"""

import os
import time
import random
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from openpyxl import load_workbook
from datetime import datetime
import re

# Конфигурация
EXCEL_FILE = "Парсер цен.xlsx"
SHEET_INPUT = "Данные для парсера ВБ"
SHEET_OUTPUT = "Парсер ВБ"

# Remote debugging port
CHROME_DEBUG_PORT = 9222


def setup_remote_chrome_driver():
    """
    Подключается к уже запущенному Chrome через remote debugging
    """
    chrome_options = Options()
    chrome_options.add_experimental_option("debuggerAddress", f"127.0.0.1:{CHROME_DEBUG_PORT}")
    
    # Минимальные опции - только подключение
    # Не добавляем excludeSwitches и другие опции которые могут конфликтовать
    
    try:
        driver = webdriver.Chrome(options=chrome_options)
        return driver
    except Exception as e:
        print(f"\n[!] ОШИБКА подключения к Chrome:")
        print(f"    {e}")
        print(f"\n💡 Убедись что Chrome запущен с параметром --remote-debugging-port={CHROME_DEBUG_PORT}")
        print(f"💡 Проверь что ChromeDriver совместим с версией Chrome")
        return None


def human_delay():
    """Случайная задержка как у человека"""
    delay = random.uniform(2, 5)
    time.sleep(delay)


def parse_wb_product_page(driver, nm_id):
    """
    Парсит карточку товара WB
    Возвращает словарь с ценами
    """
    url = f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx"
    
    try:
        print(f"\n[{nm_id}] Открываю карточку...")
        
        # Открываем в новой вкладке чтобы не потерять текущие
        driver.execute_script(f"window.open('{url}', '_blank');")
        driver.switch_to.window(driver.window_handles[-1])
        
        # Ждём загрузки
        time.sleep(3)
        
        # Проверяем на captcha
        if "Почти готово" in driver.title or "captcha" in driver.page_source.lower():
            print(f"  ⚠ Captcha! Подожди 10 сек...")
            time.sleep(10)
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            return None
        
        result = {
            'nmID': str(nm_id),
            'name': '',
            'techSizeName': '',
            'price': 0,
            'discountedPrice': 0,
            'clubDiscountedPrice': 0,
            'discount': 0,
            'clubDiscount': 0,
            'stockCount': 0
        }
        
        # 1. Название товара
        try:
            title_selectors = [
                (By.CLASS_NAME, "product-page__title"),
                (By.CSS_SELECTOR, "h1[class*='title']"),
                (By.TAG_NAME, "h1")
            ]
            
            for by, selector in title_selectors:
                try:
                    title_elem = driver.find_element(by, selector)
                    result['name'] = title_elem.text.strip()
                    print(f"  ✓ Название: {result['name'][:50]}...")
                    break
                except:
                    continue
        except:
            pass
        
        # 2. Финальная цена (ПОСЛЕ СПП)
        try:
            price_selectors = [
                (By.CLASS_NAME, "price-block__final-price"),
                (By.CSS_SELECTOR, "span[class*='final-price']"),
                (By.CSS_SELECTOR, "ins[class*='price']"),
                (By.CSS_SELECTOR, "span[class*='wallet-price']")
            ]
            
            for by, selector in price_selectors:
                try:
                    price_elem = driver.find_element(by, selector)
                    price_text = price_elem.text.strip()
                    # Извлекаем число
                    price_num = re.sub(r'[^\d]', '', price_text)
                    if price_num:
                        result['clubDiscountedPrice'] = int(price_num)
                        print(f"  ✓ Цена ПОСЛЕ СПП: {result['clubDiscountedPrice']} ₽")
                        break
                except:
                    continue
        except:
            pass
        
        # 3. Старая цена (БЕЗ скидок)
        try:
            old_price_selectors = [
                (By.CLASS_NAME, "price-block__old-price"),
                (By.CSS_SELECTOR, "del[class*='price']"),
                (By.CSS_SELECTOR, "span[class*='old-price']")
            ]
            
            for by, selector in old_price_selectors:
                try:
                    old_price_elem = driver.find_element(by, selector)
                    old_price_text = old_price_elem.text.strip()
                    price_num = re.sub(r'[^\d]', '', old_price_text)
                    if price_num:
                        result['price'] = int(price_num)
                        print(f"  ✓ Базовая цена: {result['price']} ₽")
                        break
                except:
                    continue
        except:
            pass
        
        # Если нет старой цены, базовая = финальная
        if result['price'] == 0 and result['clubDiscountedPrice'] > 0:
            result['price'] = result['clubDiscountedPrice']
        
        # 4. Процент скидки (общий)
        try:
            discount_selectors = [
                (By.CLASS_NAME, "price-block__sale-percent"),
                (By.CSS_SELECTOR, "span[class*='percent']"),
                (By.CSS_SELECTOR, "span[class*='sale']")
            ]
            
            for by, selector in discount_selectors:
                try:
                    discount_elem = driver.find_element(by, selector)
                    discount_text = discount_elem.text.strip()
                    discount_num = re.sub(r'[^\d]', '', discount_text)
                    if discount_num:
                        total_discount = int(discount_num)
                        print(f"  ✓ Общая скидка: {total_discount}%")
                        
                        # Пробуем найти СПП отдельно
                        try:
                            spp_elems = driver.find_elements(By.CSS_SELECTOR, "span[class*='club'], span[class*='spp']")
                            for spp_elem in spp_elems:
                                text = spp_elem.text.strip()
                                if 'СПП' in text or 'клуб' in text.lower():
                                    spp_num = re.sub(r'[^\d]', '', text)
                                    if spp_num:
                                        result['clubDiscount'] = int(spp_num)
                                        result['discount'] = total_discount - result['clubDiscount']
                                        break
                        except:
                            pass
                        
                        # Если СПП не нашли, весь процент = обычная скидка
                        if result['clubDiscount'] == 0:
                            result['discount'] = total_discount
                        
                        break
                except:
                    continue
        except:
            pass
        
        # 5. Расчёт цены ДО СПП
        if result['price'] > 0 and result['discount'] > 0:
            result['discountedPrice'] = int(result['price'] * (1 - result['discount'] / 100))
        else:
            result['discountedPrice'] = result['clubDiscountedPrice']
        
        # 6. Размер
        try:
            size_selectors = [
                (By.CLASS_NAME, "product-params__row"),
                (By.CSS_SELECTOR, "span[class*='size']")
            ]
            
            for by, selector in size_selectors:
                try:
                    size_elem = driver.find_element(by, selector)
                    size_text = size_elem.text.strip()
                    if 'Размер' in size_text or 'размер' in size_text:
                        result['techSizeName'] = size_text.replace('Размер:', '').strip()
                        break
                except:
                    continue
        except:
            pass
        
        # 7. Наличие
        try:
            stock_keywords = ['наличи', 'остал', 'stock']
            page_text = driver.page_source.lower()
            
            for keyword in stock_keywords:
                if keyword in page_text:
                    result['stockCount'] = 1
                    break
        except:
            pass
        
        print(f"  ✓ Цена ДО СПП (расчёт): {result['discountedPrice']} ₽")
        print(f"  ✓ Скидка: {result['discount']}%, СПП: {result['clubDiscount']}%")
        
        # Закрываем вкладку и возвращаемся к первой
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        
        return result
    
    except Exception as e:
        print(f"  ✗ Ошибка: {e}")
        # Пытаемся закрыть вкладку и вернуться
        try:
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
        except:
            pass
        return None


def main():
    print("\n" + "="*80)
    print("ПАРСЕР ЦЕН WB - ПОДКЛЮЧЕНИЕ К ОТКРЫТОМУ CHROME")
    print("="*80)
    
    # Загружаем Excel
    wb = load_workbook(EXCEL_FILE)
    ws_in = wb[SHEET_INPUT]
    ws_out = wb[SHEET_OUTPUT]
    
    # Загружаем артикулы
    articles = []
    for row in ws_in.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            articles.append(str(row[0]).strip())
    
    print(f"\n[1/3] Найдено артикулов: {len(articles)}")
    
    # ТЕСТ: парсим только первые 5 для проверки
    test_count = min(5, len(articles))
    print(f"\n⚠ ТЕСТОВЫЙ РЕЖИМ: парсим первые {test_count} артикулов")
    articles = articles[:test_count]
    
    # Подключаемся к Chrome
    print(f"\n[2/3] Подключение к Chrome (port {CHROME_DEBUG_PORT})...")
    
    # Проверяем доступность порта
    try:
        response = requests.get(f"http://127.0.0.1:{CHROME_DEBUG_PORT}/json", timeout=2)
        if response.status_code == 200:
            print(f"    ✓ Порт {CHROME_DEBUG_PORT} доступен")
        else:
            print(f"    ⚠ Порт {CHROME_DEBUG_PORT} отвечает с кодом {response.status_code}")
    except:
        print(f"    ✗ Порт {CHROME_DEBUG_PORT} НЕ доступен!")
        print(f"\n💡 Chrome НЕ запущен в debug-режиме!")
        print(f"💡 Запусти Chrome через команду:")
        print(f'\n   "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port={CHROME_DEBUG_PORT} --user-data-dir="%LOCALAPPDATA%\\Google\\Chrome\\User Data" --profile-directory="Profile 2"')
        print(f"\n   Или используй: .\\START_CHROME_DEBUG.bat")
        return
    
    driver = setup_remote_chrome_driver()
    
    if not driver:
        print("\n" + "="*80)
        print("ИНСТРУКЦИЯ: Как запустить Chrome в debug-режиме")
        print("="*80)
        print("\n1. Закрой весь Chrome (Ctrl+Shift+Q)")
        print("\n2. Нажми Win+R и вставь эту команду:\n")
        print('   "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222 --user-data-dir="%LOCALAPPDATA%\\Google\\Chrome\\User Data" --profile-directory="Profile 2"')
        print("\n3. Нажми Enter - Chrome откроется с твоими вкладками")
        print("\n4. Запусти этот скрипт снова")
        print("\n" + "="*80)
        return
    
    print("    ✓ Подключено к Chrome")
    print(f"    ✓ Открыто вкладок: {len(driver.window_handles)}")
    
    try:
        # Парсим товары
        print(f"\n[3/3] Парсинг товаров...")
        results = []
        
        for i, article in enumerate(articles, 1):
            print(f"\n{'='*60}")
            print(f"[{i}/{len(articles)}] Артикул: {article}")
            
            result = parse_wb_product_page(driver, article)
            
            if result:
                results.append(result)
                print(f"  ✓ УСПЕХ")
            else:
                print(f"  ✗ НЕ УДАЛОСЬ")
            
            # Задержка между товарами
            if i < len(articles):
                delay = random.uniform(3, 7)
                print(f"\n  [пауза {delay:.1f}с перед следующим товаром]")
                time.sleep(delay)
        
        # Сохраняем результаты
        print(f"\n{'='*80}")
        print("СОХРАНЕНИЕ РЕЗУЛЬТАТОВ")
        print(f"{'='*80}")
        
        # Очищаем лист
        if ws_out.max_row >= 1:
            ws_out.delete_rows(1, ws_out.max_row)
        
        # Заголовки
        ws_out.append([
            "Дата",
            "nmID",
            "Название (name)",
            "Размер (techSizeName)",
            "price",
            "discountedPrice",
            "clubDiscountedPrice",
            "discount %",
            "clubDiscount %",
            "Наличие"
        ])
        
        # Данные
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        for result in results:
            ws_out.append([
                timestamp,
                result['nmID'],
                result['name'],
                result['techSizeName'],
                result['price'] if result['price'] > 0 else None,
                result['discountedPrice'] if result['discountedPrice'] > 0 else None,
                result['clubDiscountedPrice'] if result['clubDiscountedPrice'] > 0 else None,
                result['discount'] if result['discount'] > 0 else None,
                result['clubDiscount'] if result['clubDiscount'] > 0 else None,
                "В наличии" if result['stockCount'] > 0 else "Нет данных"
            ])
        
        # Автофильтр
        ws_out.auto_filter.ref = ws_out.dimensions
        
        wb.save(EXCEL_FILE)
        
        print(f"\n✓ Сохранено: {len(results)} товаров")
        print(f"✓ Файл: {EXCEL_FILE}")
        
    except Exception as e:
        print(f"\n[!] КРИТИЧЕСКАЯ ОШИБКА: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        print(f"\n[!] Chrome остаётся открытым - закрой вручную когда закончишь работу")
    
    print(f"\n{'='*80}")
    print("ЗАВЕРШЕНО")
    print(f"{'='*80}\n")


if __name__ == "__main__":
    main()

