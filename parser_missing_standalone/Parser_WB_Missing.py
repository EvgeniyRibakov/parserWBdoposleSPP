# -*- coding: utf-8 -*-
"""
ПАРСЕР ЦЕН WILDBERRIES - ДОПАРСИНГ НЕДОСТАЮЩИХ АРТИКУЛОВ
Парсит цены для конкретного списка артикулов и сохраняет в Google Таблицы и Excel файл

ИНСТРУКЦИЯ:
1. Убедитесь что Chrome закрыт (или используйте remote режим)
2. Настройте google-credentials.json для работы с Google Таблицами
3. Запустите: python Parser_WB_Missing.py
4. Подтвердите логин и адрес доставки в браузере
5. Результаты сохраняются в Google Таблицы после каждых 20 товаров
6. Финальные результаты также сохраняются в data/missing_articles_results.xlsx
"""

import os
import sys
import time
import random
import re
import subprocess
import shutil
from selenium import webdriver

# Загрузка переменных окружения из .env файла
try:
    from dotenv import load_dotenv
    PROJECT_ROOT_TEMP = os.path.dirname(os.path.abspath(__file__))
    env_path = os.path.join(PROJECT_ROOT_TEMP, '.env')
    if os.path.exists(env_path):
        load_dotenv(env_path)
        print(f"[ЛОГ] Загружены настройки из .env файла")
except ImportError:
    print("[ЛОГ] python-dotenv не установлен, используются настройки по умолчанию")
except Exception as e:
    print(f"[ЛОГ] Ошибка загрузки .env: {e}, используются настройки по умолчанию")

# Настройка кодировки консоли для Windows
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass

from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, InvalidSessionIdException
from openpyxl import Workbook, load_workbook
import undetected_chromedriver as uc

# Конфигурация
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(PROJECT_ROOT, "data")

# Функция для чтения настроек из .env
def get_env_bool(key, default=False):
    value = os.getenv(key, str(default)).strip().lower()
    return value in ('true', '1', 'yes', 'on')

def get_env_int(key, default=0):
    try:
        return int(os.getenv(key, str(default)))
    except:
        return default

def get_env_str(key, default=""):
    return os.getenv(key, default)

# Настройки браузера
BROWSER_TYPE = get_env_str("BROWSER_TYPE", "chrome").lower()
HEADLESS_MODE = get_env_bool("HEADLESS_MODE", False)  # По умолчанию False (видимый режим)
USE_REMOTE_CHROME = get_env_bool("USE_REMOTE_CHROME", False)
CHROME_DEBUG_PORT = get_env_int("CHROME_DEBUG_PORT", 9222)
USE_TEMP_PROFILE = get_env_bool("USE_TEMP_PROFILE", True)
TEMP_PROFILE_DIR = os.path.join(PROJECT_ROOT, "chrome_parser_profile")
COPY_PROFILE_DATA = get_env_bool("COPY_PROFILE_DATA", True)
SOURCE_PROFILE_FOR_COPY = get_env_str("SOURCE_PROFILE_FOR_COPY", "Profile 4")
WAIT_FOR_MANUAL_LOGIN = get_env_bool("WAIT_FOR_MANUAL_LOGIN", True)
MANUAL_LOGIN_TIMEOUT = get_env_int("MANUAL_LOGIN_TIMEOUT", 120)

# Настройки парсинга
PARALLEL_TABS = get_env_int("PARALLEL_TABS", 20)

# Google Таблицы
GOOGLE_SHEETS_ENABLED = get_env_bool("GOOGLE_SHEETS_ENABLED", True)
GOOGLE_SHEET_URL = get_env_str("GOOGLE_SHEET_URL", "https://docs.google.com/spreadsheets/d/1fbMPHE43ikYM90gcSVk_kcUItjzo-OsYI3T25yOJgQU/edit")
GOOGLE_SHEET_NAME = get_env_str("GOOGLE_SHEET_NAME", "Лист1")
GOOGLE_SERVICE_ACCOUNT_FILE = get_env_str("GOOGLE_SERVICE_ACCOUNT_FILE", "google-credentials.json")

# Файл с артикулами и ссылками (ищем в корне проекта, а не в папке скрипта)
PROJECT_ROOT_PARENT = os.path.dirname(PROJECT_ROOT)  # Корень проекта (на уровень выше)
ARTICLES_EXCEL_FILE = os.path.join(PROJECT_ROOT_PARENT, get_env_str("ARTICLES_EXCEL_FILE", "Articles.xlsx"))
# Если файл не найден в корне проекта, пробуем в папке скрипта
if not os.path.exists(ARTICLES_EXCEL_FILE):
    ARTICLES_EXCEL_FILE = os.path.join(PROJECT_ROOT, get_env_str("ARTICLES_EXCEL_FILE", "Articles.xlsx"))
# Возможные имена листов
sheet_names_str = get_env_str("POSSIBLE_SHEET_NAMES", "Данные для парсера ВБ,WBarticules,WB,Артикулы,Sheet1")
POSSIBLE_SHEET_NAMES = [s.strip() for s in sheet_names_str.split(",")]

OUTPUT_EXCEL_FILE = os.path.join(DATA_DIR, "missing_articles_results.xlsx")


def load_articles_from_excel_or_google():
    """
    Загружает артикулы из Articles.xlsx или из Google Таблицы
    Возвращает список артикулов для парсинга
    """
    articles = []
    
    # Сначала пробуем загрузить из Excel файла
    if os.path.exists(ARTICLES_EXCEL_FILE):
        try:
            print(f"[ЛОГ] Загрузка артикулов из {ARTICLES_EXCEL_FILE}...")
            wb = load_workbook(ARTICLES_EXCEL_FILE, read_only=True)
            
            # Ищем нужный лист
            sheet_name = None
            for name in POSSIBLE_SHEET_NAMES:
                if name in wb.sheetnames:
                    sheet_name = name
                    break
            
            if not sheet_name:
                print(f"[!] Листы {POSSIBLE_SHEET_NAMES} не найдены в файле!")
                print(f"    Доступные листы: {wb.sheetnames}")
                wb.close()
                return []
            
            ws = wb[sheet_name]
            print(f"    ✓ Найден лист: {sheet_name}")
            
            # Читаем артикулы (колонка B, начиная со 2-й строки)
            start_row = 1
            first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            if first_row[0] and isinstance(first_row[0], str):
                first_cell_lower = str(first_row[0]).lower()
                if any(keyword in first_cell_lower for keyword in ['ссылка', 'link', 'url', 'артикул', 'article']):
                    start_row = 2
            
            for row_num in range(start_row, ws.max_row + 1):
                row = list(ws.iter_rows(min_row=row_num, max_row=row_num, max_col=2, values_only=True))[0]
                article = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                url = str(row[0]).strip() if row[0] else ""
                
                # Если есть артикул - используем его
                if article:
                    articles.append(article)
                # Если артикула нет, но есть ссылка - извлекаем из ссылки
                elif url and "wildberries.ru" in url:
                    import re
                    match = re.search(r'/catalog/(\d+)/', url)
                    if match:
                        articles.append(match.group(1))
            
            wb.close()
            print(f"    ✓ Загружено артикулов из Excel: {len(articles)}")
            return articles
            
        except Exception as e:
            print(f"[!] Ошибка чтения Excel файла: {e}")
            return []
    else:
        print(f"[ЛОГ] Файл {ARTICLES_EXCEL_FILE} не найден")
    
    # Если Excel не найден, пробуем загрузить из Google Таблицы
    if GOOGLE_SHEETS_ENABLED and GOOGLE_SHEET_URL:
        try:
            import gspread
            service_account_file = os.path.join(PROJECT_ROOT, GOOGLE_SERVICE_ACCOUNT_FILE)
            
            if os.path.exists(service_account_file):
                print(f"[ЛОГ] Загрузка артикулов из Google Таблицы...")
                gc = gspread.service_account(filename=service_account_file)
                
                if '/d/' in GOOGLE_SHEET_URL:
                    sheet_id = GOOGLE_SHEET_URL.split('/d/')[1].split('/')[0]
                    spreadsheet = gc.open_by_key(sheet_id)
                    worksheet = spreadsheet.worksheet(GOOGLE_SHEET_NAME)
                    
                    # Читаем все строки (колонка B - артикулы)
                    all_values = worksheet.get_all_values()
                    start_row = 0
                    if all_values and len(all_values) > 0:
                        first_row_lower = str(all_values[0][0]).lower() if all_values[0] else ""
                        if any(keyword in first_row_lower for keyword in ['ссылка', 'link', 'url', 'артикул', 'article']):
                            start_row = 1
                    
                    for row in all_values[start_row:]:
                        if len(row) > 1 and row[1]:
                            article = str(row[1]).strip()
                            if article:
                                articles.append(article)
                    
                    print(f"    ✓ Загружено артикулов из Google Таблицы: {len(articles)}")
                    return articles
        except Exception as e:
            print(f"[!] Ошибка чтения Google Таблицы: {e}")
            return []
    
    # Если ничего не найдено, возвращаем пустой список
    print(f"[!] Не удалось загрузить артикулы ни из Excel, ни из Google Таблицы")
    return []


def check_remote_chrome_available():
    """Проверяет доступность remote Chrome"""
    try:
        import requests
        response = requests.get(f"http://127.0.0.1:{CHROME_DEBUG_PORT}/json", timeout=2)
        return response.status_code == 200
    except:
        return False


def cleanup_profile_locks(profile_path):
    """Очищает lock-файлы профиля Chrome"""
    lock_files = [
        "SingletonLock",
        "lockfile",
        "SingletonSocket",
        "SingletonCookie",
        "Default/DevToolsActivePort"
    ]
    
    cleaned = False
    for lock_file in lock_files:
        lock_path = os.path.join(profile_path, lock_file)
        if os.path.exists(lock_path):
            try:
                os.remove(lock_path)
                cleaned = True
            except:
                pass
    
    return cleaned


def copy_profile_data(source_profile, dest_profile):
    """Копирует данные из исходного профиля в профиль парсера"""
    if not os.path.exists(source_profile):
        print(f"[ЛОГ] Источник не найден: {source_profile}")
        return False
    
    os.makedirs(dest_profile, exist_ok=True)
    
    files_to_copy = [
        "Cookies",
        "Cookies-journal",
        "Network/Cookies",
        "Network/Cookies-journal",
        "Login Data",
        "Login Data-journal",
        "Preferences",
        "Web Data",
        "Web Data-journal",
        "History",
        "History-journal"
    ]
    
    dirs_to_copy = [
        "Local Storage",
        "Session Storage",
        "IndexedDB"
    ]
    
    copied_count = 0
    for item in files_to_copy:
        src_path = os.path.join(source_profile, item)
        dest_path = os.path.join(dest_profile, item)
        
        if os.path.exists(src_path):
            try:
                os.makedirs(os.path.dirname(dest_path), exist_ok=True)
                shutil.copy2(src_path, dest_path)
                copied_count += 1
            except Exception as e:
                print(f"[ЛОГ] - Файл не найден: {item}")
        else:
            print(f"[ЛОГ] - Файл не найден: {item}")
    
    for dir_name in dirs_to_copy:
        src_dir = os.path.join(source_profile, dir_name)
        dest_dir = os.path.join(dest_profile, dir_name)
        
        if os.path.exists(src_dir):
            try:
                shutil.copytree(src_dir, dest_dir, dirs_exist_ok=True)
                copied_count += 1
            except Exception as e:
                print(f"[ЛОГ] - Директория не найдена: {dir_name}")
        else:
            print(f"[ЛОГ] - Директория не найдена: {dir_name}")
    
    return copied_count > 0


def setup_browser_driver():
    """Настраивает браузер Chrome"""
    print(f"\n{'='*60}")
    print(f"[ДИАГНОСТИКА] Настройка браузера CHROME")
    print(f"{'='*60}")
    
    if USE_REMOTE_CHROME:
        print(f"[ЛОГ] Режим: Remote подключение")
        options = ChromeOptions()
        options.add_experimental_option("debuggerAddress", f"127.0.0.1:{CHROME_DEBUG_PORT}")
        try:
            driver = webdriver.Chrome(options=options)
            return driver
        except Exception as e:
            print(f"\n[!] ОШИБКА подключения к Chrome: {e}")
            return None
    else:
        print(f"[ЛОГ] Режим: Прямой запуск браузера")
        
        if USE_TEMP_PROFILE and COPY_PROFILE_DATA:
            # Копируем данные профиля
            chrome_user_data = os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\User Data")
            source_profile = os.path.join(chrome_user_data, SOURCE_PROFILE_FOR_COPY)
            
            print(f"\n{'='*60}")
            print(f"[КОПИРОВАНИЕ] Перенос данных из {SOURCE_PROFILE_FOR_COPY}")
            print(f"{'='*60}")
            print(f"[ЛОГ] Источник: {source_profile}")
            print(f"[ЛОГ] Назначение: {TEMP_PROFILE_DIR}")
            
            copy_profile_data(source_profile, TEMP_PROFILE_DIR)
            cleanup_profile_locks(TEMP_PROFILE_DIR)
        
        # Пробуем несколько конфигураций
        attempts = [
            {'use_subprocess': True, 'version_main': None, 'user_data_dir': None},
            {'use_subprocess': True, 'version_main': None, 'user_data_dir': TEMP_PROFILE_DIR if USE_TEMP_PROFILE else None},
            {'use_subprocess': False, 'version_main': None, 'user_data_dir': TEMP_PROFILE_DIR if USE_TEMP_PROFILE else None},
        ]
        
        for attempt_num, attempt_config in enumerate(attempts, 1):
            try:
                print(f"[ЛОГ] Попытка {attempt_num}/{len(attempts)} запуска Chrome...")
                
                options = ChromeOptions()
                options.add_argument("--disable-dev-shm-usage")
                options.add_argument("--no-sandbox")
                
                driver_kwargs = {
                    'headless': HEADLESS_MODE,
                    'use_subprocess': attempt_config['use_subprocess'],
                    'version_main': attempt_config['version_main'],
                    'options': options
                }
                
                user_dir = attempt_config.get('user_data_dir')
                if user_dir is not None:
                    driver_kwargs['user_data_dir'] = user_dir
                    print(f"[ЛОГ] Использую профиль: {user_dir}")
                else:
                    print(f"[ЛОГ] Запускаю Chrome без профиля (временный профиль)")
                
                driver = uc.Chrome(**driver_kwargs)
                
                print(f"[ЛОГ] Chrome драйвер создан, жду инициализацию Chrome...")
                time.sleep(5)
                
                max_retries = 3
                driver_works = False
                for retry in range(max_retries):
                    try:
                        driver.current_url
                        print(f"[ЛОГ] ✓ Chrome драйвер создан успешно и отвечает")
                        driver_works = True
                        break
                    except Exception as check_error:
                        if retry < max_retries - 1:
                            print(f"[ЛОГ] ⚠ Попытка {retry + 1}/{max_retries}: драйвер еще не готов, жду еще 2 секунды...")
                            time.sleep(2)
                        else:
                            print(f"[ЛОГ] ⚠ Драйвер создан, но не отвечает после {max_retries} попыток")
                            try:
                                driver.quit()
                            except:
                                pass
                            driver = None
                            driver_works = False
                
                if driver_works:
                    return driver
                elif attempt_num < len(attempts):
                    print(f"[ЛОГ] Пробую следующую конфигурацию...")
                    time.sleep(2)
                    continue
                else:
                    raise Exception("Chrome драйвер не отвечает после всех попыток")
                    
            except Exception as e:
                print(f"[ЛОГ] ✗ Ошибка создания Chrome драйвера: {e}")
                if attempt_num < len(attempts):
                    print(f"[ЛОГ] Пробую следующую конфигурацию...")
                    time.sleep(2)
                    continue
                else:
                    raise
        
        return None


def parse_price_from_current_page(driver, article, product_url=None):
    """Парсит цены с текущей открытой страницы товара"""
    try:
        time.sleep(0.5)
        
        page_source_lower = driver.page_source.lower()
        if "Почти готово" in driver.title or "captcha" in page_source_lower:
            print(f"  [{article}] ⚠ Captcha обнаружена!")
            return None
        
        if "подозрительная активность" in page_source_lower:
            print(f"  [{article}] ⚠⚠⚠ WB ЗАБЛОКИРОВАЛ!")
            return None
        
        # Проверяем "Нет в наличии"
        try:
            sold_out_element = driver.find_element(By.CSS_SELECTOR, "h2[class*='soldOutProduct']")
            print(f"  [{article}] ⚠ Товар недоступен: {sold_out_element.text}")
            return {'price': 0, 'price_with_card': 0}
        except:
            pass
        
        unavailable_keywords = ['нет в наличии', 'товар недоступен', 'недоступен для заказа']
        for keyword in unavailable_keywords:
            if keyword in page_source_lower:
                print(f"  [{article}] ⚠ Товар недоступен: '{keyword}'")
                return {'price': 0, 'price_with_card': 0}
        
        # Кликаем на кнопку кошелька
        try:
            wallet_button = WebDriverWait(driver, 2).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "button[class*='priceBlockWalletPrice']"))
            )
            wallet_button.click()
            time.sleep(0.5)
        except:
            pass
        
        # Ищем цены
        price_selectors = [
            (By.CSS_SELECTOR, "h2.mo-typography_color_primary"),
            (By.CSS_SELECTOR, "h2[class*='mo-typography'][class*='color_primary']"),
            (By.CSS_SELECTOR, "ins.priceBlockFinalPrice--iToZR"),
            (By.CSS_SELECTOR, "ins[class*='priceBlockFinalPrice']"),
        ]
        
        price_with_card_selectors = [
            (By.CSS_SELECTOR, "h2.mo-typography_color_danger"),
            (By.CSS_SELECTOR, "h2[class*='mo-typography'][class*='color_danger']"),
        ]
        
        price = None
        price_with_card = None
        
        for by, selector in price_selectors:
            try:
                price_elem = WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((by, selector))
                )
                price_text = price_elem.text.strip()
                price_num = re.sub(r'[^\d]', '', price_text)
                if price_num:
                    price = int(price_num)
                    break
            except:
                continue
        
        for by, selector in price_with_card_selectors:
            try:
                price_card_elem = driver.find_element(by, selector)
                price_card_text = price_card_elem.text.strip()
                price_card_num = re.sub(r'[^\d]', '', price_card_text)
                if price_card_num:
                    price_with_card = int(price_card_num)
                    break
            except:
                continue
        
        # Если цена не найдена, перезагружаем страницу
        if not price:
            print(f"  [{article}] ⚠ Цена не найдена, перезагружаю страницу...")
            try:
                if product_url:
                    driver.get(product_url)
                else:
                    driver.get(driver.current_url)
                time.sleep(5)
                
                for by, selector in price_selectors:
                    try:
                        price_elem = WebDriverWait(driver, 8).until(
                            EC.presence_of_element_located((by, selector))
                        )
                        price_text = price_elem.text.strip()
                        price_num = re.sub(r'[^\d]', '', price_text)
                        if price_num:
                            price = int(price_num)
                            break
                    except:
                        continue
                
                if not price_with_card:
                    for by, selector in price_with_card_selectors:
                        try:
                            price_card_elem = driver.find_element(by, selector)
                            price_card_text = price_card_elem.text.strip()
                            price_card_num = re.sub(r'[^\d]', '', price_card_text)
                            if price_card_num:
                                price_with_card = int(price_card_num)
                                break
                        except:
                            continue
            except Exception as e:
                print(f"  [{article}] ⚠ Ошибка перезагрузки: {e}")
        
        if not price:
            print(f"  [{article}] ✗ Цена не найдена")
            return {'price': 0, 'price_with_card': 0}
        
        return {
            'price': price,
            'price_with_card': price_with_card if price_with_card else 0
        }
    
    except Exception as e:
        print(f"  [{article}] ✗ Ошибка парсинга: {e}")
        return {'price': 0, 'price_with_card': 0}


def process_articles_parallel(driver, articles):
    """Обрабатывает артикулы параллельно"""
    results = []
    last_saved_count = 0  # Счетчик для промежуточного сохранения
    
    # Формируем список товаров с URL
    products = []
    for article in articles:
        products.append({
            'url': f"https://www.wildberries.ru/catalog/{article}/detail.aspx",
            'article': article
        })
    
    total = len(products)
    print(f"\n{'='*80}")
    print(f"ПАРАЛЛЕЛЬНАЯ ОБРАБОТКА: {PARALLEL_TABS} вкладок одновременно")
    print(f"Всего артикулов: {total}")
    print(f"{'='*80}\n")
    
    try:
        main_window = driver.window_handles[0]
    except Exception as e:
        print(f"\n[!] ОШИБКА: Браузер закрыт: {e}")
        return results
    
    # Обрабатываем пачками
    for batch_start in range(0, total, PARALLEL_TABS):
        batch = products[batch_start : batch_start + PARALLEL_TABS]
        batch_num = batch_start // PARALLEL_TABS + 1
        total_batches = (total + PARALLEL_TABS - 1) // PARALLEL_TABS
        
        print(f"\n{'─'*80}")
        print(f"📦 ПАКЕТ {batch_num}/{total_batches} ({len(batch)} артикулов)")
        print(f"{'─'*80}")
        
        # ФАЗА 1: Открываем все вкладки
        print(f"\n[1/4] Открываю {len(batch)} вкладок...")
        driver.switch_to.window(main_window)
        
        initial_handles_count = len(driver.window_handles)
        opened_tabs_map = {}
        
        for idx, product in enumerate(batch):
            try:
                print(f"  [{batch_start + idx + 1}/{total}] Открываю: {product['article']}")
                driver.execute_script("window.open(arguments[0], '_blank');", product['url'])
                time.sleep(0.2)  # Увеличена задержка для открытия вкладки
                
                try:
                    all_handles = driver.window_handles
                    current_count = len(all_handles)
                    print(f"      [ЛОГ] Вкладок после открытия: {current_count}")
                    
                    # Проверяем что вкладка действительно открылась
                    if current_count > initial_handles_count + idx:
                        new_tab_handle = all_handles[-1]
                        opened_tabs_map[new_tab_handle] = product
                        # Переключаемся на новую вкладку чтобы она точно открылась и загрузилась
                        driver.switch_to.window(new_tab_handle)
                        time.sleep(0.2)  # Даем время на загрузку URL
                        # Возвращаемся на главную вкладку
                        driver.switch_to.window(main_window)
                    else:
                        print(f"      [ЛОГ] ⚠ Вкладка не открылась, возможно заблокирована браузером")
                except Exception as tab_error:
                    print(f"      [ЛОГ] ⚠ Ошибка при сохранении соответствия: {tab_error}")
            except Exception as e:
                print(f"  [{batch_start + idx + 1}/{total}] ⚠ Ошибка: {e}")
        
        # ФАЗА 2: Ждем загрузки
        print(f"\n[2/4] Жду полной загрузки страниц...")
        time.sleep(1.5)  # Увеличена задержка для загрузки вкладок
        
        try:
            all_handles = driver.window_handles
            tabs = [h for h in all_handles if h != main_window]
            print(f"  [ЛОГ] Всего окон: {len(all_handles)}, вкладок для парсинга: {len(tabs)}")
            
            # Если вкладки не открылись, пробуем еще раз
            if len(tabs) == 0:
                print(f"  ⚠ ВНИМАНИЕ: Вкладки не открылись! Пробую еще раз...")
                driver.switch_to.window(main_window)
                for idx, product in enumerate(batch):
                    try:
                        driver.execute_script(f"window.open('{product['url']}', '_blank');")
                        time.sleep(0.2)
                        if len(driver.window_handles) > initial_handles_count + idx + 1:
                            driver.switch_to.window(driver.window_handles[-1])
                            time.sleep(0.1)
                            driver.switch_to.window(main_window)
                    except Exception as e:
                        print(f"  [ЛОГ] Ошибка при повторном открытии вкладки {idx+1}: {e}")
                time.sleep(0.5)
                try:
                    all_handles = driver.window_handles
                    tabs = [h for h in all_handles if h != main_window]
                    print(f"  [ЛОГ] После повторной попытки: {len(tabs)} вкладок")
                except:
                    tabs = []
        except Exception as e:
            print(f"  ⚠ Ошибка получения вкладок: {e}")
            tabs = []
        
        if not opened_tabs_map and len(tabs) == len(batch):
            for idx, tab_handle in enumerate(tabs):
                if idx < len(batch):
                    opened_tabs_map[tab_handle] = batch[idx]
        
        print(f"  ✓ Все {len(tabs)} вкладок загружены")
        
        # ФАЗА 3: Парсим цены
        print(f"\n[3/4] Парсинг цен...")
        
        tab_to_product = {}
        if opened_tabs_map:
            tab_to_product = opened_tabs_map.copy()
        else:
            for tab_handle in tabs:
                try:
                    driver.switch_to.window(tab_handle)
                    time.sleep(0.1)
                    current_url = driver.current_url
                    for product in batch:
                        if product['article'] in current_url or product['url'] in current_url:
                            tab_to_product[tab_handle] = product
                            break
                except:
                    continue
        
        for idx, product in enumerate(batch):
            try:
                matching_tab = None
                for tab_handle, tab_product in tab_to_product.items():
                    if tab_product['article'] == product['article']:
                        matching_tab = tab_handle
                        break
                
                if not matching_tab:
                    if idx < len(tabs):
                        matching_tab = tabs[idx]
                        try:
                            driver.switch_to.window(matching_tab)
                            current_url = driver.current_url
                            if product['article'] not in current_url and product['url'] not in current_url:
                                # Ищем правильную вкладку
                                found = False
                                for tab_handle in tabs:
                                    try:
                                        driver.switch_to.window(tab_handle)
                                        tab_url = driver.current_url
                                        if product['article'] in tab_url or product['url'] in tab_url:
                                            matching_tab = tab_handle
                                            found = True
                                            break
                                    except:
                                        continue
                                if not found:
                                    print(f"  [{batch_start + idx + 1}/{total}] ✗ Не найдена вкладка для {product['article']}")
                                    results.append({
                                        'url': product['url'],
                                        'article': product['article'],
                                        'price': 0,
                                        'price_with_card': 0
                                    })
                                    continue
                        except:
                            pass
                    else:
                        print(f"  [{batch_start + idx + 1}/{total}] ⚠ Вкладка не найдена для {product['article']}")
                        results.append({
                            'url': product['url'],
                            'article': product['article'],
                            'price': 0,
                            'price_with_card': 0
                        })
                        continue
                
                driver.switch_to.window(matching_tab)
                
                # Финальная проверка соответствия
                try:
                    current_url = driver.current_url
                    if product['article'] not in current_url and product['url'] not in current_url:
                        print(f"  [{batch_start + idx + 1}/{total}] ⚠ КРИТИЧНО: На вкладке неверный товар!")
                        print(f"      Ожидается: {product['article']}")
                        print(f"      На вкладке: {current_url[:80]}...")
                        found = False
                        for tab_handle in tabs:
                            try:
                                driver.switch_to.window(tab_handle)
                                tab_url = driver.current_url
                                if product['article'] in tab_url or product['url'] in tab_url:
                                    matching_tab = tab_handle
                                    found = True
                                    break
                            except:
                                continue
                        if not found:
                            print(f"  [{batch_start + idx + 1}/{total}] ✗ Не найдена правильная вкладка")
                            results.append({
                                'url': product['url'],
                                'article': product['article'],
                                'price': 0,
                                'price_with_card': 0
                            })
                            continue
                except:
                    pass
                
                price_data = parse_price_from_current_page(driver, product['article'], product['url'])
                
                if price_data is None:
                    price_data = {'price': 0, 'price_with_card': 0}
                
                if isinstance(price_data, (int, float)):
                    price_data = {'price': int(price_data), 'price_with_card': 0}
                
                results.append({
                    'url': product['url'],
                    'article': product['article'],
                    'price': price_data['price'],
                    'price_with_card': price_data.get('price_with_card', 0)
                })
                
                price = price_data['price']
                price_card = price_data.get('price_with_card', 0)
                if price_card and price_card > 0:
                    status = f"{price} ₽ / {price_card} ₽ (с картой)"
                else:
                    status = f"{price} ₽" if price > 0 else "недоступен" if price == 0 else "ошибка"
                print(f"  [{batch_start + idx + 1}/{total}] {product['article']}: {status}")
            
            except Exception as e:
                print(f"  [{batch_start + idx + 1}/{total}] {product['article']}: ✗ ошибка - {e}")
                results.append({
                    'url': product['url'],
                    'article': product['article'],
                    'price': 0,
                    'price_with_card': 0
                })
        
        # ФАЗА 4: Закрываем вкладки
        print(f"\n[4/4] Закрываю вкладки...")
        for tab_handle in tabs:
            try:
                driver.switch_to.window(tab_handle)
                driver.close()
            except:
                pass
        
        try:
            driver.switch_to.window(main_window)
        except:
            main_window = driver.window_handles[0]
            driver.switch_to.window(main_window)
        
        # Сохранение в Google Таблицы после каждого пакета (каждые 20 товаров)
        if GOOGLE_SHEETS_ENABLED and GOOGLE_SHEET_URL and len(results) > last_saved_count:
            new_results = results[last_saved_count:]
            print(f"\n📊 Запись в Google Таблицы ({len(new_results)} новых товаров)...")
            if save_results_to_google_sheets(new_results, GOOGLE_SHEET_URL, GOOGLE_SHEET_NAME, append_only=True):
                print(f"✓ Сохранено в Google Таблицы")
                last_saved_count = len(results)
            else:
                print(f"⚠ Не удалось сохранить в Google Таблицы")
        
        if batch_start + PARALLEL_TABS < total:
            time.sleep(0.5)
    
    return results


def save_results_to_google_sheets(results, sheet_url, sheet_name="Лист1", append_only=True):
    """
    Сохраняет результаты в Google Таблицы через Service Account
    Использует google-credentials.json для авторизации
    """
    if not GOOGLE_SHEETS_ENABLED:
        print(f"  ⚠ Google Sheets отключен (GOOGLE_SHEETS_ENABLED = False)")
        return False
    
    if not sheet_url:
        print(f"  ⚠ Google Sheet URL не указан (GOOGLE_SHEET_URL пустой)")
        return False
    
    try:
        import gspread
    except ImportError:
        print(f"\n[!] Для записи в Google Таблицы установите:")
        print(f"    pip install gspread google-auth google-auth-httplib2")
        return False
    
    try:
        # Извлекаем ID таблицы из URL
        if '/d/' in sheet_url:
            sheet_id = sheet_url.split('/d/')[1].split('/')[0]
        else:
            print(f"[!] Неверный формат ссылки на Google Sheet")
            return False
        
        # Подключаемся к Google Sheets через Service Account
        service_account_file = os.path.join(PROJECT_ROOT, GOOGLE_SERVICE_ACCOUNT_FILE)
        
        if not os.path.exists(service_account_file):
            print(f"  ⚠ Файл {service_account_file} не найден!")
            print(f"     Создайте Service Account в Google Cloud Console и скачайте credentials")
            return False
        
        gc = gspread.service_account(filename=service_account_file)
        spreadsheet = gc.open_by_key(sheet_id)
        
        # Получаем или создаем лист
        try:
            worksheet = spreadsheet.worksheet(sheet_name)
        except gspread.exceptions.WorksheetNotFound:
            worksheet = spreadsheet.add_worksheet(title=sheet_name, rows=1000, cols=10)
        
        # Записываем заголовки если их нет
        if len(worksheet.get_all_values()) == 0:
            worksheet.append_row(["ссылка на товар", "артикул", "цена", "цена с картой"])
        
        # Записываем данные (добавляем в конец если append_only=True)
        batch_size = 100  # Google Sheets API ограничение
        for i in range(0, len(results), batch_size):
            batch = results[i:i+batch_size]
            rows = []
            for result in batch:
                rows.append([
                    result['url'],
                    result['article'],
                    result['price'],
                    result.get('price_with_card', 0)
                ])
            worksheet.append_rows(rows)
        
        return True
        
    except Exception as e:
        print(f"\n[!] ОШИБКА при сохранении в Google Таблицы: {e}")
        import traceback
        traceback.print_exc()
        return False


def save_results_to_excel(results, output_file):
    """Сохраняет результаты в Excel файл"""
    try:
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Результаты"
        
        # Заголовки
        ws.append(["ссылка на товар", "артикул", "цена", "цена с картой"])
        
        # Данные
        for result in results:
            ws.append([
                result['url'],
                result['article'],
                result['price'],
                result.get('price_with_card', 0)
            ])
        
        wb.save(output_file)
        print(f"\n✓ Результаты сохранены в: {output_file}")
        return True
    except Exception as e:
        print(f"\n[!] ОШИБКА при сохранении: {e}")
        return False


def main():
    print("\n" + "="*80)
    print("ПАРСЕР ЦЕН WB - ДОПАРСИНГ НЕДОСТАЮЩИХ АРТИКУЛОВ")
    print("="*80)
    
    # Отладочный вывод настроек
    print(f"\n[ЛОГ] HEADLESS_MODE = {HEADLESS_MODE} (из .env: {os.getenv('HEADLESS_MODE', 'не задано')})")
    
    # Загружаем артикулы из файла или Google Таблицы
    print(f"\n[1/3] Загрузка артикулов...")
    MISSING_ARTICLES = load_articles_from_excel_or_google()
    
    if not MISSING_ARTICLES:
        print("\n[!] Нет артикулов для обработки!")
        print(f"    Проверьте файл {ARTICLES_EXCEL_FILE} или Google Таблицу")
        return
    
    print(f"\n✓ Конфигурация проверена")
    print(f"  Всего артикулов для обработки: {len(MISSING_ARTICLES)}")
    
    # Запускаем Chrome
    print(f"\n[1/3] Запуск Chrome...")
    driver = None
    results = []
    
    try:
        driver = setup_browser_driver()
        
        if not driver:
            print("\n[!] Не удалось запустить Chrome!")
            return
        
        print("    ✓ Chrome запущен")
        
        # Пауза для ручной авторизации
        if WAIT_FOR_MANUAL_LOGIN and not HEADLESS_MODE:
            print(f"\n{'='*80}")
            print("⏸  ПАУЗА ДЛЯ АВТОРИЗАЦИИ")
            print(f"{'='*80}")
            print(f"\n📋 ИНСТРУКЦИЯ:")
            print(f"   1. Открываю сайт WB в браузере...")
            print(f"   2. Авторизуйтесь в своем аккаунте")
            print(f"   3. Установите правильный адрес доставки")
            print(f"   4. После этого вернитесь сюда и нажмите ENTER")
            print(f"\n⏱  Таймаут: {MANUAL_LOGIN_TIMEOUT} секунд")
            print(f"{'='*80}\n")
            
            try:
                print(f"[ЛОГ] Открываю https://www.wildberries.ru/ для авторизации...")
                driver.get("https://www.wildberries.ru/")
                time.sleep(2)
                print(f"[ЛОГ] ✓ Страница WB открыта")
            except Exception as e:
                print(f"[ЛОГ] ⚠ Ошибка открытия WB: {e}")
            
            try:
                input(f"\n⏸ Нажмите ENTER когда авторизуетесь и установите адрес доставки...")
            except KeyboardInterrupt:
                print(f"\n[!] Прервано пользователем")
                try:
                    if driver:
                        driver.quit()
                except:
                    pass
                return
            
            # Открываем тестовую вкладку для подтверждения разрешения на открытие вкладок
            print(f"\n{'='*80}")
            print("⏸  ПОДТВЕРЖДЕНИЕ РАЗРЕШЕНИЯ НА ОТКРЫТИЕ ВКЛАДОК")
            print(f"{'='*80}")
            print(f"\n📋 ИНСТРУКЦИЯ:")
            print(f"   1. Сейчас открою тестовую вкладку с товаром...")
            print(f"   2. В браузере появится запрос: 'Разрешить этому сайту открывать вкладки?'")
            print(f"   3. Нажмите 'РАЗРЕШИТЬ' или 'ALLOW' в браузере")
            print(f"   4. После этого вернитесь сюда и нажмите ENTER")
            print(f"{'='*80}\n")
            
            try:
                # Открываем тестовую вкладку с первым товаром из списка
                test_url = f"https://www.wildberries.ru/catalog/{MISSING_ARTICLES[0]}/detail.aspx"
                print(f"[ЛОГ] Открываю тестовую вкладку: {test_url}")
                driver.execute_script("window.open(arguments[0], '_blank');", test_url)
                time.sleep(2)
                
                # Переключаемся на новую вкладку
                if len(driver.window_handles) > 1:
                    driver.switch_to.window(driver.window_handles[-1])
                    print(f"[ЛОГ] ✓ Тестовая вкладка открыта")
                    time.sleep(1)
                    # Возвращаемся на главную вкладку
                    driver.switch_to.window(driver.window_handles[0])
                else:
                    print(f"[ЛОГ] ⚠ Вкладка не открылась, возможно браузер заблокировал")
            except Exception as e:
                print(f"[ЛОГ] ⚠ Ошибка открытия тестовой вкладки: {e}")
            
            try:
                input(f"\n⏸ Нажмите ENTER после того как разрешите открытие вкладок в браузере...")
            except KeyboardInterrupt:
                print(f"\n[!] Прервано пользователем")
                try:
                    if driver:
                        driver.quit()
                except:
                    pass
                return
            
            # Закрываем тестовую вкладку если она открыта
            try:
                if len(driver.window_handles) > 1:
                    driver.switch_to.window(driver.window_handles[-1])
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    print(f"[ЛОГ] ✓ Тестовая вкладка закрыта")
            except:
                pass
        
        # Парсинг
        print(f"\n[2/3] Парсинг цен для {len(MISSING_ARTICLES)} артикулов...")
        results = process_articles_parallel(driver, MISSING_ARTICLES)
        
        print(f"\n✓ Парсинг завершен: собрано {len(results)} товаров")
        
    except Exception as e:
        print(f"\n[!] КРИТИЧЕСКАЯ ОШИБКА: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        # Сохраняем результаты
        print(f"\n{'='*80}")
        print("СОХРАНЕНИЕ РЕЗУЛЬТАТОВ")
        print(f"{'='*80}")
        
        if len(results) > 0:
            # Сохраняем в Google Таблицы (если еще не сохранено)
            if GOOGLE_SHEETS_ENABLED and GOOGLE_SHEET_URL:
                print(f"\n📊 Финальная запись в Google Таблицы ({len(results)} товаров)...")
                if save_results_to_google_sheets(results, GOOGLE_SHEET_URL, GOOGLE_SHEET_NAME, append_only=True):
                    print(f"✓ Данные загружены в Google Таблицы")
                    print(f"  Ссылка: {GOOGLE_SHEET_URL}")
                else:
                    print(f"⚠ Не удалось сохранить в Google Таблицы")
            
            # Также сохраняем в Excel для резервной копии
            if save_results_to_excel(results, OUTPUT_EXCEL_FILE):
                print(f"✓ Данные сохранены в Excel файл")
            else:
                print(f"⚠ Не удалось сохранить в Excel")
        else:
            print(f"\n⚠ Нет данных для сохранения")
        
        if driver:
            print(f"\n[Закрываю Chrome через 5 секунд...]")
            time.sleep(5)
            try:
                # Пытаемся закрыть все окна перед quit
                try:
                    for handle in driver.window_handles:
                        driver.switch_to.window(handle)
                        driver.close()
                except:
                    pass
                
                # Закрываем драйвер
                driver.quit()
            except Exception as e:
                # Игнорируем ошибки при закрытии (драйвер уже закрыт)
                pass
            finally:
                # Принудительно очищаем ссылку на драйвер
                driver = None
    
    print(f"\n{'='*80}")
    print("ЗАВЕРШЕНО")
    print(f"{'='*80}\n")


if __name__ == "__main__":
    main()


