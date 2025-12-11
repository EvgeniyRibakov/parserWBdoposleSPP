# -*- coding: utf-8 -*-
"""
БЫСТРЫЙ ПАРСЕР ЦЕН WILDBERRIES - ТОЛЬКО API
Получает ВСЕ данные через API за секунды (без браузера!)
"""

import os
import json
import requests
from datetime import datetime, timedelta
from openpyxl import load_workbook
from dotenv import load_dotenv
import time

# === КОНФИГУРАЦИЯ ===
# Пути относительно корня проекта
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(PROJECT_ROOT, "data")

EXCEL_FILE = os.path.join(DATA_DIR, "Парсер цен.xlsx")
SHEET_INPUT_WB = "Данные для парсера ВБ"
SHEET_OUTPUT_WB = "Парсер ВБ"

# API ENDPOINTS
WB_PRICES_API_URL = "https://discounts-prices-api.wildberries.ru/api/v2/list/goods/filter"
WB_CONTENT_API_URL = "https://content-api.wildberries.ru/content/v2/get/cards/list"
WB_STOCKS_API_URL = "https://seller-analytics-api.wildberries.ru/api/v2/stocks-report/products/products"

# Названия кабинетов (для .env файла)
CABINET_NAMES = ["COSMO", "MMA", "MAB", "MAU", "DREAMLAB", "BEAUTYLAB"]

# === ФУНКЦИИ ===

def load_api_keys_from_env():
    """
    Загружает API ключи WB из .env файла
    Формат в .env:
    COSMO=eyJhbGc...
    MMA=eyJhbGc...
    и т.д.
    """
    load_dotenv()  # Загружаем .env файл
    
    api_keys = []
    cabinet_info = []
    
    for cabinet_name in CABINET_NAMES:
        api_key = os.getenv(cabinet_name, "").strip()
        if api_key:
            api_keys.append(api_key)
            cabinet_info.append(cabinet_name)
    
    print(f"\n[API] Загружено из .env файла:")
    print(f"    Кабинетов: {len(api_keys)}")
    for name in cabinet_info:
        print(f"      ✓ {name}")
    
    return api_keys, cabinet_info


def get_product_info(articles, api_keys_list, cabinet_names=None):
    """
    Получает информацию о товарах через Content API
    Возвращает словарь {артикул: {название, nmID, vendorCode, cabinet}}
    """
    print("\n[API] Загрузка информации о товарах (названия, ID)...")
    
    if not api_keys_list:
        print("[!] API ключи не найдены!")
        return {}
    
    product_info = {}
    cabinet_map = {}  # nmID -> cabinet
    
    # Пробуем каждый API ключ
    for idx, api_key in enumerate(api_keys_list, 1):
        cabinet_name = cabinet_names[idx-1] if cabinet_names and idx-1 < len(cabinet_names) else f"Кабинет {idx}"
        print(f"\n[API] {cabinet_name} ({idx}/{len(api_keys_list)})...")
        
        try:
            headers = {
                "Authorization": api_key,
                "Content-Type": "application/json"
            }
            
            # Конвертируем артикулы в set для поиска
            articles_set = {str(art).strip() for art in articles}
            
            # Content API: получаем список карточек с пагинацией (максимум 100 за раз)
            cursor_updatedAt = ""
            cursor_nmID = 0
            total_found_this_cabinet = 0
            page = 0
            
            while True:
                page += 1
                
                payload = {
                    "settings": {
                        "cursor": {
                            "limit": 100
                        },
                        "filter": {
                            "withPhoto": -1
                        }
                    }
                }
                
                # Добавляем курсор для пагинации (если не первая страница)
                if cursor_updatedAt and cursor_nmID:
                    payload["settings"]["cursor"]["updatedAt"] = cursor_updatedAt
                    payload["settings"]["cursor"]["nmID"] = cursor_nmID
                
                response = requests.post(WB_CONTENT_API_URL, headers=headers, json=payload, timeout=30)
                
                if response.status_code == 200:
                    data = response.json()
                    
                    cards = data.get("cards", [])
                    if not cards and "data" in data:
                        cards = data.get("data", {}).get("cards", [])
                    
                    if not cards:
                        # Нет больше карточек
                        break
                    
                    # Обрабатываем карточки
                    for card in cards:
                        nm_id = str(card.get("nmID", ""))
                        vendor_code = str(card.get("vendorCode", ""))
                        
                        # Проверяем совпадение по nmID или vendorCode
                        if nm_id in articles_set or vendor_code in articles_set:
                            # Берем название (может быть в разных полях)
                            title = card.get("title") or card.get("object") or f"Товар {nm_id}"
                            
                            # Используем nmID как ключ
                            if nm_id:
                                product_info[nm_id] = {
                                    "title": title,
                                    "nmID": nm_id,
                                    "vendorCode": vendor_code,
                                    "cabinet": cabinet_name
                                }
                                total_found_this_cabinet += 1
                    
                    # Получаем курсор для следующей страницы
                    cursor_data = data.get("cursor", {})
                    cursor_updatedAt = cursor_data.get("updatedAt", "")
                    cursor_nmID = cursor_data.get("nmID", 0)
                    
                    # Если курсор пустой - больше страниц нет
                    if not cursor_updatedAt or not cursor_nmID:
                        break
                    
                    # Если нашли все нужные товары - можно остановиться
                    if len([x for x in product_info if str(x) in articles_set]) >= len(articles_set):
                        break
                    
                    time.sleep(0.2)  # Пауза между запросами пагинации
                
                else:
                    print(f"[!] Ошибка Content API: {response.status_code}")
                    print(f"    {response.text[:200]}")
                    break
            
            print(f"    Обработано страниц: {page}, найдено товаров: {total_found_this_cabinet}")
            time.sleep(0.3)
        
        except Exception as e:
            print(f"[!] Ошибка при запросе Content API (кабинет {idx}): {e}")
    
    print(f"\n[API] Итого загружено информации о {len(product_info)} товарах")
    return product_info


def get_prices_full_info(articles, api_keys_list, cabinet_names=None):
    """
    Получает ВСЕ цены через Prices API - ДО и ПОСЛЕ СПП!
    Возвращает словарь {артикул: {price_original, price_before_spp, price_after_spp, discount, spp, stocks}}
    
    Структура цен WB API:
    - price: базовая цена (без скидок)
    - discountedPrice: цена после обычных скидок (ДО СПП)
    - clubDiscountedPrice: финальная цена (ПОСЛЕ СПП и скидок кошелька)
    - stocks: остатки товара
    """
    print("\n[API] Загрузка цен ДО и ПОСЛЕ СПП через API...")
    
    if not api_keys_list:
        print("[!] API ключи не найдены!")
        return {}
    
    prices_info = {}
    debug_done = False  # Флаг для вывода DEBUG только 1 раз
    
    for idx, api_key in enumerate(api_keys_list, 1):
        cabinet_name = cabinet_names[idx-1] if cabinet_names and idx-1 < len(cabinet_names) else f"Кабинет {idx}"
        print(f"\n[API] {cabinet_name} ({idx}/{len(api_keys_list)})...")
        
        try:
            headers = {
                "Authorization": api_key,
                "Content-Type": "application/json"
            }
            
            # Обрабатываем батчами по 1000
            batch_size = 1000
            
            for i in range(0, len(articles), batch_size):
                batch = articles[i:i + batch_size]
                nm_ids = [int(art) for art in batch if str(art).isdigit()]
                
                if not nm_ids:
                    continue
                
                # Правильный формат для Prices API
                payload = {
                    "limit": 1000,
                    "offset": 0,
                    "nmList": nm_ids  # ВАЖНО: nmList а не filterNmID!
                }
                
                response = requests.post(WB_PRICES_API_URL, headers=headers, json=payload, timeout=30)
                
                if response.status_code == 200:
                    data = response.json()
                    
                    # Парсим товары
                    goods_list = []
                    if "data" in data and "listGoods" in data["data"]:
                        goods_list = data["data"]["listGoods"]
                    elif "listGoods" in data:
                        goods_list = data["listGoods"]
                    
                    
                    # Обрабатываем товары
                    for item in goods_list:
                        nm_id = str(item.get("nmID", ""))
                        
                        # Берем данные из первого размера
                        sizes = item.get("sizes", [])
                        if sizes and len(sizes) > 0:
                            size_data = sizes[0]
                            
                            # Все данные из Prices API
                            price_original = size_data.get("price", 0)  # price
                            price_discounted = size_data.get("discountedPrice", 0)  # discountedPrice
                            price_club = size_data.get("clubDiscountedPrice", 0)  # clubDiscountedPrice
                            tech_size_name = size_data.get("techSizeName", "")  # techSizeName
                            
                            # Проценты скидок
                            discount_percent = item.get("discount", 0)  # discount
                            club_discount_percent = item.get("clubDiscount", 0)  # clubDiscount
                            
                            # Если нет цены после скидок, используем базовую
                            if not price_discounted and price_original:
                                price_discounted = price_original
                            
                            # Если нет клубной цены, используем цену после скидок
                            if not price_club and price_discounted:
                                price_club = price_discounted
                            
                            if nm_id:
                                prices_info[nm_id] = {
                                    "price": float(price_original) if price_original else 0,
                                    "discountedPrice": float(price_discounted) if price_discounted else 0,
                                    "clubDiscountedPrice": float(price_club) if price_club else 0,
                                    "techSizeName": tech_size_name,
                                    "discount": float(discount_percent) if discount_percent else 0,
                                    "clubDiscount": float(club_discount_percent) if club_discount_percent else 0
                                }
                    
                    print(f"    Батч {i//batch_size + 1}: загружено цен для {len(goods_list)} товаров")
                
                else:
                    print(f"[!] Ошибка Prices API: {response.status_code}")
                    print(f"    {response.text[:200]}")
                
                time.sleep(0.3)
        
        except Exception as e:
            print(f"[!] Ошибка при запросе Prices API (кабинет {idx}): {e}")
            import traceback
            traceback.print_exc()
    
    print(f"\n[API] Итого загружено цен для {len(prices_info)} товаров")
    return prices_info


def get_stocks_info(api_keys_list, cabinet_names=None, articles=None):
    """
    Получает остатки товаров через /api/v2/stocks-report/products/products
    Возвращает словарь {nmID: {stockCount, minPrice, maxPrice}}
    """
    print("\n[API] Загрузка остатков через Stocks API...")
    
    if not api_keys_list:
        print("[!] API ключи не найдены!")
        return {}
    
    stocks_info = {}
    
    # Формируем список nmIDs для фильтрации
    nm_ids = [int(art) for art in articles if str(art).isdigit()] if articles else []
    
    for idx, api_key in enumerate(api_keys_list, 1):
        cabinet_name = cabinet_names[idx-1] if cabinet_names and idx-1 < len(cabinet_names) else f"Кабинет {idx}"
        print(f"\n[API] {cabinet_name} ({idx}/{len(api_keys_list)})...")
        
        try:
            headers = {
                "Authorization": api_key,
                "Content-Type": "application/json"
            }
            
            # Минимальный payload - только nmIDs для фильтрации
            payload = {}
            if nm_ids:
                payload["nmIDs"] = nm_ids[:1000]  # Ограничение 1000
            
            response = requests.post(WB_STOCKS_API_URL, headers=headers, json=payload, timeout=60)
            
            if response.status_code == 200:
                data = response.json()
                
                # Парсим товары
                products = []
                if isinstance(data, list):
                    products = data
                elif isinstance(data, dict):
                    products = data.get("products", []) or data.get("data", [])
                
                for product in products:
                    nm_id = str(product.get("nmID", "") or product.get("nmId", ""))
                    
                    if nm_id:
                        stocks_info[nm_id] = {
                            "stockCount": product.get("stockCount", 0) or 0,
                            "minPrice": product.get("minPrice", 0) or 0,
                            "maxPrice": product.get("maxPrice", 0) or 0
                        }
                
                print(f"    Загружено остатков для {len(products)} товаров")
            
            elif response.status_code == 401:
                print(f"    [!] Ошибка 401: Неверный API ключ")
            elif response.status_code == 400:
                error_text = response.text[:500]
                print(f"[!] Ошибка 400: {error_text}")
                # Пробуем без фильтра
                if nm_ids:
                    print(f"    Пробуем запрос без фильтра nmIDs...")
                    response2 = requests.post(WB_STOCKS_API_URL, headers=headers, json={}, timeout=60)
                    if response2.status_code == 200:
                        data = response2.json()
                        products = data if isinstance(data, list) else data.get("products", [])
                        for product in products:
                            nm_id = str(product.get("nmID", "") or product.get("nmId", ""))
                            if nm_id:
                                stocks_info[nm_id] = {
                                    "stockCount": product.get("stockCount", 0) or 0,
                                    "minPrice": product.get("minPrice", 0) or 0,
                                    "maxPrice": product.get("maxPrice", 0) or 0
                                }
                        print(f"    Загружено остатков для {len(products)} товаров")
            else:
                print(f"[!] Ошибка Stocks API: {response.status_code}")
                print(f"    {response.text[:300]}")
            
            time.sleep(0.3)
        
        except Exception as e:
            print(f"[!] Ошибка при запросе Stocks API ({cabinet_name}): {e}")
    
    print(f"\n[API] Итого загружено остатков для {len(stocks_info)} товаров")
    return stocks_info




def parse_wb_fast_api(wb, api_keys, cabinet_names=None):
    """
    БЫСТРЫЙ парсинг WB - ТОЛЬКО через API!
    Получает: название, nmID, цену до СПП, цену после СПП
    """
    print("\n" + "="*80)
    print("БЫСТРЫЙ ПАРСИНГ WB - ТОЛЬКО API (БЕЗ БРАУЗЕРА!)")
    print("="*80)
    
    # Загрузка артикулов
    ws_in = wb[SHEET_INPUT_WB]
    ws_out = wb[SHEET_OUTPUT_WB]
    
    articles = []
    for row in ws_in.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            article = str(row[0]).strip()
            articles.append(article)
    
    total = len(articles)
    print(f"\n[1/6] Найдено артикулов: {total}")
    
    if total == 0:
        print("[!] Нет артикулов для обработки!")
        return
    
    start_time = time.time()
    
    # Шаг 1: Получаем информацию о товарах (названия)
    print("\n[2/6] Получение информации о товарах через Content API...")
    product_info_dict = get_product_info(articles, api_keys, cabinet_names)
    
    
    # Шаг 2: Получаем цены (до и после СПП)
    print("\n[3/6] Получение цен через Prices API...")
    prices_dict = get_prices_full_info(articles, api_keys, cabinet_names)
    
    
    # Шаг 3: Получаем остатки через Stocks API
    print("\n[4/6] Получение остатков через Stocks API...")
    stocks_dict = get_stocks_info(api_keys, cabinet_names, articles)
    
    
    # Шаг 4: Очищаем старые данные и обновляем заголовки
    print(f"\n[5/6] Очистка старых записей...")
    
    # Удаляем ВСЕ строки (включая заголовки) для обновления структуры
    if ws_out.max_row >= 1:
        ws_out.delete_rows(1, ws_out.max_row)
        print(f"    ✓ Удалены старые данные и заголовки")
    
    # Создаем новые заголовки с названиями полей из API
    ws_out.append([
        "Дата", 
        "Кабинет (cabinet)", 
        "nmID", 
        "Название (title)",
        "Размер (techSizeName)",
        "price", 
        "discountedPrice", 
        "clubDiscountedPrice",
        "discount %",
        "clubDiscount %",
        "stockCount",
        "minPrice",
        "maxPrice"
    ])
    
    # Включаем автофильтр на заголовки
    ws_out.auto_filter.ref = ws_out.dimensions
    
    # Шаг 5: Объединяем данные и сохраняем
    print(f"\n[6/6] Сохранение результатов...")
    print("="*80)
    
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    success = 0
    failed = 0
    
    for i, article in enumerate(articles, 1):
        # Получаем данные
        info = product_info_dict.get(article, {})
        prices = prices_dict.get(article, {})
        stocks_data = stocks_dict.get(article, {})
        
        title = info.get("title", "Не найдено")
        nm_id = info.get("nmID", article)
        cabinet = info.get("cabinet", "Неизвестно")
        
        # Все данные из Prices API
        price_base = prices.get("price", 0)  # price
        price_discounted = prices.get("discountedPrice", 0)  # discountedPrice
        price_club = prices.get("clubDiscountedPrice", 0)  # clubDiscountedPrice
        tech_size_name = prices.get("techSizeName", "")  # techSizeName
        discount_percent = prices.get("discount", 0)  # discount
        club_discount_percent = prices.get("clubDiscount", 0)  # clubDiscount
        
        # Остатки и цены из Stocks API
        stock_count = stocks_data.get("stockCount", 0)
        min_price = stocks_data.get("minPrice", 0)
        max_price = stocks_data.get("maxPrice", 0)
        
        # Прогресс каждые 50 товаров
        if i % 50 == 0:
            print(f"[{i}/{total}] Обработано товаров...")
        
        if price_base or price_discounted or price_club:
            # Сохраняем все данные
            new_row = [
                timestamp,
                cabinet,
                nm_id,
                title,
                tech_size_name if tech_size_name else "",
                price_base if price_base else None,
                price_discounted if price_discounted else None,
                price_club if price_club else None,
                discount_percent if discount_percent else None,
                club_discount_percent if club_discount_percent else None,
                stock_count if stock_count else 0,
                min_price if min_price else None,
                max_price if max_price else None
            ]
            ws_out.append(new_row)
            success += 1
        else:
            failed += 1
            new_row = [
                timestamp,
                cabinet,
                nm_id,
                title,
                "",
                None,
                None,
                None,
                None,
                None,
                0,
                None,
                None
            ]
            ws_out.append(new_row)
    
    # Итоги
    elapsed = time.time() - start_time
    print(f"\n{'='*80}")
    print("ГОТОВО!")
    print(f"{'='*80}")
    print(f"Всего артикулов: {total}")
    print(f"Успешно обработано: {success}")
    print(f"Не найдено: {failed}")
    print(f"Время выполнения: {elapsed:.1f} сек ({elapsed/60:.2f} мин)")
    print(f"Скорость: {total/elapsed:.1f} артикулов/сек")
    print(f"{'='*80}")
    
    wb.save(EXCEL_FILE)
    print(f"\n[SAVE] ✓ Результаты сохранены в '{EXCEL_FILE}'")


def main():
    print("\n" + "!"*80)
    print("БЫСТРЫЙ ПАРСЕР WB - ТОЛЬКО API")
    print("!"*80)
    print("\nОсобенности:")
    print("  ✓ Работает БЕЗ браузера - только API запросы")
    print("  ✓ В 50-100 раз быстрее обычного парсинга")
    print("  ✓ Получает: название, ID, цену ДО СПП, цену ПОСЛЕ СПП")
    print("  ✓ Обрабатывает все 6 магазинов одновременно")
    print("\nТребования:")
    print("  1. ЗАКРОЙТЕ Excel файл перед запуском")
    print("  2. API ключи WB в файле .env (6 кабинетов)")
    print("  3. Артикулы в листе 'Данные для парсера ВБ' (столбец A)")
    print("!"*80)
    
    input("\n💡 Нажмите Enter чтобы начать...")
    
    # Загружаем API ключи из .env
    api_keys, cabinet_names = load_api_keys_from_env()
    
    if not api_keys:
        print("\n[!] ОШИБКА: Не найдено ни одного API ключа в .env файле!")
        print("\n📝 Создайте файл .env в той же папке со скриптом:")
        print("    COSMO=ваш_api_ключ_1")
        print("    MMA=ваш_api_ключ_2")
        print("    MAB=ваш_api_ключ_3")
        print("    MAU=ваш_api_ключ_4")
        print("    DREAMLAB=ваш_api_ключ_5")
        print("    BEAUTYLAB=ваш_api_ключ_6")
        return
    
    # Загружаем Excel
    try:
        wb = load_workbook(EXCEL_FILE)
    except Exception as e:
        print(f"\n[!] Ошибка открытия файла '{EXCEL_FILE}': {e}")
        print("    Убедитесь что файл существует и закрыт!")
        return
    
    try:
        # Быстрый парсинг через API
        parse_wb_fast_api(wb, api_keys, cabinet_names)
        
        print("\n" + "="*80)
        print("✓ ВСЕ ЗАДАЧИ ВЫПОЛНЕНЫ УСПЕШНО!")
        print("="*80)
        
    except Exception as e:
        print(f"\n[!] ОШИБКА: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        wb.close()
        print("\n[DONE] Завершено!")


if __name__ == "__main__":
    main()

