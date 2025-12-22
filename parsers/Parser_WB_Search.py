# -*- coding: utf-8 -*-
"""
ПАРСЕР ЦЕН WILDBERRIES - ПРОСТОЙ ПАРСЕР ЦЕН
Открывает карточки товаров напрямую по ссылкам и извлекает цену
Сохраняет результаты в Google Таблицы

ИНСТРУКЦИЯ:
1. Убедитесь что файл Articles.xlsx содержит ссылки (колонка A) и артикулы (колонка B)
2. Убедитесь что Chrome закрыт (или используйте remote режим)
3. Запустите: python Parser_WB_Search.py
4. Парсер читает ссылки из файла Articles.xlsx
5. Результаты сохраняются в Google Таблицы (настроено в конфигурации)

РЕЖИМЫ РАБОТЫ:
- Обычный режим (USE_REMOTE_CHROME = False): запускает браузер с вашим профилем
- Remote режим (USE_REMOTE_CHROME = True): подключается к уже запущенному браузеру
  Для remote режима сначала запустите START_CHROME_DEBUG.bat

ВЫБОР БРАУЗЕРА:
- Chrome (BROWSER_TYPE = 'chrome') - по умолчанию
- Edge (BROWSER_TYPE = 'edge') - может работать стабильнее с профилями
"""

import os
import sys
import time
import random
import re
import subprocess
import shutil
import threading
from selenium import webdriver

# Загрузка переменных окружения из .env файла
try:
    from dotenv import load_dotenv
    # Загружаем .env файл из корня проекта
    PROJECT_ROOT_TEMP = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    env_path = os.path.join(PROJECT_ROOT_TEMP, '.env')
    if os.path.exists(env_path):
        load_dotenv(env_path)
        print(f"[ЛОГ] Загружены настройки из .env файла")
except ImportError:
    print("[ЛОГ] python-dotenv не установлен, используются настройки по умолчанию")
except Exception as e:
    print(f"[ЛОГ] Ошибка загрузки .env: {e}, используются настройки по умолчанию")

# Настройка кодировки консоли для Windows
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from openpyxl import load_workbook, Workbook
from selenium.common.exceptions import InvalidSessionIdException
import requests
import undetected_chromedriver as uc

# Конфигурация
# Пути относительно корня проекта
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(PROJECT_ROOT, "data")

# Функция для чтения настроек из .env с fallback на значения по умолчанию
def get_env_bool(key, default=False):
    """Читает булево значение из .env"""
    value = os.getenv(key, str(default)).strip().lower()
    return value in ('true', '1', 'yes', 'on')

def get_env_int(key, default=0):
    """Читает целое число из .env"""
    try:
        return int(os.getenv(key, str(default)))
    except:
        return default

def get_env_float(key, default=0.0):
    """Читает число с плавающей точкой из .env"""
    try:
        return float(os.getenv(key, str(default)))
    except:
        return default

def get_env_str(key, default=""):
    """Читает строку из .env"""
    return os.getenv(key, default)

def get_env_tuple(key_min, key_max, default_tuple):
    """Читает кортеж из двух значений .env"""
    min_val = get_env_float(key_min, default_tuple[0])
    max_val = get_env_float(key_max, default_tuple[1])
    return (min_val, max_val)

# Файл с артикулами и ссылками
ARTICLES_EXCEL_FILE = os.path.join(PROJECT_ROOT, get_env_str("ARTICLES_EXCEL_FILE", "Articles.xlsx"))
# Возможные имена листов
sheet_names_str = get_env_str("POSSIBLE_SHEET_NAMES", "Данные для парсера ВБ,WBarticules,WB,Артикулы,Sheet1")
POSSIBLE_SHEET_NAMES = [s.strip() for s in sheet_names_str.split(",")]
OUTPUT_EXCEL_FILE = os.path.join(PROJECT_ROOT, get_env_str("OUTPUT_EXCEL_FILE", "data/prices_results.xlsx"))

# Пути к Chrome
CHROME_USER_DATA_DIR = os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\User Data")
CHROME_PROFILE_NAME = get_env_str("CHROME_PROFILE_NAME", "Default")

# Пути к Edge
EDGE_USER_DATA_DIR = os.path.expandvars(r"%LOCALAPPDATA%\Microsoft\Edge\User Data")
EDGE_PROFILE_NAME = get_env_str("EDGE_PROFILE_NAME", "Default")

# Использовать remote Chrome/Edge
USE_REMOTE_CHROME = get_env_bool("USE_REMOTE_CHROME", False)
CHROME_DEBUG_PORT = get_env_int("CHROME_DEBUG_PORT", 9222)

# Использовать временный профиль для парсинга
USE_TEMP_PROFILE = get_env_bool("USE_TEMP_PROFILE", True)
TEMP_PROFILE_DIR = os.path.join(PROJECT_ROOT, "chrome_parser_profile")

# Копировать данные из профиля Chrome в рабочий профиль
COPY_PROFILE_DATA = get_env_bool("COPY_PROFILE_DATA", True)
SOURCE_PROFILE_FOR_COPY = get_env_str("SOURCE_PROFILE_FOR_COPY", "Profile 4")

# Выбор браузера
BROWSER_TYPE = get_env_str("BROWSER_TYPE", "chrome").lower()

# Режим работы браузера
HEADLESS_MODE = get_env_bool("HEADLESS_MODE", True)

# Пауза для ручной авторизации при первом запуске
WAIT_FOR_MANUAL_LOGIN = get_env_bool("WAIT_FOR_MANUAL_LOGIN", True)
MANUAL_LOGIN_TIMEOUT = get_env_int("MANUAL_LOGIN_TIMEOUT", 120)

# Промежуточное сохранение результатов
SAVE_INTERMEDIATE_RESULTS = get_env_bool("SAVE_INTERMEDIATE_RESULTS", True)
SAVE_EVERY_N_PRODUCTS = get_env_int("SAVE_EVERY_N_PRODUCTS", 10)

# Параллельная обработка товаров
PARALLEL_TABS = get_env_int("PARALLEL_TABS", 20)
DELAY_BETWEEN_TABS = get_env_tuple("DELAY_BETWEEN_TABS_MIN", "DELAY_BETWEEN_TABS_MAX", (0.0, 0.1))  # Минимальные задержки
DELAY_BETWEEN_BATCHES = get_env_tuple("DELAY_BETWEEN_BATCHES_MIN", "DELAY_BETWEEN_BATCHES_MAX", (0.5, 1.0))  # Сокращены до минимума
TEST_MODE = get_env_bool("TEST_MODE", False)
TEST_PRODUCTS_COUNT = get_env_int("TEST_PRODUCTS_COUNT", 50)

# Google Таблицы
GOOGLE_SHEETS_ENABLED = get_env_bool("GOOGLE_SHEETS_ENABLED", True)
GOOGLE_SHEET_URL = get_env_str("GOOGLE_SHEET_URL", "https://docs.google.com/spreadsheets/d/1fbMPHE43ikYM90gcSVk_kcUItjzo-OsYI3T25yOJgQU/edit")
GOOGLE_SHEET_NAME = get_env_str("GOOGLE_SHEET_NAME", "Лист1")
GOOGLE_SERVICE_ACCOUNT_FILE = get_env_str("GOOGLE_SERVICE_ACCOUNT_FILE", "google-credentials.json")
GOOGLE_CREDENTIALS_FILE = get_env_str("GOOGLE_CREDENTIALS_FILE", "google_credentials.json")


def check_chrome_running():
    """Проверяет, запущен ли Chrome"""
    try:
        print(f"[ЛОГ] Проверка запущенных процессов Chrome...")
        result = subprocess.run(['tasklist', '/FI', 'IMAGENAME eq chrome.exe'], 
                              capture_output=True, text=True, timeout=5)
        is_running = 'chrome.exe' in result.stdout
        if is_running:
            print(f"[ЛОГ] Chrome процессы найдены:")
            # Подсчитываем количество процессов
            lines = [line for line in result.stdout.split('\n') if 'chrome.exe' in line]
            print(f"[ЛОГ]   Найдено процессов: {len(lines)}")
            for line in lines[:5]:  # Показываем первые 5
                print(f"[ЛОГ]   {line.strip()}")
        else:
            print(f"[ЛОГ] Chrome процессы не найдены")
        return is_running
    except Exception as e:
        print(f"[ЛОГ] Ошибка проверки Chrome процессов: {e}")
        return False


def check_edge_running():
    """Проверяет, запущен ли Edge"""
    try:
        print(f"[ЛОГ] Проверка запущенных процессов Edge...")
        result = subprocess.run(['tasklist', '/FI', 'IMAGENAME eq msedge.exe'], 
                              capture_output=True, text=True, timeout=5)
        is_running = 'msedge.exe' in result.stdout
        if is_running:
            print(f"[ЛОГ] Edge процессы найдены:")
            # Подсчитываем количество процессов
            lines = [line for line in result.stdout.split('\n') if 'msedge.exe' in line]
            print(f"[ЛОГ]   Найдено процессов: {len(lines)}")
            for line in lines[:5]:  # Показываем первые 5
                print(f"[ЛОГ]   {line.strip()}")
        else:
            print(f"[ЛОГ] Edge процессы не найдены")
        return is_running
    except Exception as e:
        print(f"[ЛОГ] Ошибка проверки Edge процессов: {e}")
        return False


def check_remote_chrome_available():
    """Проверяет, доступен ли Chrome в remote режиме"""
    try:
        import requests
        url = f"http://127.0.0.1:{CHROME_DEBUG_PORT}/json"
        print(f"[ЛОГ] Проверка remote Chrome: {url}")
        response = requests.get(url, timeout=2)
        print(f"[ЛОГ] Ответ: статус {response.status_code}")
        if response.status_code == 200:
            print(f"[ЛОГ] Remote Chrome доступен")
        return response.status_code == 200
    except Exception as e:
        print(f"[ЛОГ] Remote Chrome недоступен: {e}")
        return False


def copy_profile_data(source_profile, target_profile, copy_cookies=True, copy_storage=True):
    """
    Копирует данные из одного профиля Chrome в другой
    source_profile: путь к исходному профилю (Profile 4)
    target_profile: путь к целевому профилю
    """
    print(f"\n{'='*60}")
    print(f"[КОПИРОВАНИЕ] Перенос данных из Profile 4")
    print(f"{'='*60}")
    print(f"[ЛОГ] Источник: {source_profile}")
    print(f"[ЛОГ] Назначение: {target_profile}")
    
    if not os.path.exists(source_profile):
        print(f"[!] ОШИБКА: Исходный профиль не найден!")
        return False
    
    if not os.path.exists(target_profile):
        print(f"[ЛОГ] Создаю целевую директорию...")
        os.makedirs(target_profile, exist_ok=True)
    
    files_to_copy = []
    
    if copy_cookies:
        # Файлы с cookies и сессиями
        files_to_copy.extend([
            "Cookies",
            "Cookies-journal",
            "Network\\Cookies",
            "Network\\Cookies-journal",
            "Login Data",  # Сохраненные пароли и логины
            "Login Data-journal",
        ])
    
    if copy_storage:
        # Local Storage и другие данные
        files_to_copy.extend([
            "Local Storage",
            "Session Storage",
            "IndexedDB",
            "Preferences",  # Настройки профиля (ВАЖНО для адреса!)
            "Web Data",  # Автозаполнение форм (адреса, данные)
            "Web Data-journal",
            "History",  # История
            "History-journal",
        ])
    
    copied_count = 0
    for file_name in files_to_copy:
        source_file = os.path.join(source_profile, file_name)
        target_file = os.path.join(target_profile, file_name)
        
        if os.path.exists(source_file):
            try:
                # Создаём родительскую директорию если нужно
                target_dir = os.path.dirname(target_file)
                if target_dir and not os.path.exists(target_dir):
                    os.makedirs(target_dir, exist_ok=True)
                
                # Копируем файл или директорию
                if os.path.isdir(source_file):
                    if os.path.exists(target_file):
                        shutil.rmtree(target_file)
                    shutil.copytree(source_file, target_file)
                    print(f"[ЛОГ] ✓ Скопирована директория: {file_name}")
                else:
                    shutil.copy2(source_file, target_file)
                    file_size = os.path.getsize(source_file)
                    print(f"[ЛОГ] ✓ Скопирован файл: {file_name} ({file_size} байт)")
                
                copied_count += 1
            except Exception as e:
                print(f"[ЛОГ] ✗ Ошибка копирования {file_name}: {e}")
        else:
            print(f"[ЛОГ] - Файл не найден: {file_name}")
    
    print(f"\n[ЛОГ] Итого скопировано: {copied_count} элементов")
    print(f"{'='*60}\n")
    
    return copied_count > 0


def cleanup_profile_locks(profile_path):
    """Очищает lock-файлы профиля Chrome"""
    lock_files = [
        "SingletonLock",
        "lockfile",
        "SingletonSocket",
        "SingletonCookie"
    ]
    
    cleaned = False
    print(f"[ЛОГ] Очистка lock-файлов в: {profile_path}")
    
    for lock_file in lock_files:
        lock_path = os.path.join(profile_path, lock_file)
        if os.path.exists(lock_path):
            try:
                file_size = os.path.getsize(lock_path)
                print(f"[ЛОГ]   Удаляю: {lock_file} (размер: {file_size} байт)")
                os.remove(lock_path)
                cleaned = True
                print(f"[ЛОГ]   ✓ Удалено успешно")
            except Exception as e:
                print(f"[ЛОГ]   ✗ Ошибка удаления {lock_file}: {e}")
        else:
            print(f"[ЛОГ]   {lock_file} не найден")
    
    # Также очищаем DevToolsActivePort если есть
    devtools_port = os.path.join(profile_path, "DevToolsActivePort")
    if os.path.exists(devtools_port):
        try:
            file_size = os.path.getsize(devtools_port)
            print(f"[ЛОГ]   Удаляю: DevToolsActivePort (размер: {file_size} байт)")
            os.remove(devtools_port)
            cleaned = True
            print(f"[ЛОГ]   ✓ Удалено успешно")
        except Exception as e:
            print(f"[ЛОГ]   ✗ Ошибка удаления DevToolsActivePort: {e}")
    else:
        print(f"[ЛОГ]   DevToolsActivePort не найден")
    
    print(f"[ЛОГ] Результат очистки: {'очищено' if cleaned else 'нечего очищать'}")
    return cleaned


def setup_browser_driver():
    """
    Настраивает браузер (Chrome или Edge)
    Автоматически определяет режим работы
    """
    print(f"\n{'='*60}")
    print(f"[ДИАГНОСТИКА] Настройка браузера {BROWSER_TYPE.upper()}")
    print(f"{'='*60}")
    
    # Автоматическое определение режима
    auto_remote = False
    if not USE_REMOTE_CHROME:
        print(f"[ЛОГ] USE_REMOTE_CHROME = {USE_REMOTE_CHROME}")
        # Проверяем, доступен ли remote Chrome
        print(f"[ЛОГ] Проверка доступности remote Chrome на порту {CHROME_DEBUG_PORT}...")
        if check_remote_chrome_available():
            print(f"    [Авто] Обнаружен Chrome в remote режиме, переключаюсь...")
            auto_remote = True
        else:
            print(f"[ЛОГ] Remote Chrome недоступен")
    
    if USE_REMOTE_CHROME or auto_remote:
        # Подключение к уже запущенному браузеру
        print(f"[ЛОГ] Режим: Remote подключение")
        if BROWSER_TYPE == 'edge':
            options = EdgeOptions()
        else:
            options = ChromeOptions()
        
        options.add_experimental_option("debuggerAddress", f"127.0.0.1:{CHROME_DEBUG_PORT}")
        print(f"    [Режим] Подключение к {BROWSER_TYPE.upper()} (port {CHROME_DEBUG_PORT})")
        
        try:
            if BROWSER_TYPE == 'edge':
                driver = webdriver.Edge(options=options)
            else:
                driver = webdriver.Chrome(options=options)
            return driver
        except Exception as e:
            print(f"\n[!] ОШИБКА подключения к {BROWSER_TYPE.upper()}: {e}")
            print(f"\n💡 Убедись что браузер запущен через START_CHROME_DEBUG.bat")
            return None
    else:
        # Используем профиль пользователя
        print(f"[ЛОГ] Режим: Прямой запуск браузера")
        
        if BROWSER_TYPE == 'edge':
            # Edge использует другой путь к профилям
            profile_path = os.path.join(EDGE_USER_DATA_DIR, EDGE_PROFILE_NAME)
            options = EdgeOptions()
            
            print(f"[ЛОГ] Edge User Data Dir: {EDGE_USER_DATA_DIR}")
            print(f"[ЛОГ] Edge Profile Name: {EDGE_PROFILE_NAME}")
            print(f"[ЛОГ] Полный путь к профилю: {profile_path}")
            print(f"[ЛОГ] User Data Dir существует: {os.path.exists(EDGE_USER_DATA_DIR)}")
            print(f"[ЛОГ] Профиль существует: {os.path.exists(profile_path)}")
            
            # Проверяем, запущен ли Edge
            edge_running = check_edge_running()
            print(f"[ЛОГ] Edge запущен: {edge_running}")
            
            if edge_running:
                print(f"    ⚠ Edge уже запущен!")
                print(f"    [Авто] Пытаюсь очистить lock-файлы профиля...")
                
                # Автоматически очищаем lock-файлы
                cleaned = cleanup_profile_locks(profile_path)
                if cleaned:
                    print(f"    ✓ Lock-файлы очищены, пробую запустить...")
                    time.sleep(1)
                else:
                    print(f"    ⚠ Lock-файлы не найдены")
            else:
                # Очищаем старые lock-файлы на всякий случай
                print(f"[ЛОГ] Очистка lock-файлов профиля...")
                cleanup_profile_locks(profile_path)
            
            options.add_argument(f"--user-data-dir={EDGE_USER_DATA_DIR}")
            options.add_argument(f"--profile-directory={EDGE_PROFILE_NAME}")
            print(f"    [Режим] Запуск Edge с профилем '{EDGE_PROFILE_NAME}'")
        else:
            # Chrome
            profile_path = os.path.join(CHROME_USER_DATA_DIR, CHROME_PROFILE_NAME)
            options = ChromeOptions()
            
            print(f"[ЛОГ] Chrome User Data Dir: {CHROME_USER_DATA_DIR}")
            print(f"[ЛОГ] Chrome Profile Name: {CHROME_PROFILE_NAME}")
            print(f"[ЛОГ] Полный путь к профилю: {profile_path}")
            print(f"[ЛОГ] User Data Dir существует: {os.path.exists(CHROME_USER_DATA_DIR)}")
            print(f"[ЛОГ] Профиль существует: {os.path.exists(profile_path)}")
            
            # Проверяем наличие Chrome.exe
            chrome_paths = [
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
                os.path.expandvars(r"%PROGRAMFILES%\Google\Chrome\Application\chrome.exe"),
                os.path.expandvars(r"%PROGRAMFILES(X86)%\Google\Chrome\Application\chrome.exe")
            ]
            chrome_found = False
            chrome_exe_path = None
            for path in chrome_paths:
                if os.path.exists(path):
                    chrome_found = True
                    chrome_exe_path = path
                    print(f"[ЛОГ] Chrome.exe найден: {path}")
                    break
            
            if not chrome_found:
                print(f"[ЛОГ] ⚠ Chrome.exe не найден в стандартных путях!")
                print(f"[ЛОГ] Проверенные пути:")
                for path in chrome_paths:
                    print(f"[ЛОГ]   - {path}")
            else:
                # НЕ устанавливаем binary_location - пусть Selenium найдет сам
                print(f"[ЛОГ] Chrome найден: {chrome_exe_path}")
            
            # Проверяем, запущен ли Chrome
            chrome_running = check_chrome_running()
            print(f"[ЛОГ] Chrome запущен (по tasklist): {chrome_running}")
            
            # Проверяем lock-файлы до очистки
            lock_files_before = []
            lock_files_to_check = ["SingletonLock", "lockfile", "SingletonSocket", "SingletonCookie", "DevToolsActivePort"]
            for lock_file in lock_files_to_check:
                lock_path = os.path.join(profile_path, lock_file)
                if os.path.exists(lock_path):
                    lock_files_before.append(lock_file)
                    print(f"[ЛОГ] Найден lock-файл: {lock_file} ({lock_path})")
            
            if chrome_running:
                print(f"    ⚠ Chrome уже запущен!")
                print(f"    [Авто] Пытаюсь очистить lock-файлы профиля...")
                
                # Автоматически очищаем lock-файлы
                cleaned = cleanup_profile_locks(profile_path)
                if cleaned:
                    print(f"    ✓ Lock-файлы очищены, пробую запустить...")
                    time.sleep(1)
                else:
                    print(f"    ⚠ Lock-файлы не найдены")
            else:
                # Очищаем старые lock-файлы на всякий случай
                print(f"[ЛОГ] Очистка lock-файлов профиля...")
                cleanup_profile_locks(profile_path)
            
            options.add_argument(f"--user-data-dir={CHROME_USER_DATA_DIR}")
            options.add_argument(f"--profile-directory={CHROME_PROFILE_NAME}")
            print(f"    [Режим] Запуск Chrome с профилем '{CHROME_PROFILE_NAME}'")
        
        # Дополнительные опции для стабильности
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--remote-debugging-port=9223")
        # КРИТИЧНО: отключаем расширения - они блокируют запуск через Selenium
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-plugins")
        options.add_argument("--disable-popup-blocking")
        
        # Специальные опции для headless режима
        if HEADLESS_MODE:
            options.add_argument("--disable-gpu")
            options.add_argument("--disable-software-rasterizer")
            options.add_argument("--window-size=1920,1080")
            options.add_argument("--disable-background-timer-throttling")
            options.add_argument("--disable-backgrounding-occluded-windows")
            options.add_argument("--disable-renderer-backgrounding")
            print(f"[ЛОГ] Добавлены опции для headless режима")
        
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        
        # Логируем все аргументы
        print(f"[ЛОГ] Аргументы командной строки Chrome:")
        for arg in options.arguments:
            print(f"[ЛОГ]   - {arg}")
        
        # Логируем experimental options
        print(f"[ЛОГ] Experimental options:")
        for key, value in options.experimental_options.items():
            print(f"[ЛОГ]   - {key}: {value}")
        
        # Устанавливаем драйвер - ОДНА ПОПЫТКА
        print(f"\n[{BROWSER_TYPE.upper()}Driver] Установка/проверка драйвера...")
        print(f"[ЛОГ] Инициализация {BROWSER_TYPE}DriverManager...")
        
        try:
            if BROWSER_TYPE == 'edge':
                driver_path = EdgeChromiumDriverManager().install()
                print(f"[ЛОГ] EdgeDriver путь: {driver_path}")
                service = EdgeService(driver_path)
                print(f"[ЛОГ] Создание Edge WebDriver...")
                driver = webdriver.Edge(service=service, options=options)
            else:
                print(f"[ЛОГ] Используем UNDETECTED CHROMEDRIVER...")
                
                # Копируем данные из Profile 4 если нужно
                if COPY_PROFILE_DATA and USE_TEMP_PROFILE:
                    source_profile_path = os.path.join(CHROME_USER_DATA_DIR, SOURCE_PROFILE_FOR_COPY)
                    target_profile_path = TEMP_PROFILE_DIR
                    
                    print(f"[ЛОГ] Будет создан профиль парсера с данными из '{SOURCE_PROFILE_FOR_COPY}'")
                    
                    # Копируем данные из Profile 4
                    if os.path.exists(source_profile_path):
                        copy_profile_data(source_profile_path, target_profile_path)
                        # Очищаем lock-файлы в профиле парсера
                        print(f"[ЛОГ] Очистка lock-файлов в профиле парсера...")
                        cleanup_profile_locks(TEMP_PROFILE_DIR)
                        time.sleep(1)  # Небольшая задержка после копирования
                    else:
                        print(f"[!] Профиль '{SOURCE_PROFILE_FOR_COPY}' не найден, запускаю без копирования")
                
                if USE_TEMP_PROFILE:
                    mode_text = "headless (фоновый)" if HEADLESS_MODE else "видимый"
                    print(f"[ЛОГ] Запуск Chrome с профилем: {TEMP_PROFILE_DIR}...")
                    print(f"[ЛОГ] Режим: {mode_text}")
                    
                    # Для headless режима используем use_subprocess=True для стабильности
                    use_subprocess = HEADLESS_MODE
                    
                    # Проверяем, не мешают ли запущенные процессы Chrome
                    chrome_running = check_chrome_running()
                    if chrome_running and HEADLESS_MODE:
                        print(f"[ЛОГ] ⚠ Chrome уже запущен. Это может мешать headless режиму.")
                        print(f"[ЛОГ] Рекомендуется закрыть Chrome перед запуском парсера.")
                        print(f"[ЛОГ] Пробую запустить несмотря на это...")
                        time.sleep(2)  # Даем время на освобождение ресурсов
                    
                    # Пробуем несколько конфигураций для надежности
                    attempts = [
                        {'use_subprocess': use_subprocess, 'version_main': 143},
                        {'use_subprocess': True, 'version_main': 143},
                        {'use_subprocess': True, 'version_main': None},  # Автоопределение версии
                    ]
                    
                    driver = None
                    # Меняем порядок попыток - сначала пробуем без профиля (самый простой вариант)
                    attempts = [
                        {'use_subprocess': True, 'version_main': None, 'user_data_dir': None},  # Без профиля - самый простой
                        {'use_subprocess': True, 'version_main': None, 'user_data_dir': TEMP_PROFILE_DIR},  # С профилем
                        {'use_subprocess': False, 'version_main': None, 'user_data_dir': TEMP_PROFILE_DIR},  # Без subprocess
                        {'use_subprocess': True, 'version_main': 143, 'user_data_dir': TEMP_PROFILE_DIR},   # С указанной версией
                    ]
                    
                    for attempt_num, attempt_config in enumerate(attempts, 1):
                        try:
                            print(f"[ЛОГ] Попытка {attempt_num}/{len(attempts)} запуска Chrome...")
                            user_dir_info = attempt_config.get('user_data_dir', 'временный')
                            print(f"[ЛОГ] Параметры: use_subprocess={attempt_config['use_subprocess']}, version_main={attempt_config['version_main']}, user_data_dir={user_dir_info}")
                            print(f"[ЛОГ] Запускаю Chrome... (таймаут 45 секунд)")
                            
                            # Запускаем Chrome напрямую (без потока для большей надежности)
                            try:
                                # Создаем минимальные опции (undetected-chromedriver сам добавляет нужные)
                                options = ChromeOptions()
                                # Только критичные опции для стабильности
                                options.add_argument("--disable-dev-shm-usage")
                                options.add_argument("--no-sandbox")
                                # НЕ добавляем --remote-debugging-port - uc.Chrome сам управляет портом
                                # НЕ добавляем прокси опции - они могут конфликтовать с uc.Chrome
                                
                                print(f"[ЛОГ] Создаю Chrome драйвер...")
                                # undetected-chromedriver сам управляет профилем и портами
                                driver_kwargs = {
                                    'headless': HEADLESS_MODE,
                                    'use_subprocess': attempt_config['use_subprocess'],
                                    'version_main': attempt_config['version_main'],
                                    'options': options
                                }
                                
                                # Добавляем user_data_dir только если указан (не None)
                                user_dir = attempt_config.get('user_data_dir')
                                if user_dir is not None:
                                    driver_kwargs['user_data_dir'] = user_dir
                                    print(f"[ЛОГ] Использую профиль: {user_dir}")
                                else:
                                    print(f"[ЛОГ] Запускаю Chrome без профиля (временный профиль)")
                                
                                driver = uc.Chrome(**driver_kwargs)
                                
                                # Даем больше времени на полную инициализацию Chrome
                                print(f"[ЛОГ] Chrome драйвер создан, жду инициализацию Chrome...")
                                time.sleep(5)  # Увеличена задержка для полной инициализации Chrome
                                
                                # Проверяем что драйвер действительно работает
                                max_retries = 3
                                driver_works = False
                                for retry in range(max_retries):
                                    try:
                                        driver.current_url  # Простая проверка
                                        print(f"[ЛОГ] ✓ Chrome драйвер создан успешно и отвечает")
                                        driver_works = True
                                        break
                                    except Exception as check_error:
                                        if retry < max_retries - 1:
                                            print(f"[ЛОГ] ⚠ Попытка {retry + 1}/{max_retries}: драйвер еще не готов, жду еще 2 секунды...")
                                            time.sleep(2)
                                        else:
                                            print(f"[ЛОГ] ⚠ Драйвер создан, но не отвечает после {max_retries} попыток: {check_error}")
                                            try:
                                                driver.quit()
                                            except:
                                                pass
                                            driver = None
                                            driver_works = False
                                
                                if driver_works:
                                    break  # Успешно запустили, выходим из цикла попыток
                                elif attempt_num < len(attempts):
                                    print(f"[ЛОГ] Пробую следующую конфигурацию...")
                                    time.sleep(2)
                                    continue
                                else:
                                    raise Exception("Chrome драйвер не отвечает после всех попыток")
                                        
                            except Exception as e:
                                error_msg = str(e)
                                print(f"[ЛОГ] ✗ Ошибка создания Chrome драйвера: {error_msg[:200]}")
                                # Пытаемся убить зависшие процессы Chrome
                                try:
                                    subprocess.run(['taskkill', '/F', '/IM', 'chrome.exe'], 
                                                 capture_output=True, timeout=5)
                                    time.sleep(2)
                                except:
                                    pass
                                
                                if attempt_num < len(attempts):
                                    print(f"[ЛОГ] Пробую следующую конфигурацию...")
                                    time.sleep(2)
                                    continue
                                else:
                                    raise
                            
                            if not driver:
                                print(f"[ЛОГ] ⚠ Chrome не запустился. Пробую следующую конфигурацию...")
                                if attempt_num < len(attempts):
                                    time.sleep(2)
                                    continue
                                else:
                                    raise Exception("Chrome не запустился после всех попыток")
                            print(f"[ЛОГ] Chrome процесс запущен, проверяю работоспособность...")
                            
                            # Проверяем что драйвер работает
                            try:
                                driver.current_url  # Простая проверка работоспособности
                                print(f"[ЛОГ] ✓ Chrome запущен с профилем парсера (данные из Profile 4)")
                                break  # Успешно запустили, выходим из цикла
                            except Exception as check_error:
                                print(f"[ЛОГ] ⚠ Драйвер создан, но не отвечает: {check_error}")
                                try:
                                    driver.quit()
                                except:
                                    pass
                                driver = None
                                if attempt_num < len(attempts):
                                    print(f"[ЛОГ] Пробую следующую конфигурацию...")
                                    time.sleep(2)
                                    continue
                                else:
                                    raise Exception("Драйвер не отвечает после всех попыток")
                                    
                        except (ConnectionResetError, ConnectionError, ConnectionAbortedError) as conn_error:
                            error_msg = str(conn_error)
                            print(f"[ЛОГ] ✗ Попытка {attempt_num} не удалась: {type(conn_error).__name__}: {error_msg[:200]}")
                            
                            if attempt_num < len(attempts):
                                print(f"[ЛОГ] Ошибка подключения. Очищаю lock-файлы и пробую еще раз...")
                                cleanup_profile_locks(TEMP_PROFILE_DIR)
                                time.sleep(3)
                                continue
                            else:
                                raise
                                
                        except Exception as e:
                            error_msg = str(e)
                            print(f"[ЛОГ] ✗ Попытка {attempt_num} не удалась: {error_msg[:200]}")
                            
                            # Если ошибка связана с подключением, пробуем еще раз с задержкой
                            if any(keyword in error_msg.lower() for keyword in ["cannot connect", "not reachable", "connection", "reset", "refused"]):
                                if attempt_num < len(attempts):
                                    print(f"[ЛОГ] Ошибка подключения. Очищаю lock-файлы и пробую еще раз...")
                                    cleanup_profile_locks(TEMP_PROFILE_DIR)
                                    time.sleep(3)
                                    continue
                                else:
                                    raise
                            elif attempt_num < len(attempts):
                                print(f"[ЛОГ] Пробую следующую конфигурацию...")
                                time.sleep(2)
                                continue
                            else:
                                raise
                    
                    if driver is None:
                        raise Exception("Не удалось запустить Chrome после всех попыток")
                else:
                    mode_text = "headless (фоновый)" if HEADLESS_MODE else "видимый"
                    print(f"[ЛОГ] Запуск Chrome БЕЗ профиля (временный)...")
                    print(f"[ЛОГ] Режим: {mode_text}")
                    
                    # Пробуем несколько конфигураций для надежности
                    driver = None
                    # Меняем порядок попыток - сначала пробуем без профиля (самый простой вариант)
                    attempts_no_profile = [
                        {'use_subprocess': True, 'version_main': None, 'user_data_dir': None},  # Без профиля - самый простой
                        {'use_subprocess': True, 'version_main': None},   # С временным профилем
                        {'use_subprocess': False, 'version_main': None},  # Без subprocess
                        {'use_subprocess': True, 'version_main': 143},    # С указанной версией
                    ]
                    
                    for attempt_num, attempt_config in enumerate(attempts_no_profile, 1):
                        try:
                            print(f"[ЛОГ] Попытка {attempt_num}/{len(attempts_no_profile)} запуска Chrome...")
                            user_dir_info = attempt_config.get('user_data_dir', 'временный')
                            print(f"[ЛОГ] Параметры: use_subprocess={attempt_config['use_subprocess']}, version_main={attempt_config['version_main']}, user_data_dir={user_dir_info}")
                            print(f"[ЛОГ] ⚙ Отключаю прокси/хост браузера (если был настроен)...")
                            print(f"[ЛОГ] Запускаю Chrome... (таймаут 45 секунд)")
                            
                            # Запускаем Chrome напрямую (без потока для большей надежности)
                            try:
                                # Создаем минимальные опции (undetected-chromedriver сам добавляет нужные)
                                options = ChromeOptions()
                                # Только критичные опции для стабильности
                                options.add_argument("--disable-dev-shm-usage")
                                options.add_argument("--no-sandbox")
                                # НЕ добавляем --remote-debugging-port - uc.Chrome сам управляет портом
                                
                                print(f"[ЛОГ] Создаю Chrome драйвер...")
                                driver_kwargs = {
                                    'headless': HEADLESS_MODE,
                                    'use_subprocess': attempt_config['use_subprocess'],
                                    'version_main': attempt_config['version_main'],
                                    'options': options
                                }
                                
                                # Добавляем user_data_dir только если указан (не None)
                                user_dir = attempt_config.get('user_data_dir')
                                if user_dir is not None:
                                    driver_kwargs['user_data_dir'] = user_dir
                                    print(f"[ЛОГ] Использую профиль: {user_dir}")
                                else:
                                    print(f"[ЛОГ] Запускаю Chrome без профиля (временный профиль)")
                                
                                driver = uc.Chrome(**driver_kwargs)
                                
                                # Даем больше времени на полную инициализацию Chrome
                                print(f"[ЛОГ] Chrome драйвер создан, жду инициализацию Chrome...")
                                time.sleep(5)  # Увеличена задержка для полной инициализации Chrome
                                
                                # Проверяем что драйвер действительно работает
                                max_retries = 3
                                driver_works = False
                                for retry in range(max_retries):
                                    try:
                                        driver.current_url  # Простая проверка
                                        print(f"[ЛОГ] ✓ Chrome драйвер создан успешно и отвечает")
                                        print(f"[ЛОГ] ✓ Chrome запущен с временным профилем")
                                        driver_works = True
                                        break  # Успешно запустили, выходим из цикла
                                    except Exception as check_error:
                                        if retry < max_retries - 1:
                                            print(f"[ЛОГ] ⚠ Попытка {retry + 1}/{max_retries}: драйвер еще не готов, жду еще 2 секунды...")
                                            time.sleep(2)
                                        else:
                                            print(f"[ЛОГ] ⚠ Драйвер создан, но не отвечает после {max_retries} попыток: {check_error}")
                                            try:
                                                driver.quit()
                                            except:
                                                pass
                                            driver = None
                                            driver_works = False
                                
                                if driver_works:
                                    break  # Успешно запустили, выходим из цикла попыток
                                elif attempt_num < len(attempts_no_profile):
                                    print(f"[ЛОГ] Пробую следующую конфигурацию...")
                                    time.sleep(2)
                                    continue
                                else:
                                    raise Exception("Chrome драйвер не отвечает после всех попыток")
                                        
                            except Exception as e:
                                error_msg = str(e)
                                print(f"[ЛОГ] ✗ Ошибка создания Chrome драйвера: {error_msg[:200]}")
                                # Пытаемся убить зависшие процессы Chrome
                                try:
                                    subprocess.run(['taskkill', '/F', '/IM', 'chrome.exe'], 
                                                 capture_output=True, timeout=5)
                                    time.sleep(2)
                                except:
                                    pass
                                
                                if attempt_num < len(attempts_no_profile):
                                    print(f"[ЛОГ] Пробую следующую конфигурацию...")
                                    time.sleep(2)
                                    continue
                                else:
                                    raise
                                    
                        except (ConnectionResetError, ConnectionError, ConnectionAbortedError) as conn_error:
                            error_msg = str(conn_error)
                            print(f"[ЛОГ] ✗ Попытка {attempt_num} не удалась: {type(conn_error).__name__}: {error_msg[:200]}")
                            
                            if attempt_num < len(attempts_no_profile):
                                print(f"[ЛОГ] Ошибка подключения. Пробую еще раз...")
                                time.sleep(3)
                                continue
                            else:
                                raise
                                
                        except Exception as e:
                            error_msg = str(e)
                            print(f"[ЛОГ] ✗ Попытка {attempt_num} не удалась: {error_msg[:200]}")
                            
                            # Если ошибка связана с подключением, пробуем еще раз с задержкой
                            if any(keyword in error_msg.lower() for keyword in ["cannot connect", "not reachable", "connection", "reset", "refused"]):
                                if attempt_num < len(attempts_no_profile):
                                    print(f"[ЛОГ] Ошибка подключения. Пробую еще раз...")
                                    time.sleep(3)
                                    continue
                                else:
                                    raise
                            elif attempt_num < len(attempts_no_profile):
                                print(f"[ЛОГ] Пробую следующую конфигурацию...")
                                time.sleep(2)
                                continue
                            else:
                                raise
                    
                    if driver is None:
                        raise Exception("Не удалось запустить Chrome после всех попыток")
            
            # Проверяем что driver создан
            if driver is None:
                raise Exception("Драйвер не был создан")
            
            print(f"[ЛОГ] ✓ WebDriver создан успешно")
            try:
                print(f"[ЛОГ] Session ID: {driver.session_id}")
                print(f"[ЛОГ] Capabilities: {driver.capabilities}")
            except Exception as e:
                print(f"[ЛОГ] ⚠ Не удалось получить информацию о сессии: {e}")
            
            # Скрываем webdriver
            try:
                driver.execute_cdp_cmd('Network.setUserAgentOverride', {
                    "userAgent": driver.execute_script("return navigator.userAgent").replace('Headless', '')
                })
            except Exception as e:
                print(f"[ЛОГ] ⚠ Не удалось установить User-Agent: {e}")
            
            return driver
            
        except (ConnectionResetError, ConnectionError, ConnectionAbortedError) as conn_error:
            import traceback
            print(f"\n{'='*60}")
            print(f"[ОШИБКА] Ошибка подключения к Chrome")
            print(f"{'='*60}")
            print(f"[ЛОГ] Тип: {type(conn_error).__name__}")
            print(f"[ЛОГ] Сообщение: {str(conn_error)}")
            print(f"{'='*60}\n")
            
            print(f"\n💡 ВОЗМОЖНЫЕ ПРИЧИНЫ:")
            print(f"   1. Chrome запустился, но соединение было разорвано")
            print(f"   2. Антивирус или файрвол блокирует соединение")
            print(f"   3. Порт 9223 (remote-debugging-port) занят другим процессом")
            print(f"   4. Профиль поврежден или имеет проблемы с правами доступа")
            print(f"\n💡 РЕШЕНИЯ:")
            print(f"   1. Закройте ВСЕ окна Chrome: taskkill /F /IM chrome.exe")
            print(f"   2. Подождите 10 секунд и попробуйте снова")
            print(f"   3. Перезагрузите компьютер (если Chrome завис)")
            print(f"   4. Проверьте антивирус (может блокировать)")
            print(f"   5. Попробуйте запустить Chrome вручную и закройте его")
            print(f"   6. Удалите папку chrome_parser_profile и дайте парсеру создать новую")
            return None
            
        except Exception as e:
            import traceback
            print(f"\n{'='*60}")
            print(f"[ОШИБКА] Детальная информация")
            print(f"{'='*60}")
            print(f"[ЛОГ] Тип: {type(e).__name__}")
            print(f"[ЛОГ] Сообщение: {str(e)}")
            print(f"\n[ЛОГ] Полный traceback:")
            traceback.print_exc()
            print(f"{'='*60}\n")
            
            print(f"\n💡 ВОЗМОЖНЫЕ ПРИЧИНЫ:")
            print(f"   1. Профиль '{CHROME_PROFILE_NAME}' используется другим процессом Chrome")
            print(f"   2. Профиль поврежден или имеет проблемы с правами доступа")
            print(f"   3. Несовместимость версий Chrome ({chrome_exe_path if BROWSER_TYPE == 'chrome' else 'Edge'}) и ChromeDriver")
            print(f"   4. Антивирус блокирует запуск Chrome через Selenium")
            print(f"\n💡 РЕШЕНИЯ:")
            print(f"   1. Закройте ВСЕ окна Chrome: taskkill /F /IM chrome.exe")
            print(f"   2. Подождите 10 секунд и попробуйте снова")
            print(f"   3. Попробуйте другой профиль (измените CHROME_PROFILE_NAME)")
            print(f"   4. Используйте Edge: BROWSER_TYPE = 'edge'")
            print(f"   5. Удалите папку chrome_parser_profile и дайте парсеру создать новую")
            return None


def human_delay(min_sec=1, max_sec=3):
    """Случайная задержка как у человека"""
    delay = random.uniform(min_sec, max_sec)
    time.sleep(delay)


def parse_price_from_current_page(driver, article, product_url=None):
    """
    Парсит цены с текущей открытой страницы товара
    НЕ открывает и НЕ закрывает вкладки - это делает вызывающая функция
    Возвращает словарь {'price': обычная цена, 'price_with_card': цена с картой}
    или 0 если товара нет в наличии
    """
    try:
        # Минимальная задержка для загрузки страницы (основная задержка будет при повторной попытке если данные не найдены)
        time.sleep(0.5)
        
        # Проверяем на captcha и блокировку WB
        page_source_lower = driver.page_source.lower()
        if "Почти готово" in driver.title or "captcha" in page_source_lower:
            print(f"  [{article}] ⚠ Captcha обнаружена!")
            return None  # None = нужна повторная попытка
        
        # Проверяем на блокировку WB из-за подозрительной активности
        if "подозрительная активность" in page_source_lower or "suspicious activity" in page_source_lower:
            print(f"  [{article}] ⚠⚠⚠ WB ЗАБЛОКИРОВАЛ из-за подозрительной активности!")
            print(f"  [{article}] Рекомендуется увеличить задержки или уменьшить PARALLEL_TABS")
            return None  # None = нужна повторная попытка
        
        # КРИТИЧНО: Проверяем наличие элемента "Нет в наличии"
        try:
            sold_out_element = driver.find_element(By.CSS_SELECTOR, "h2[class*='soldOutProduct']")
            print(f"  [{article}] ⚠ Товар недоступен: {sold_out_element.text}")
            return {'price': 0, 'price_with_card': 0}
        except:
            pass  # Элемент не найден - товар в наличии
        
        # Дополнительная проверка по ключевым словам
        page_text = driver.page_source.lower()
        unavailable_keywords = ['нет в наличии', 'товар недоступен', 'недоступен для заказа', 'закончился', 'распродан']
        
        for keyword in unavailable_keywords:
            if keyword in page_text:
                print(f"  [{article}] ⚠ Товар недоступен: '{keyword}'")
                return {'price': 0, 'price_with_card': 0}
        
        # Кликаем на кнопку кошелька (если есть)
        try:
            wallet_button = WebDriverWait(driver, 2).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "button[class*='priceBlockWalletPrice']"))
            )
            wallet_button.click()
            time.sleep(0.5)  # Минимальная задержка для появления финальной цены
        except:
            pass  # Кнопки кошелька нет - это нормально
        
        # Ищем элемент с обычной ценой (mo-typography_color_primary)
        price_selectors = [
            (By.CSS_SELECTOR, "h2.mo-typography_color_primary"),
            (By.CSS_SELECTOR, "h2[class*='mo-typography'][class*='color_primary']"),
            (By.CSS_SELECTOR, "ins.priceBlockFinalPrice--iToZR"),
            (By.CSS_SELECTOR, "ins[class*='priceBlockFinalPrice']"),
            (By.CSS_SELECTOR, "ins.mo-typography[class*='priceBlockFinalPrice']"),
            (By.CSS_SELECTOR, "ins[class*='priceBlockFinalPrice'][class*='mo-typography']"),
            (By.CSS_SELECTOR, "ins[class*='FinalPrice']"),
            (By.CSS_SELECTOR, "span[class*='final-price']"),
            (By.CSS_SELECTOR, "ins[class*='price']"),
        ]
        
        # Ищем элемент с ценой с картой (mo-typography_color_danger - красная цена)
        price_with_card_selectors = [
            (By.CSS_SELECTOR, "h2.mo-typography_color_danger"),
            (By.CSS_SELECTOR, "h2[class*='mo-typography'][class*='color_danger']"),
        ]
        
        price = None
        price_with_card = None
        
        # Ищем обе цены одновременно для ускорения
        # Сначала пробуем найти обычную цену (приоритет)
        for by, selector in price_selectors:
            try:
                price_elem = WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((by, selector))
                )
                price_text = price_elem.text.strip()
                price_num = re.sub(r'[^\d]', '', price_text)
                if price_num:
                    price = int(price_num)
                    break
            except:
                continue
        
        # Параллельно ищем цену с картой (не блокируем поиск обычной цены)
        for by, selector in price_with_card_selectors:
            try:
                price_card_elem = driver.find_element(by, selector)
                price_card_text = price_card_elem.text.strip()
                price_card_num = re.sub(r'[^\d]', '', price_card_text)
                if price_card_num:
                    price_with_card = int(price_card_num)
                    break
            except:
                continue
        
        # Если обычная цена не найдена, переходим на страницу, ждем 5 секунд и проверяем заново
        if not price:
            print(f"  [{article}] ⚠ Обычная цена не найдена с первой попытки, перезагружаю страницу и жду 5 секунд...")
            try:
                # Переходим на страницу товара заново
                current_url = driver.current_url
                if product_url:
                    driver.get(product_url)
                else:
                    driver.get(current_url)  # Перезагружаем текущую страницу
                time.sleep(5)  # Ждем 5 секунд для загрузки страницы товара
            except Exception as e:
                print(f"  [{article}] ⚠ Ошибка перезагрузки страницы: {e}, просто жду 5 секунд...")
                time.sleep(5)  # Если не удалось перейти, просто ждем
            
            # Повторная попытка найти обычную цену
            for by, selector in price_selectors:
                try:
                    price_elem = WebDriverWait(driver, 8).until(
                        EC.presence_of_element_located((by, selector))
                    )
                    price_text = price_elem.text.strip()
                    price_num = re.sub(r'[^\d]', '', price_text)
                    if price_num:
                        price = int(price_num)
                        print(f"  [{article}] ✓ Обычная цена найдена со второй попытки: {price} ₽")
                        break
                except:
                    continue
            
            # Повторная попытка найти цену с картой
            if not price_with_card:
                for by, selector in price_with_card_selectors:
                    try:
                        price_card_elem = driver.find_element(by, selector)
                        price_card_text = price_card_elem.text.strip()
                        price_card_num = re.sub(r'[^\d]', '', price_card_text)
                        if price_card_num:
                            price_with_card = int(price_card_num)
                            break
                    except:
                        continue
        
        if not price:
            print(f"  [{article}] ✗ Обычная цена не найдена даже после повторной попытки")
            return {'price': 0, 'price_with_card': 0}
        
        # Возвращаем словарь с обеими ценами
        # Если цена с картой не найдена - возвращаем 0
        return {
            'price': price,
            'price_with_card': price_with_card if price_with_card else 0
        }
    
    except Exception as e:
        print(f"  [{article}] ✗ Ошибка парсинга: {e}")
        return {'price': 0, 'price_with_card': None}


def process_products_parallel(driver, products):
    """
    Обрабатывает товары параллельно по PARALLEL_TABS штук
    Возвращает кортеж (список результатов, количество сохраненных товаров)
    """
    results = []
    last_saved_count = 0  # Счетчик последнего сохранения
    try:
        main_window = driver.window_handles[0]
    except (InvalidSessionIdException, Exception) as e:
        print(f"\n[!] ОШИБКА: Браузер закрыт или сессия потеряна: {e}")
        print(f"    Возвращаю уже собранные результаты: {len(results)} товаров")
        return (results, last_saved_count)
    
    total = len(products)
    
    print(f"\n{'='*80}")
    print(f"ПАРАЛЛЕЛЬНАЯ ОБРАБОТКА: {PARALLEL_TABS} вкладок одновременно")
    print(f"{'='*80}\n")
    
    # Обрабатываем товары пачками
    try:
        for batch_start in range(0, total, PARALLEL_TABS):
            batch = products[batch_start : batch_start + PARALLEL_TABS]
            batch_num = batch_start // PARALLEL_TABS + 1
            total_batches = (total + PARALLEL_TABS - 1) // PARALLEL_TABS
            
            print(f"\n{'─'*80}")
            print(f"📦 ПАКЕТ {batch_num}/{total_batches} ({len(batch)} товаров)")
            print(f"{'─'*80}")
            
            # ФАЗА 1: Открыть все вкладки пакета
            print(f"\n[1/4] Открываю {len(batch)} вкладок...")
            
            # Убеждаемся что мы на главной вкладке
            try:
                driver.switch_to.window(main_window)
            except:
                main_window = driver.window_handles[0]
                driver.switch_to.window(main_window)
            
            # Открываем все вкладки и сохраняем соответствие между вкладками и товарами
            initial_handles_count = len(driver.window_handles)
            print(f"  [ЛОГ] Вкладок до открытия: {initial_handles_count}")
            
            # Словарь для сохранения соответствия: handle -> product
            opened_tabs_map = {}
            
            for idx, product in enumerate(batch):
                try:
                    print(f"  [{batch_start + idx + 1}/{total}] Открываю: {product['article']}")
                    # Открываем вкладку
                    driver.execute_script("window.open(arguments[0], '_blank');", product['url'])
                    time.sleep(0.1)  # Небольшая задержка для открытия вкладки
                    
                    # Получаем handle новой вкладки (последняя открытая)
                    try:
                        all_handles = driver.window_handles
                        if len(all_handles) > initial_handles_count + idx:
                            new_tab_handle = all_handles[-1]
                            # Сохраняем соответствие между вкладкой и товаром
                            opened_tabs_map[new_tab_handle] = product
                            # Переключаемся на новую вкладку чтобы она точно открылась и загрузилась
                            driver.switch_to.window(new_tab_handle)
                            time.sleep(0.1)  # Даем время на загрузку URL
                            # Возвращаемся на главную вкладку
                            driver.switch_to.window(main_window)
                    except Exception as tab_error:
                        print(f"      [ЛОГ] ⚠ Ошибка при сохранении соответствия вкладки: {tab_error}")
                    
                    current_handles = len(driver.window_handles)
                    print(f"      [ЛОГ] Вкладок после открытия: {current_handles}")
                except Exception as e:
                    print(f"  [{batch_start + idx + 1}/{total}] ⚠ Ошибка: {e}")
                    import traceback
                    traceback.print_exc()
            
            # ФАЗА 2: Ждем загрузки всех вкладок
            print(f"\n[2/4] Жду полной загрузки страниц...")
            time.sleep(0.5)  # Минимальная задержка для открытия вкладок
            
            # Получаем все вкладки кроме главной
            try:
                all_handles = driver.window_handles
                tabs = [h for h in all_handles if h != main_window]
                print(f"  [ЛОГ] Всего окон: {len(all_handles)}, вкладок для парсинга: {len(tabs)}")
                
                # Если у нас есть сохраненное соответствие, используем его
                # Иначе создаем соответствие по порядку (fallback)
                if not opened_tabs_map and len(tabs) == len(batch):
                    # Fallback: предполагаем что порядок совпадает
                    for idx, tab_handle in enumerate(tabs):
                        if idx < len(batch):
                            opened_tabs_map[tab_handle] = batch[idx]
            except Exception as e:
                print(f"  ⚠ Ошибка получения вкладок: {e}")
                tabs = []
            
            # Минимальная задержка для загрузки страниц (основная задержка будет при парсинге каждой страницы)
            time.sleep(1)
            
            if len(tabs) == 0:
                print(f"  ⚠ ВНИМАНИЕ: Вкладки не открылись! Пробую еще раз...")
                # Пробуем открыть еще раз с переключением на каждую вкладку
                driver.switch_to.window(main_window)
                for idx, product in enumerate(batch):
                    try:
                        driver.execute_script(f"window.open('{product['url']}', '_blank');")
                        time.sleep(0.1)
                        # Переключаемся на новую вкладку чтобы она точно открылась
                        if len(driver.window_handles) > initial_handles_count + idx + 1:
                            driver.switch_to.window(driver.window_handles[-1])
                            time.sleep(0.05)
                            driver.switch_to.window(main_window)
                    except Exception as e:
                        print(f"  [ЛОГ] Ошибка при повторном открытии вкладки {idx+1}: {e}")
                time.sleep(0.5)
                try:
                    all_handles = driver.window_handles
                    tabs = [h for h in all_handles if h != main_window]
                    print(f"  [ЛОГ] После повторной попытки: {len(tabs)} вкладок")
                except:
                    tabs = []
            
            print(f"  ✓ Все {len(tabs)} вкладок загружены")
            
            # ФАЗА 3: Парсим цены из всех вкладок
            print(f"\n[3/4] Парсинг цен...")
            
            # Сначала создаем словарь соответствия вкладок и товаров по URL
            # Это гарантирует, что мы парсим цену правильного товара
            tab_to_product = {}
            
            # Используем сохраненное соответствие если есть
            if opened_tabs_map:
                tab_to_product = opened_tabs_map.copy()
                print(f"  [ЛОГ] Использую сохраненное соответствие вкладок и товаров")
            else:
                # Создаем соответствие по URL на каждой вкладке
                for tab_handle in tabs:
                    try:
                        driver.switch_to.window(tab_handle)
                        time.sleep(0.1)  # Небольшая задержка для загрузки URL
                        current_url = driver.current_url
                        
                        # Ищем соответствующий товар по URL
                        for product in batch:
                            # Проверяем совпадение по артикулу в URL
                            if product['article'] in current_url or product['url'] in current_url:
                                tab_to_product[tab_handle] = product
                                break
                    except Exception as e:
                        print(f"  [ЛОГ] Ошибка при проверке вкладки {tab_handle}: {e}")
                        continue
            
            # Теперь парсим цены, гарантируя соответствие товара и вкладки
            for idx, product in enumerate(batch):
                try:
                    # Ищем вкладку с правильным товаром
                    matching_tab = None
                    for tab_handle, tab_product in tab_to_product.items():
                        if tab_product['article'] == product['article']:
                            matching_tab = tab_handle
                            break
                    
                    if not matching_tab:
                        # Если не нашли вкладку, пробуем найти по порядку (fallback)
                        if idx < len(tabs):
                            matching_tab = tabs[idx]
                            # Проверяем что это правильный товар
                            try:
                                driver.switch_to.window(matching_tab)
                                current_url = driver.current_url
                                if product['article'] not in current_url and product['url'] not in current_url:
                                    print(f"  [{batch_start + idx + 1}/{total}] ⚠ Вкладка не соответствует товару {product['article']}, ищу правильную...")
                                    # Пробуем найти правильную вкладку среди всех открытых
                                    found_correct_tab = False
                                    for tab_handle in tabs:
                                        try:
                                            driver.switch_to.window(tab_handle)
                                            tab_url = driver.current_url
                                            if product['article'] in tab_url or product['url'] in tab_url:
                                                matching_tab = tab_handle
                                                found_correct_tab = True
                                                break
                                        except:
                                            continue
                                    
                                    if not found_correct_tab:
                                        print(f"  [{batch_start + idx + 1}/{total}] ✗ Не удалось найти вкладку с товаром {product['article']}")
                                        results.append({
                                            'url': product['url'],
                                            'article': product['article'],
                                            'price': 0,
                                            'price_with_card': 0
                                        })
                                        continue
                            except:
                                pass
                        else:
                            print(f"  [{batch_start + idx + 1}/{total}] ⚠ Вкладка для товара {product['article']} не найдена")
                            results.append({
                                'url': product['url'],
                                'article': product['article'],
                                'price': 0,
                                'price_with_card': 0
                            })
                            continue
                    
                    # Переключаемся на правильную вкладку
                    driver.switch_to.window(matching_tab)
                    
                    # КРИТИЧНО: Финальная проверка - убеждаемся что на вкладке правильный товар
                    try:
                        current_url = driver.current_url
                        if product['article'] not in current_url and product['url'] not in current_url:
                            print(f"  [{batch_start + idx + 1}/{total}] ⚠ КРИТИЧНО: На вкладке неверный товар!")
                            print(f"      Ожидается: {product['article']} ({product['url'][:50]}...)")
                            print(f"      На вкладке: {current_url[:80]}...")
                            # Пробуем найти правильную вкладку среди всех открытых
                            found_correct_tab = False
                            for tab_handle in tabs:
                                try:
                                    driver.switch_to.window(tab_handle)
                                    tab_url = driver.current_url
                                    if product['article'] in tab_url or product['url'] in tab_url:
                                        matching_tab = tab_handle
                                        found_correct_tab = True
                                        print(f"      ✓ Найдена правильная вкладка для {product['article']}")
                                        break
                                except:
                                    continue
                            
                            if not found_correct_tab:
                                print(f"  [{batch_start + idx + 1}/{total}] ✗ Не удалось найти вкладку с товаром {product['article']}")
                                results.append({
                                    'url': product['url'],
                                    'article': product['article'],
                                    'price': 0,
                                    'price_with_card': 0
                                })
                                continue
                    except Exception as url_check_error:
                        print(f"  [ЛОГ] ⚠ Не удалось проверить URL вкладки: {url_check_error}")
                        # Продолжаем парсинг, но с предупреждением
                    
                    # Парсим цену с правильной вкладки
                    price_data = parse_price_from_current_page(driver, product['article'], product['url'])
                    
                    # Если captcha - пропускаем
                    if price_data is None:
                        price_data = {'price': 0, 'price_with_card': 0}
                    
                    # Если вернулось число (старый формат), преобразуем в словарь
                    if isinstance(price_data, (int, float)):
                        price_data = {'price': int(price_data), 'price_with_card': 0}
                    
                    results.append({
                        'url': product['url'],
                        'article': product['article'],
                        'price': price_data['price'],
                        'price_with_card': price_data.get('price_with_card')
                    })
                    
                    price = price_data['price']
                    price_card = price_data.get('price_with_card', 0)
                    if price_card and price_card > 0:
                        status = f"{price} ₽ / {price_card} ₽ (с картой)"
                    else:
                        status = f"{price} ₽" if price > 0 else "недоступен" if price == 0 else "ошибка"
                    print(f"  [{batch_start + idx + 1}/{total}] {product['article']}: {status}")
                
                except Exception as e:
                    print(f"  [{batch_start + idx + 1}/{total}] {product['article']}: ✗ ошибка - {e}")
                    results.append({
                        'url': product['url'],
                        'article': product['article'],
                        'price': 0,
                        'price_with_card': 0
                    })
                    
                    # Промежуточное сохранение каждые 10 товаров (даже при ошибках)
                    if SAVE_INTERMEDIATE_RESULTS and len(results) - last_saved_count >= SAVE_EVERY_N_PRODUCTS:
                        print(f"\n💾 Промежуточное сохранение ({len(results)} товаров)...")
                        if GOOGLE_SHEETS_ENABLED and GOOGLE_SHEET_URL:
                            print(f"📊 Запись в Google Таблицы ({len(results)} товаров)...")
                            if save_results_to_google_sheets(results, GOOGLE_SHEET_URL, GOOGLE_SHEET_NAME):
                                print(f"✓ Сохранено в Google Таблицы")
                                last_saved_count = len(results)  # Обновляем счетчик
                            else:
                                print(f"⚠ Не удалось сохранить в Google Таблицы")
                        else:
                            print(f"⚠ Google Таблицы не настроены (GOOGLE_SHEETS_ENABLED = False или URL не указан)")
            
            # ФАЗА 4: Закрыть все вкладки пакета
            print(f"\n[4/4] Закрываю вкладки...")
            for tab_handle in tabs:
                try:
                    driver.switch_to.window(tab_handle)
                    driver.close()
                except:
                    pass
            
            # Возвращаемся на главную вкладку и обновляем main_window
            try:
                # Обновляем main_window - берем первую доступную вкладку
                if driver.window_handles:
                    main_window = driver.window_handles[0]
                    driver.switch_to.window(main_window)
                else:
                    print(f"  ⚠ Все вкладки закрыты!")
            except Exception as e:
                print(f"  ⚠ Ошибка переключения на главную вкладку: {e}")
                # Пробуем получить любую доступную вкладку
                try:
                    if driver.window_handles:
                        main_window = driver.window_handles[0]
                        driver.switch_to.window(main_window)
                except:
                    pass
            
            # Дополнительная проверка сохранения в конце пакета (если накопилось >= 10 товаров с последнего сохранения)
            if SAVE_INTERMEDIATE_RESULTS and len(results) - last_saved_count >= SAVE_EVERY_N_PRODUCTS:
                print(f"\n💾 Промежуточное сохранение в конце пакета ({len(results)} товаров)...")
                if GOOGLE_SHEETS_ENABLED and GOOGLE_SHEET_URL:
                    # Сохраняем только новые результаты (с last_saved_count до конца)
                    new_results = results[last_saved_count:]
                    print(f"📊 Запись в Google Таблицы ({len(new_results)} новых товаров)...")
                    if save_results_to_google_sheets(new_results, GOOGLE_SHEET_URL, GOOGLE_SHEET_NAME, append_only=True):
                        print(f"✓ Сохранено в Google Таблицы")
                        last_saved_count = len(results)  # Обновляем счетчик
                    else:
                        print(f"⚠ Не удалось сохранить в Google Таблицы")
                else:
                    print(f"⚠ Google Таблицы не настроены (GOOGLE_SHEETS_ENABLED = False или URL не указан)")
            
            # Минимальная задержка между пакетами (сокращена для ускорения)
            if batch_start + PARALLEL_TABS < total:
                delay = 0.5  # Минимальная задержка вместо 2-4 секунд
                print(f"\n⏸ Пауза {delay:.1f}с перед следующим пакетом...\n")
                time.sleep(delay)
    
    except (InvalidSessionIdException, Exception) as e:
        print(f"\n[!] КРИТИЧЕСКАЯ ОШИБКА в process_products_parallel: {e}")
        print(f"    Возвращаю уже собранные результаты: {len(results)} товаров")
        import traceback
        traceback.print_exc()
        return (results, last_saved_count)  # Возвращаем то, что успели собрать
    
    return (results, last_saved_count)


def get_price_from_product_page(driver, product_url, article):
    """
    Открывает карточку товара по ссылке и извлекает цену
    Возвращает цену или 0 если товара нет в наличии
    """
    try:
        print(f"\n[{article}] Открываю карточку в новой вкладке...")
        print(f"  URL: {product_url}")
        
        # Открываем в новой вкладке того же окна
        driver.execute_script("window.open(arguments[0], '_blank');", product_url)
        
        # Переключаемся на новую вкладку
        driver.switch_to.window(driver.window_handles[-1])
        
        human_delay(2, 4)
        
        # Проверяем на captcha
        if "Почти готово" in driver.title or "captcha" in driver.page_source.lower():
            print(f"  ⚠ Captcha! Жду 10 сек...")
            time.sleep(10)
            driver.get(product_url)
            human_delay(2, 4)
        
        # КРИТИЧНО: Проверяем наличие элемента "Нет в наличии"
        # <h2 class="... soldOutProduct--vCzrv">Нет в наличии</h2>
        try:
            sold_out_element = driver.find_element(By.CSS_SELECTOR, "h2[class*='soldOutProduct']")
            print(f"  ⚠ Товар недоступен: найден элемент 'soldOutProduct' - {sold_out_element.text}")
            # Закрываем вкладку и пропускаем товар
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            return 0
        except:
            pass  # Элемент не найден - товар в наличии
        
        # Дополнительная проверка по ключевым словам (fallback)
        page_text = driver.page_source.lower()
        unavailable_keywords = [
            'нет в наличии',
            'товар недоступен',
            'недоступен для заказа',
            'закончился',
            'распродан'
        ]
        
        is_unavailable = False
        for keyword in unavailable_keywords:
            if keyword in page_text:
                is_unavailable = True
                print(f"  ⚠ Товар недоступен: найдено '{keyword}'")
                break
        
        if is_unavailable:
            # Закрываем вкладку если товар недоступен
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            return 0
        
        # НОВАЯ ЛОГИКА: Сначала кликаем на кнопку кошелька (если есть)
        # Это открывает финальную цену с учетом всех скидок
        try:
            # Ищем кнопку с кошельком (класс priceBlockWalletPrice)
            wallet_button = WebDriverWait(driver, 3).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "button[class*='priceBlockWalletPrice']"))
            )
            print(f"  ⚠ Найдена кнопка кошелька, кликаю...")
            wallet_button.click()
            human_delay(1, 2)  # Ждем появления финальной цены
        except:
            # Кнопки кошелька нет - это нормально, продолжаем
            print(f"  ℹ Кнопка кошелька не найдена, ищу обычную цену")
        
        # Ищем элемент с финальной ценой
        # Приоритет 1: h2 с классом mo-typography_color_primary (появляется после клика на кошелек)
        # Приоритет 2: ins.priceBlockFinalPrice (обычная цена)
        price_selectors = [
            # Финальная цена после клика на кошелек
            (By.CSS_SELECTOR, "h2.mo-typography_color_primary"),
            (By.CSS_SELECTOR, "h2[class*='mo-typography'][class*='color_primary']"),
            # Обычная цена
            (By.CSS_SELECTOR, "ins.priceBlockFinalPrice--iToZR"),
            (By.CSS_SELECTOR, "ins[class*='priceBlockFinalPrice']"),
            (By.CSS_SELECTOR, "ins.mo-typography[class*='priceBlockFinalPrice']"),
            (By.CSS_SELECTOR, "ins[class*='priceBlockFinalPrice'][class*='mo-typography']"),
            # Fallback селекторы
            (By.CSS_SELECTOR, "ins[class*='FinalPrice']"),
            (By.CSS_SELECTOR, "span[class*='final-price']"),
            (By.CSS_SELECTOR, "ins[class*='price']"),
        ]
        
        price = None
        for by, selector in price_selectors:
            try:
                price_elem = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((by, selector))
                )
                price_text = price_elem.text.strip()
                # Извлекаем число (убираем все нецифровые символы, включая nbsp)
                price_num = re.sub(r'[^\d]', '', price_text)
                if price_num:
                    price = int(price_num)
                    print(f"  ✓ Цена найдена: {price} ₽ (селектор: {selector})")
                    break
            except:
                continue
        
        if not price:
            print(f"  ⚠ Цена не найдена - возможно товар недоступен")
            # Закрываем вкладку перед возвратом
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            return 0
        
        # Закрываем вкладку после успешного парсинга
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        
        return price
    
    except InvalidSessionIdException:
        print(f"  ✗ Сессия Chrome разорвана")
        raise  # Пробрасываем дальше для переподключения
    except Exception as e:
        print(f"  ✗ Ошибка: {e}")
        # Закрываем вкладку в случае ошибки (если она открыта)
        try:
            if len(driver.window_handles) > 1:
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
        except:
            pass
        return 0


def save_results_to_excel(results, output_file):
    """Сохраняет результаты в Excel файл"""
    try:
        from openpyxl import Workbook
        
        # Создаём новый Excel файл
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "Цены"
        
        # Заголовки
        ws_out.append(["ссылка на товар", "артикул", "цена", "цена с картой"])
        
        # Данные
        for result in results:
            ws_out.append([
                result['url'],
                result['article'],
                result['price'],
                result.get('price_with_card', 0)  # 0 если цена с картой не найдена
            ])
        
        # Автофильтр
        ws_out.auto_filter.ref = ws_out.dimensions
        
        # Сохраняем файл
        wb_out.save(output_file)
        wb_out.close()
        
        return True
    except Exception as e:
        print(f"\n[!] ОШИБКА при сохранении: {e}")
        return False


def get_last_processed_row_count(sheet_url, sheet_name="Лист1"):
    """
    Получает количество уже обработанных строк в Google Таблице
    Возвращает количество строк (без учета заголовка), или 0 если таблица пустая
    Выдает ошибку если таблица недоступна
    """
    if not GOOGLE_SHEETS_ENABLED:
        raise Exception("Google Sheets отключен (GOOGLE_SHEETS_ENABLED = False)")
    
    if not sheet_url:
        raise Exception("Google Sheet URL не указан (GOOGLE_SHEET_URL пустой)")
    
    try:
        import gspread
    except ImportError:
        raise Exception("Библиотека gspread не установлена. Установите: pip install gspread google-auth")
    
    try:
        # Извлекаем ID таблицы из URL
        if '/d/' in sheet_url:
            sheet_id = sheet_url.split('/d/')[1].split('/')[0]
        else:
            raise Exception("Неверный формат ссылки на Google Sheet")
        
        # Подключаемся к Google Sheets
        service_account_file = os.path.join(PROJECT_ROOT, GOOGLE_SERVICE_ACCOUNT_FILE)
        
        if os.path.exists(service_account_file):
            gc = gspread.service_account(filename=service_account_file)
        else:
            raise Exception(f"Service Account файл не найден: {service_account_file}")
        
        spreadsheet = gc.open_by_key(sheet_id)
        
        # Получаем лист
        try:
            worksheet = spreadsheet.worksheet(sheet_name)
        except gspread.exceptions.WorksheetNotFound:
            raise Exception(f"Лист '{sheet_name}' не найден в Google Таблице")
        
        # Получаем все значения
        all_values = worksheet.get_all_values()
        
        # Если таблица пустая или только заголовок - возвращаем 0
        if len(all_values) <= 1:
            if len(all_values) == 0:
                raise Exception(f"Google Таблица пустая. Сначала создайте заголовки.")
            # Только заголовок
            return 0
        
        # Возвращаем количество строк без заголовка
        return len(all_values) - 1
        
    except Exception as e:
        raise Exception(f"Ошибка при чтении Google Таблицы: {e}")


def get_processed_articles_from_google_sheets(sheet_url, sheet_name="Лист1"):
    """
    Получает множество всех уже обработанных артикулов из Google Таблицы
    Возвращает set артикулов (строки), или пустой set если таблица пустая
    """
    if not GOOGLE_SHEETS_ENABLED:
        return set()
    
    if not sheet_url:
        return set()
    
    try:
        import gspread
    except ImportError:
        print(f"[ЛОГ] gspread не установлен, не могу проверить недостающие артикулы")
        return set()
    
    try:
        # Извлекаем ID таблицы из URL
        if '/d/' in sheet_url:
            sheet_id = sheet_url.split('/d/')[1].split('/')[0]
        else:
            return set()
        
        # Подключаемся к Google Sheets
        service_account_file = os.path.join(PROJECT_ROOT, GOOGLE_SERVICE_ACCOUNT_FILE)
        
        if not os.path.exists(service_account_file):
            return set()
        
        gc = gspread.service_account(filename=service_account_file)
        spreadsheet = gc.open_by_key(sheet_id)
        
        # Получаем лист
        try:
            worksheet = spreadsheet.worksheet(sheet_name)
        except gspread.exceptions.WorksheetNotFound:
            return set()
        
        # Получаем все значения (начиная со 2-й строки, первая - заголовок)
        all_values = worksheet.get_all_values()
        
        if len(all_values) <= 1:
            return set()
        
        # Артикулы находятся во 2-й колонке (индекс 1)
        processed_articles = set()
        for row in all_values[1:]:  # Пропускаем заголовок
            if len(row) > 1 and row[1]:  # Проверяем что есть артикул
                article = str(row[1]).strip()
                if article:
                    processed_articles.add(article)
        
        return processed_articles
        
    except Exception as e:
        print(f"[ЛОГ] Ошибка при чтении артикулов из Google Таблицы: {e}")
        return set()


def find_missing_articles(all_products, processed_articles):
    """
    Находит товары из all_products, которых нет в processed_articles
    Возвращает список словарей с ключами 'url' и 'article'
    """
    missing_products = []
    for product in all_products:
        article = str(product['article']).strip()
        if article and article not in processed_articles:
            missing_products.append(product)
    return missing_products


def save_results_to_google_sheets(results, sheet_url, sheet_name="Цены", append_only=False):
    """
    Сохраняет результаты в Google Таблицы автоматически через gspread с OAuth2
    
    Инструкция по настройке (один раз):
    1. Создайте Google Sheet и скопируйте ссылку
    2. Вставьте ссылку в GOOGLE_SHEET_URL
    3. Установите GOOGLE_SHEETS_ENABLED = True
    4. При первом запуске откроется браузер для авторизации (один раз)
    5. После авторизации создастся файл google_credentials.json
    6. В дальнейшем авторизация не потребуется
    """
    if not GOOGLE_SHEETS_ENABLED:
        print(f"  ⚠ Google Sheets отключен (GOOGLE_SHEETS_ENABLED = False)")
        return False
    
    if not sheet_url:
        print(f"  ⚠ Google Sheet URL не указан (GOOGLE_SHEET_URL пустой)")
        return False
    
    try:
        import gspread
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from google.auth.transport.requests import Request
        import pickle
        import json
    except ImportError:
        print(f"\n[!] Для записи в Google Таблицы установите:")
        print(f"    pip install gspread google-auth google-auth-oauthlib google-auth-httplib2")
        return False
    
    try:
        # Извлекаем ID таблицы из URL
        # Формат: https://docs.google.com/spreadsheets/d/SHEET_ID/edit
        if '/d/' in sheet_url:
            sheet_id = sheet_url.split('/d/')[1].split('/')[0]
        else:
            print(f"[!] Неверный формат ссылки на Google Sheet")
            print(f"    Пример: https://docs.google.com/spreadsheets/d/1ABC.../edit")
            return False
        
        # Подключаемся к Google Sheets
        service_account_file = os.path.join(PROJECT_ROOT, GOOGLE_SERVICE_ACCOUNT_FILE)
        
        # Пробуем использовать Service Account (самый простой способ)
        if os.path.exists(service_account_file):
            print(f"  📊 Подключение к Google Таблице через Service Account...")
            print(f"     Файл: {service_account_file}")
            gc = gspread.service_account(filename=service_account_file)
            print(f"     Подключение успешно!")
            spreadsheet = gc.open_by_key(sheet_id)
            print(f"     Таблица открыта: {spreadsheet.title}")
        else:
            # Используем OAuth2 (требует один раз авторизоваться через браузер)
            # OAuth2 авторизация (более безопасно)
            SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
            creds_file = os.path.join(PROJECT_ROOT, GOOGLE_CREDENTIALS_FILE)
            token_file = os.path.join(PROJECT_ROOT, 'google_token.pickle')
            
            creds = None
            
            # Пробуем загрузить сохраненные credentials
            if os.path.exists(token_file):
                with open(token_file, 'rb') as token:
                    creds = pickle.load(token)
            
            # Если нет валидных credentials, запрашиваем авторизацию
            if not creds or not creds.valid:
                if creds and creds.expired and creds.refresh_token:
                    creds.refresh(Request())
                else:
                    # Создаем файл credentials.json если его нет
                    if not os.path.exists(creds_file):
                        print(f"\n{'='*60}")
                        print("НАСТРОЙКА GOOGLE SHEETS API (OAuth2)")
                        print(f"{'='*60}")
                        print(f"\nДля автоматической записи в Google Таблицы нужна авторизация.")
                        print(f"\nИнструкция:")
                        print(f"1. Перейдите: https://console.cloud.google.com/")
                        print(f"2. Создайте проект (или выберите существующий)")
                        print(f"3. Включите Google Sheets API")
                        print(f"4. Создайте OAuth 2.0 Client ID (Desktop app)")
                        print(f"5. Скачайте credentials.json и сохраните как '{GOOGLE_CREDENTIALS_FILE}' в корне проекта")
                        print(f"\nИли используйте упрощенный способ:")
                        print(f"   - Установите GOOGLE_USE_PUBLIC_ACCESS = True")
                        print(f"   - Создайте публичную Google Sheet с правами редактирования")
                        print(f"\nПропускаю запись в Google Sheets...")
                        return False
                    
                    flow = InstalledAppFlow.from_client_secrets_file(creds_file, SCOPES)
                    creds = flow.run_local_server(port=0)
                
                # Сохраняем credentials для следующего раза
                with open(token_file, 'wb') as token:
                    pickle.dump(creds, token)
            
            gc = gspread.authorize(creds)
            spreadsheet = gc.open_by_key(sheet_id)
        
        # Получаем или создаем лист
        try:
            worksheet = spreadsheet.worksheet(sheet_name)
        except gspread.exceptions.WorksheetNotFound:
            worksheet = spreadsheet.add_worksheet(title=sheet_name, rows=1000, cols=10)
        
        # Записываем заголовки если их нет
        if len(worksheet.get_all_values()) == 0:
            worksheet.append_row(["ссылка на товар", "артикул", "цена", "цена с картой"])
        
        # Если append_only=False, очищаем лист (кроме заголовков) и перезаписываем все
        if not append_only:
            if len(worksheet.get_all_values()) > 1:
                worksheet.delete_rows(2, len(worksheet.get_all_values()))
        
        # Записываем данные (добавляем в конец если append_only=True, иначе перезаписываем)
        print(f"\n📊 Запись в Google Таблицы...")
        batch_size = 100  # Google Sheets API ограничение
        for i in range(0, len(results), batch_size):
            batch = results[i:i+batch_size]
            rows = []
            for result in batch:
                rows.append([
                    result['url'],
                    result['article'],
                    result['price'],
                    result.get('price_with_card', 0)
                ])
            worksheet.append_rows(rows)
            print(f"  Записано: {min(i+batch_size, len(results))}/{len(results)}")
        
        print(f"✓ Данные успешно загружены в Google Таблицы")
        print(f"  Ссылка: {sheet_url}")
        return True
        
    except Exception as e:
        print(f"\n[!] ОШИБКА при сохранении в Google Таблицы: {e}")
        import traceback
        traceback.print_exc()
        return False


def save_results_to_csv_for_google_sheets(results, output_file):
    """
    Сохраняет результаты в CSV файл для последующего импорта в Google Таблицы
    Это самый простой способ без API ключа
    """
    try:
        import csv
        
        csv_file = output_file.replace('.xlsx', '_for_google_sheets.csv')
        
        with open(csv_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            # Заголовки
            writer.writerow(["ссылка на товар", "артикул", "цена", "цена с картой"])
            # Данные
            for result in results:
                writer.writerow([
                    result['url'],
                    result['article'],
                    result['price'],
                    result.get('price_with_card', 0)
                ])
        
        print(f"\n✓ CSV файл для Google Таблиц сохранен: {csv_file}")
        print(f"  Импортируйте его в Google Sheets: Файл → Импортировать → Загрузить")
        return True
    except Exception as e:
        print(f"\n[!] ОШИБКА при сохранении CSV: {e}")
        return False


def main():
    print("\n" + "="*80)
    print("ПАРСЕР ЦЕН WB - ПРОСТОЙ ПАРСЕР")
    print("="*80)
    
    # Проверяем путь к профилю (если не используем remote и не используем временный профиль)
    if not USE_REMOTE_CHROME and not USE_TEMP_PROFILE:
        if not os.path.exists(CHROME_USER_DATA_DIR):
            print(f"\n[!] ОШИБКА: Не найден Chrome User Data: {CHROME_USER_DATA_DIR}")
            return
        
        profile_path = os.path.join(CHROME_USER_DATA_DIR, CHROME_PROFILE_NAME)
        if not os.path.exists(profile_path):
            print(f"\n[!] ОШИБКА: Не найден профиль: {profile_path}")
            print(f"    Доступные профили:")
            for item in os.listdir(CHROME_USER_DATA_DIR):
                if item.startswith('Profile') or item == 'Default':
                    print(f"      - {item}")
            return
    
    print(f"\n✓ Конфигурация проверена")
    
    # Загружаем Excel с артикулами и ссылками
    print(f"\n[1/3] Загрузка данных из {ARTICLES_EXCEL_FILE}...")
    try:
        wb = load_workbook(ARTICLES_EXCEL_FILE)
    except Exception as e:
        print(f"\n[!] ОШИБКА открытия Excel: {e}")
        print(f"    Убедись что файл '{ARTICLES_EXCEL_FILE}' закрыт!")
        return
    
    # Определяем правильный лист
    sheet_name = None
    for possible_name in POSSIBLE_SHEET_NAMES:
        if possible_name in wb.sheetnames:
            sheet_name = possible_name
            break
    
    if not sheet_name:
        # Используем первый лист если ничего не найдено
        if wb.sheetnames:
            sheet_name = wb.sheetnames[0]
            print(f"[ЛОГ] Используется первый доступный лист: '{sheet_name}'")
        else:
            print(f"\n[!] ОШИБКА: В файле нет листов!")
            wb.close()
            return
    else:
        print(f"[ЛОГ] Используется лист: '{sheet_name}'")
    
    ws_in = wb[sheet_name]
    
    # Загружаем ссылки и артикулы из Articles.xlsx
    # Формат: колонка A - ссылка, колонка B - артикул
    # Начинаем со 2-й строки (первая может быть заголовком)
    products = []
    all_products = []  # Сохраняем все товары для проверки недостающих артикулов
    start_row = 1
    
    # Проверяем первую строку - если это заголовки, начинаем со 2-й
    first_row = list(ws_in.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    if first_row[0] and isinstance(first_row[0], str):
        first_cell_lower = str(first_row[0]).lower()
        if any(keyword in first_cell_lower for keyword in ['ссылка', 'link', 'url', 'артикул', 'article']):
            start_row = 2
            print(f"[ЛОГ] Обнаружены заголовки, пропускаю первую строку")
    
    # Читаем все строки до конца файла (включая пустые, чтобы не пропустить данные)
    for row_num in range(start_row, ws_in.max_row + 1):
        row = list(ws_in.iter_rows(min_row=row_num, max_row=row_num, max_col=2, values_only=True))[0]
        url = str(row[0]).strip() if row[0] else ""
        article = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        
        # Пропускаем полностью пустые строки
        if not url and not article:
            continue
        
        # Если есть артикул во втором столбце - используем его
        if article:
            # Если есть ссылка - используем её, иначе генерируем
            if url and "wildberries.ru" in url:
                product = {'url': url, 'article': article}
                products.append(product)
                all_products.append(product)
            else:
                # Генерируем ссылку из артикула
                product = {'url': f"https://www.wildberries.ru/catalog/{article}/detail.aspx", 'article': article}
                products.append(product)
                all_products.append(product)
        # Если артикула нет, но есть ссылка - извлекаем артикул из ссылки
        elif url and "wildberries.ru" in url:
            import re
            match = re.search(r'/catalog/(\d+)/', url)
            if match:
                article = match.group(1)
                product = {'url': url, 'article': article}
                products.append(product)
                all_products.append(product)
    
    print(f"    ✓ Найдено товаров: {len(products)}")
    
    if len(products) == 0:
        print("\n[!] Нет товаров для обработки!")
        print(f"    Проверьте файл {ARTICLES_EXCEL_FILE}, лист '{sheet_name}'")
        print(f"    Должны быть ссылки в колонке A и артикулы в колонке B")
        wb.close()
        return
    
    # ТЕСТОВЫЙ РЕЖИМ: ограничиваем количество товаров
    if TEST_MODE:
        products = products[:TEST_PRODUCTS_COUNT]
        print(f"⚠️  ТЕСТОВЫЙ РЕЖИМ: обработка первых {len(products)} товаров")
    
    # Проверяем Google Таблицу для продолжения с места остановки
    skip_count = 0
    total_products = len(products)
    if GOOGLE_SHEETS_ENABLED and GOOGLE_SHEET_URL:
        try:
            print(f"\n[1.5/3] Проверка Google Таблицы для продолжения...")
            last_row_count = get_last_processed_row_count(GOOGLE_SHEET_URL, GOOGLE_SHEET_NAME)
            skip_count = last_row_count
            if skip_count > 0:
                print(f"  ✓ Найдено уже обработанных товаров: {skip_count}")
                if skip_count >= total_products:
                    print(f"\n⚠️  ВСЕ ТОВАРЫ УЖЕ ОБРАБОТАНЫ!")
                    print(f"    В Google Таблице записано {skip_count} товаров")
                    print(f"    Всего товаров в файле: {total_products}")
                    print(f"    Парсинг не требуется.")
                    wb.close()
                    return
                # Пропускаем уже обработанные товары и продолжаем с места остановки
                products = products[skip_count:]
                print(f"  → Продолжаю с товара #{skip_count + 1} (пропущено {skip_count} товаров)")
                print(f"  → Осталось обработать: {len(products)} товаров из {total_products}")
            else:
                print(f"  ✓ Таблица пустая, начинаю с начала")
                print(f"  → Буду обрабатывать все {total_products} товаров")
        except Exception as e:
            print(f"\n[!] КРИТИЧЕСКАЯ ОШИБКА при проверке Google Таблицы:")
            print(f"    {e}")
            print(f"\n    Парсинг остановлен. Исправьте проблему и запустите снова.")
            wb.close()
            return
    
    # Запускаем Chrome
    print(f"\n[2/3] Запуск Chrome...")
    
    driver = None
    results = []  # Инициализируем результаты вне try, чтобы сохранить в finally
    last_saved_count = 0  # Счетчик сохраненных товаров для финального сохранения
    # Сохраняем все товары для проверки недостающих артикулов (доступно в finally)
    all_products_for_check = all_products.copy() if 'all_products' in locals() else []
    try:
        driver = setup_browser_driver()
        
        if not driver:
            print("\n[!] Не удалось запустить Chrome!")
            if USE_REMOTE_CHROME:
                print(f"\n💡 Убедись что Chrome запущен через START_CHROME_DEBUG.bat")
            wb.close()
            return
        
        print("    ✓ Chrome запущен")
        
        # Пауза для ручной авторизации (только в видимом режиме)
        if WAIT_FOR_MANUAL_LOGIN and not HEADLESS_MODE:
            print(f"\n{'='*80}")
            print("⏸  ПАУЗА ДЛЯ АВТОРИЗАЦИИ")
            print(f"{'='*80}")
            print(f"\n📋 ИНСТРУКЦИЯ:")
            print(f"   1. Открываю сайт WB в браузере...")
            print(f"   2. Авторизуйтесь в своем аккаунте")
            print(f"   3. Установите правильный адрес доставки")
            print(f"   4. После этого вернитесь сюда и нажмите ENTER")
            print(f"\n⏱  Таймаут: {MANUAL_LOGIN_TIMEOUT} секунд")
            print(f"   (или нажмите ENTER когда будете готовы)")
            print(f"\n{'='*80}\n")
            
            # Открываем WB для авторизации
            try:
                print(f"[ЛОГ] Открываю https://www.wildberries.ru/ для авторизации...")
                driver.get("https://www.wildberries.ru/")
                time.sleep(2)  # Даем время на загрузку
                print(f"[ЛОГ] ✓ Страница WB открыта")
            except Exception as e:
                print(f"[ЛОГ] ⚠ Ошибка открытия WB: {e}")
                print(f"[ЛОГ] Откройте WB вручную в браузере")
            
            # Ждем пока пользователь авторизуется
            try:
                input(f"\n⏸ Нажмите ENTER когда авторизуетесь и установите адрес доставки...")
            except KeyboardInterrupt:
                print(f"\n[!] Прервано пользователем")
                driver.quit()
                return
        elif WAIT_FOR_MANUAL_LOGIN and HEADLESS_MODE:
            print(f"\n⚠️  ВНИМАНИЕ: Headless режим активен!")
            print(f"   Авторизация через браузер невозможна (браузер не виден).")
            print(f"   Убедитесь, что профиль уже авторизован или используйте видимый режим для первой авторизации.\n")
            # В headless режиме просто проверяем, что профиль работает
            try:
                print(f"[ЛОГ] Проверяю доступность WB...")
                driver.get("https://www.wildberries.ru/")
                time.sleep(2)
                print(f"[ЛОГ] ✓ WB доступен, продолжаю парсинг...")
            except Exception as e:
                print(f"\n[!] Ошибка при проверке WB: {e}")
                print(f"    Продолжаю парсинг...")
        
        # Парсим товары (параллельно)
        print(f"\n[3/3] Парсинг цен...")
        print("="*80)
        
        # Используем параллельную обработку
        parsed_data = process_products_parallel(driver, products)
        # Объединяем результаты (на случай если уже были частичные результаты)
        if parsed_data:
            results, last_saved_count = parsed_data
            print(f"\n✓ Парсинг завершен: собрано {len(results)} товаров")
        else:
            print(f"\n⚠ Парсинг не вернул результатов (возможно произошла ошибка)")
            results = []
            last_saved_count = 0
        
    except Exception as e:
        print(f"\n[!] КРИТИЧЕСКАЯ ОШИБКА: {e}")
        print(f"    Сохраню уже собранные результаты: {len(results)} товаров")
        import traceback
        traceback.print_exc()
    
    finally:
        # Сохраняем результаты в Google Таблицы (всегда, даже при ошибках)
        print(f"\n{'='*80}")
        print("ФИНАЛЬНОЕ СОХРАНЕНИЕ РЕЗУЛЬТАТОВ")
        print(f"{'='*80}")
        
        if len(results) > 0:
            # Сохраняем в Google Таблицы (единственный способ сохранения)
            if GOOGLE_SHEETS_ENABLED and GOOGLE_SHEET_URL:
                # Сохраняем только несохраненные товары (если есть)
                unsaved_results = results[last_saved_count:]
                if len(unsaved_results) > 0:
                    print(f"\n📊 Финальная запись в Google Таблицы ({len(unsaved_results)} несохраненных товаров)...")
                    if save_results_to_google_sheets(unsaved_results, GOOGLE_SHEET_URL, GOOGLE_SHEET_NAME, append_only=True):
                        print(f"✓ Данные загружены в Google Таблицы")
                        print(f"  Ссылка: {GOOGLE_SHEET_URL}")
                    else:
                        print(f"⚠ Не удалось сохранить в Google Таблицы")
                else:
                    print(f"\n✓ Все товары уже сохранены в Google Таблицы")
            else:
                print(f"\n⚠ Google Таблицы не настроены!")
                print(f"   Установите GOOGLE_SHEETS_ENABLED = True и укажите GOOGLE_SHEET_URL")
        else:
            print(f"\n⚠ Нет данных для сохранения (results пустой)")
            print(f"   Возможные причины:")
            print(f"   - Парсинг не начался из-за ошибки")
            print(f"   - Браузер закрылся до начала парсинга")
            print(f"   - Ошибка в process_products_parallel")
        
        # Проверяем недостающие артикулы и допарсиваем их
        if driver and GOOGLE_SHEETS_ENABLED and GOOGLE_SHEET_URL:
            try:
                print(f"\n{'='*80}")
                print("ПРОВЕРКА НЕДОСТАЮЩИХ АРТИКУЛОВ")
                print(f"{'='*80}")
                
                # Получаем все артикулы из Google Таблицы
                print(f"\n[1/3] Получаю список обработанных артикулов из Google Таблицы...")
                processed_articles = get_processed_articles_from_google_sheets(GOOGLE_SHEET_URL, GOOGLE_SHEET_NAME)
                print(f"  ✓ Найдено обработанных артикулов в Google Таблице: {len(processed_articles)}")
                
                # Используем сохраненный список всех товаров
                if all_products_for_check:
                    print(f"\n[2/3] Использую загруженные товары из файла: {len(all_products_for_check)}")
                    
                    # Находим недостающие артикулы
                    print(f"\n[3/3] Ищу недостающие артикулы...")
                    missing_products = find_missing_articles(all_products_for_check, processed_articles)
                    
                    if missing_products:
                        print(f"  ⚠ Найдено недостающих артикулов: {len(missing_products)}")
                        print(f"  → Допарсиваю недостающие товары...")
                        
                        # Допарсиваем недостающие товары
                        missing_results, _ = process_products_parallel(driver, missing_products)
                        
                        if missing_results:
                            print(f"\n📊 Сохранение допарсенных товаров ({len(missing_results)} товаров)...")
                            if save_results_to_google_sheets(missing_results, GOOGLE_SHEET_URL, GOOGLE_SHEET_NAME, append_only=True):
                                print(f"✓ Допарсенные товары сохранены в Google Таблицы")
                            else:
                                print(f"⚠ Не удалось сохранить допарсенные товары")
                    else:
                        print(f"  ✓ Все артикулы из файла уже обработаны!")
                else:
                    print(f"  ⚠ Не удалось найти список всех товаров для проверки")
                    
            except Exception as e:
                print(f"\n⚠ Ошибка при проверке недостающих артикулов: {e}")
                import traceback
                traceback.print_exc()
        
        if driver:
            print(f"\n[Закрываю Chrome через 5 секунд...]")
            time.sleep(5)
            driver.quit()
        
        if 'wb' in locals():
            wb.close()
    
    print(f"\n{'='*80}")
    print("ЗАВЕРШЕНО")
    print(f"{'='*80}\n")


def test_google_sheets():
    """Тестовая функция для проверки записи в Google Таблицы"""
    if not GOOGLE_SHEETS_ENABLED or not GOOGLE_SHEET_URL:
        print("\n[!] Google Sheets не настроен. Установите GOOGLE_SHEETS_ENABLED = True и GOOGLE_SHEET_URL")
        return
    
    print("\n" + "="*80)
    print("ТЕСТ ЗАПИСИ В GOOGLE ТАБЛИЦЫ")
    print("="*80)
    
    # Создаем тестовые данные
    test_results = [{
        'url': 'https://test.com',
        'article': 'ПРИВЕТ',
        'price': 12345,
        'price_with_card': 0
    }]
    
    print(f"\n📊 Пробую записать 'привет' в Google Таблицы...")
    print(f"   URL: {GOOGLE_SHEET_URL}")
    print(f"   Лист: {GOOGLE_SHEET_NAME}")
    
    if save_results_to_google_sheets(test_results, GOOGLE_SHEET_URL, GOOGLE_SHEET_NAME):
        print(f"\n✓ ТЕСТ УСПЕШЕН! Проверьте Google Таблицу - там должно быть 'привет'")
    else:
        print(f"\n✗ ТЕСТ НЕ УДАЛСЯ. Проверьте настройки и файл google-credentials.json")


if __name__ == "__main__":
    import sys
    # Если запущен с аргументом --test-google, запускаем тест
    if len(sys.argv) > 1 and sys.argv[1] == "--test-google":
        test_google_sheets()
    else:
        main()
