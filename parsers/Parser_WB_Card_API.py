# -*- coding: utf-8 -*-
"""
ПАРСЕР ЦЕН WILDBERRIES - ЧЕРЕЗ CARD API
Получает цены напрямую с карточек товаров через внутренний JSON API
"""

import os
import requests
import json
import time
from openpyxl import load_workbook
from datetime import datetime

# Конфигурация
# Пути относительно корня проекта
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(PROJECT_ROOT, "data")

EXCEL_FILE = os.path.join(DATA_DIR, "Парсер цен.xlsx")
SHEET_INPUT = "Данные для парсера ВБ"
SHEET_OUTPUT = "Результаты парсинга ВБ"

# WB Basket API - более надёжный способ получить данные товаров
def get_basket_number(nm_id):
    """Определяет номер корзины для артикула"""
    vol = nm_id // 100000
    part = nm_id // 1000
    return vol, part

def get_wb_card_data(nm_ids, spp=30):
    """
    Получает данные через Basket API (по одному товару)
    
    URL формат: https://basket-XX.wbbasket.ru/vol{vol}/part{part}/{nmID}/info/ru/card.json
    """
    
    results = {}
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept': 'application/json',
        'Accept-Language': 'ru-RU,ru;q=0.9'
    }
    
    for nm_id in nm_ids:
        try:
            vol, part = get_basket_number(int(nm_id))
            
            # Определяем номер корзины (01-20)
            basket_num = str((vol % 20) + 1).zfill(2)
            
            url = f"https://basket-{basket_num}.wbbasket.ru/vol{vol}/part{part}/{nm_id}/info/ru/card.json"
            
            response = requests.get(url, headers=headers, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                parsed = parse_basket_response(data, nm_id)
                if parsed:
                    results[str(nm_id)] = parsed
            else:
                print(f"  [{nm_id}] Ошибка {response.status_code}")
            
            time.sleep(0.1)  # Пауза между запросами
        
        except Exception as e:
            print(f"  [{nm_id}] Ошибка: {e}")
    
    return results


def parse_basket_response(data, nm_id):
    """
    Парсит ответ от Basket API
    
    Структура:
    {
        "nm_id": nmID,
        "name": "название",
        "brand": "бренд",
        "brand_id": id,
        "site_brand_id": id,
        "supplier_id": id,
        "sale": процент_скидки,
        "priceU": цена * 100,
        "salePriceU": финальная_цена * 100,
        "extended": {
            "basicSale": обычная_скидка,
            "clientSale": скидка_для_клиента,
            "basicPriceU": базовая_цена * 100
        },
        "sizes": [...]
    }
    """
    
    try:
        # Базовые данные
        name = data.get('name', '')
        
        # Цены (в копейках, делим на 100)
        price_u = data.get('priceU', 0) / 100
        sale_price_u = data.get('salePriceU', 0) / 100
        
        # Расширенные данные
        extended = data.get('extended', {})
        basic_price_u = extended.get('basicPriceU', price_u * 100) / 100
        
        # Скидки
        sale = data.get('sale', 0)  # Общая скидка %
        basic_sale = extended.get('basicSale', 0)  # Обычная скидка %
        client_sale = extended.get('clientSale', 0)  # Клубная скидка %
        
        # Размеры
        sizes = data.get('sizes', [])
        size_name = sizes[0].get('origName', '') if sizes else ''
        
        # Остатки
        total_stock = 0
        if sizes:
            for size in sizes:
                stocks = size.get('stocks', [])
                for stock in stocks:
                    total_stock += stock.get('qty', 0)
        
        # Рассчитываем цены
        # Базовая цена
        base_price = basic_price_u if basic_price_u > 0 else price_u
        
        # Цена до СПП (с обычной скидкой)
        if basic_sale > 0:
            price_before_spp = round(base_price * (1 - basic_sale / 100), 2)
        else:
            price_before_spp = price_u
        
        # Финальная цена (после СПП)
        final_price = sale_price_u
        
        # СПП %
        spp_percent = client_sale
        
        return {
            'name': name,
            'techSizeName': size_name,
            'price': base_price,
            'discountedPrice': price_before_spp,
            'clubDiscountedPrice': final_price,
            'discount': basic_sale,
            'clubDiscount': spp_percent,
            'stockCount': total_stock
        }
    
    except Exception as e:
        print(f"  [!] Ошибка парсинга {nm_id}: {e}")
        return None


def main():
    print("\n" + "="*80)
    print("ПАРСЕР ЦЕН WB - ЧЕРЕЗ CARD API")
    print("="*80)
    
    # Загружаем Excel
    wb = load_workbook(EXCEL_FILE)
    ws_in = wb[SHEET_INPUT]
    ws_out = wb[SHEET_OUTPUT]
    
    # Загружаем артикулы
    articles = []
    for row in ws_in.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            articles.append(str(row[0]).strip())
    
    print(f"\n[1/3] Найдено артикулов: {len(articles)}")
    
    # Парсим партиями по 100
    batch_size = 100
    all_results = {}
    
    print(f"\n[2/3] Парсинг через Card API...")
    
    for i in range(0, len(articles), batch_size):
        batch = articles[i:i + batch_size]
        batch_num = i // batch_size + 1
        total_batches = (len(articles) + batch_size - 1) // batch_size
        
        print(f"  Батч {batch_num}/{total_batches}: {len(batch)} артикулов...")
        
        batch_results = get_wb_card_data(batch)
        all_results.update(batch_results)
        
        print(f"    Получено: {len(batch_results)} товаров")
        
        time.sleep(0.5)  # Пауза между запросами
    
    # Сохраняем результаты
    print(f"\n[3/3] Сохранение в Excel...")
    
    # Очищаем старые данные
    if ws_out.max_row >= 1:
        ws_out.delete_rows(1, ws_out.max_row)
    
    # Заголовки
    ws_out.append([
        "Дата",
        "nmID",
        "Название (name)",
        "Размер (techSizeName)",
        "price",
        "discountedPrice",
        "clubDiscountedPrice",
        "discount %",
        "clubDiscount %",
        "stockCount"
    ])
    
    # Данные
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    success = 0
    failed = 0
    
    for article in articles:
        data = all_results.get(article, {})
        
        if data:
            ws_out.append([
                timestamp,
                article,
                data.get('name', ''),
                data.get('techSizeName', ''),
                data.get('price', None),
                data.get('discountedPrice', None),
                data.get('clubDiscountedPrice', None),
                data.get('discount', None),
                data.get('clubDiscount', None),
                data.get('stockCount', 0)
            ])
            success += 1
        else:
            ws_out.append([
                timestamp,
                article,
                'Не найдено',
                '',
                None,
                None,
                None,
                None,
                None,
                0
            ])
            failed += 1
    
    # Автофильтр
    ws_out.auto_filter.ref = ws_out.dimensions
    
    wb.save(EXCEL_FILE)
    
    print(f"\n{'='*80}")
    print(f"ГОТОВО!")
    print(f"{'='*80}")
    print(f"Найдено: {success}")
    print(f"Не найдено: {failed}")
    print(f"{'='*80}\n")


if __name__ == "__main__":
    # Тест на 3 артикулах
    test_ids = [583658258, 214690928, 242175842]
    
    print("ТЕСТ Card API на 3 артикулах:\n")
    
    results = get_wb_card_data(test_ids)
    
    for nm_id, data in results.items():
        print(f"\n{'='*60}")
        print(f"nmID: {nm_id}")
        print(f"Название: {data.get('name', '')}")
        print(f"Размер: {data.get('techSizeName', '')}")
        print(f"Базовая цена (price): {data.get('price', 0)} ₽")
        print(f"ДО СПП (discountedPrice): {data.get('discountedPrice', 0)} ₽")
        print(f"ПОСЛЕ СПП (clubDiscountedPrice): {data.get('clubDiscountedPrice', 0)} ₽")
        print(f"Скидка %: {data.get('discount', 0)}")
        print(f"СПП %: {data.get('clubDiscount', 0)}")
        print(f"Остаток: {data.get('stockCount', 0)} шт")
    
    print(f"\n{'='*60}\n")
    print("Тест завершён! Запустите main() для полного парсинга.")

