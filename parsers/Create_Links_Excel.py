# -*- coding: utf-8 -*-
"""
СОЗДАНИЕ EXCEL ФАЙЛА СО ССЫЛКАМИ НА ТОВАРЫ
Читает артикулы из Articles.xlsx и создаёт файл links_to_products.xlsx со ссылками
Файл используется парсером Parser_WB_Search.py для парсинга цен
"""

import os
import sys
from openpyxl import load_workbook, Workbook

# Настройка кодировки консоли для Windows
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass

# Пути относительно корня проекта
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(PROJECT_ROOT, "data")

# Конфигурация
EXCEL_FILE = os.path.join(PROJECT_ROOT, "Articles.xlsx")
# Возможные имена листов (будет использован первый найденный)
POSSIBLE_SHEET_NAMES = ["Данные для парсера ВБ", "WBarticules", "WB", "Артикулы", "Sheet1"]
OUTPUT_EXCEL_FILE = os.path.join(DATA_DIR, "links_to_products.xlsx")
SHEET_LINKS = "Ссылки на товары"
WB_URL_TEMPLATE = "https://www.wildberries.ru/catalog/{}/detail.aspx"


def main():
    print("\n" + "=" * 80)
    print("СОЗДАНИЕ EXCEL ФАЙЛА СО ССЫЛКАМИ НА ТОВАРЫ WB")
    print("=" * 80)
    
    # Создаем директорию data если её нет
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)
        print(f"[ЛОГ] Создана директория: {DATA_DIR}")

    # Загружаем исходный Excel с артикулами
    print(f"\n[1/3] Загрузка артикулов из {EXCEL_FILE}...")
    try:
        wb = load_workbook(EXCEL_FILE)
    except Exception as e:
        print(f"\n[!] ОШИБКА открытия Excel: {e}")
        print(f"    Убедись что файл '{EXCEL_FILE}' закрыт!")
        print(f"    Файл должен находиться в корне проекта")
        return

    # Определяем правильный лист
    sheet_name = None
    for possible_name in POSSIBLE_SHEET_NAMES:
        if possible_name in wb.sheetnames:
            sheet_name = possible_name
            break
    
    if not sheet_name:
        # Используем первый лист если ничего не найдено
        if wb.sheetnames:
            sheet_name = wb.sheetnames[0]
            print(f"[ЛОГ] Используется первый доступный лист: '{sheet_name}'")
        else:
            print(f"\n[!] ОШИБКА: В файле нет листов!")
            wb.close()
            return
    else:
        print(f"[ЛОГ] Используется лист: '{sheet_name}'")

    ws_in = wb[sheet_name]

    # Загружаем данные: проверяем оба столбца (ссылка и артикул)
    articles = []
    links = []
    
    for row in ws_in.iter_rows(min_row=1, max_col=2, values_only=True):
        url = str(row[0]).strip() if row[0] else ""
        article = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        
        # Если есть артикул во втором столбце - используем его
        if article and article.isdigit():
            articles.append(article)
            # Если есть ссылка - используем её, иначе генерируем
            if url and "wildberries.ru" in url:
                links.append(url)
            else:
                links.append(WB_URL_TEMPLATE.format(article))
        # Если артикула нет, но есть ссылка - извлекаем артикул из ссылки
        elif url and "wildberries.ru" in url:
            # Извлекаем артикул из URL: /catalog/12345678/detail.aspx
            import re
            match = re.search(r'/catalog/(\d+)/', url)
            if match:
                article = match.group(1)
                articles.append(article)
                links.append(url)

    print(f"    ✓ Найдено артикулов: {len(articles)}")

    if len(articles) == 0:
        print("\n[!] Нет артикулов для обработки!")
        print(f"    Проверьте файл {EXCEL_FILE}, лист '{sheet_name}'")
        print(f"    Артикулы должны быть в первом столбце, начиная со 2-й строки")
        wb.close()
        return

    wb.close()

    # Создаём новый Excel файл со ссылками
    print(f"\n[2/3] Создание файла со ссылками...")

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = SHEET_LINKS

    # Заголовки
    ws_out.append(["ссылка на товар", "артикул"])

    # Генерируем ссылки для каждого артикула
    print(f"    Генерация ссылок для {len(articles)} товаров...")
    for article in articles:
        product_url = WB_URL_TEMPLATE.format(article)
        ws_out.append([product_url, article])

    # Автофильтр
    ws_out.auto_filter.ref = ws_out.dimensions

    # Сохраняем файл
    print(f"\n[3/3] Сохранение файла...")
    wb_out.save(OUTPUT_EXCEL_FILE)
    wb_out.close()

    print(f"\n{'=' * 80}")
    print("ФАЙЛ СОЗДАН УСПЕШНО")
    print(f"{'=' * 80}")
    print(f"✓ Обработано артикулов: {len(articles)}")
    print(f"✓ Файл сохранён: {OUTPUT_EXCEL_FILE}")
    print(f"✓ Лист: {SHEET_LINKS}")
    print(f"\n📋 СЛЕДУЮЩИЙ ШАГ:")
    print(f"   Запустите парсер цен:")
    print(f"   python parsers/Parser_WB_Search.py")
    print(f"{'=' * 80}\n")


if __name__ == "__main__":
    main()