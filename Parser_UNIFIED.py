# -*- coding: utf-8 -*-
"""
ПАРСЕР ЦЕН WILDBERRIES
Парсинг цен до СПП и после СПП для нескольких магазинов WB
"""

import time
import json
import requests
from datetime import datetime
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# === КОНФИГУРАЦИЯ ===
EXCEL_FILE = "Парсер цен.xlsx"
SHEET_INPUT_WB = "Данные для парсера ВБ"
SHEET_OUTPUT_WB = "Парсер ВБ"
SHEET_SETTINGS = "Настройка"

WB_URL_TEMPLATE = "https://www.wildberries.ru/catalog/{}/detail.aspx"
WB_API_URL = "https://discounts-prices-api.wildberries.ru/api/v2/list/goods/filter"

PAGE_TIMEOUT_WB = 5
PAUSE_BETWEEN = 0.5

# === ФУНКЦИИ ДЛЯ РАБОТЫ С API ===

def load_api_keys(wb):
    """
    Загружает API ключи WB из листа Настройка
    Структура листа:
    D: API Key WB
    
    Возвращает список API ключей WB
    """
    try:
        ws_settings = wb[SHEET_SETTINGS]
        
        wb_api_keys = []      # Список API ключей WB
        
        # Читаем все строки начиная со 2-й (1-я строка - заголовки)
        for row in ws_settings.iter_rows(min_row=2, values_only=True):
            if len(row) >= 4:  # Проверяем что есть столбец D
                api_key_wb = str(row[3]).strip() if row[3] else ""
                
                # Добавляем WB ключ если он есть
                if api_key_wb:
                    wb_api_keys.append(api_key_wb)
        
        print(f"\n[API] Загружено WB кабинетов: {len(wb_api_keys)}")
        
        return wb_api_keys
        
    except Exception as e:
        print(f"[!] Ошибка загрузки API ключей: {e}")
        import traceback
        traceback.print_exc()
        return []

def get_wb_prices_api(articles, api_keys_list):
    """
    Получает цены до СПП для WB через API
    Обрабатывает несколько API ключей для разных кабинетов
    Возвращает словарь {артикул: цена_до_спп}
    """
    print("\n[API WB] Загрузка цен до СПП через API...")
    
    if not api_keys_list or len(api_keys_list) == 0:
        print("[!] API ключи WB не найдены!")
        return {}
    
    prices_before_spp = {}
    
    # Обрабатываем каждый API ключ (каждый кабинет)
    for idx, api_key in enumerate(api_keys_list, 1):
        print(f"\n[API WB] Кабинет {idx}/{len(api_keys_list)}...")
        
        try:
            headers = {
                "Authorization": api_key,
                "Content-Type": "application/json"
            }
            
            # WB API позволяет запрашивать до 1000 артикулов за раз
            batch_size = 1000
            
            for i in range(0, len(articles), batch_size):
                batch = articles[i:i + batch_size]
                
                # Конвертируем артикулы в числа
                nm_ids = [int(art) for art in batch if art.isdigit()]
                
                payload = {
                    "limit": 1000,
                    "offset": 0,
                    "nmList": nm_ids
                }
                
                response = requests.post(WB_API_URL, headers=headers, json=payload, timeout=30)
                
                if response.status_code == 200:
                    data = response.json()
                    
                    # Парсим ответ
                    if "data" in data and "listGoods" in data["data"]:
                        for item in data["data"]["listGoods"]:
                            nm_id = str(item.get("nmID", ""))
                            
                            # Берем discountedPrice из sizes
                            sizes = item.get("sizes", [])
                            if sizes and len(sizes) > 0:
                                # Берем первый размер
                                discounted_price = sizes[0].get("discountedPrice", 0)
                                if discounted_price:
                                    prices_before_spp[nm_id] = float(discounted_price)
                    
                    print(f"    Найдено {len([x for x in data.get('data', {}).get('listGoods', []) if x])} товаров в этом кабинете")
                else:
                    print(f"[!] Ошибка API WB (кабинет {idx}): {response.status_code}")
                    if response.status_code != 404:  # 404 = товары не найдены (нормально)
                        print(f"    Ответ: {response.text[:200]}")
                
                time.sleep(0.3)  # Пауза между запросами
        
        except Exception as e:
            print(f"[!] Ошибка при работе с API WB (кабинет {idx}): {e}")
    
    print(f"\n[API WB] Итого загружено {len(prices_before_spp)} цен из всех кабинетов")
    return prices_before_spp


# === ФУНКЦИИ ДЛЯ БРАУЗЕРА ===

def start_browser_wb(headless=False):
    """Запускает браузер для WB"""
    options = webdriver.ChromeOptions()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    
    # Отключаем картинки для скорости (только для headless)
    if headless:
        prefs = {"profile.managed_default_content_settings.images": 2}
        options.add_experimental_option("prefs", prefs)
    
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(5)
    return driver


# === ФУНКЦИИ ПАРСИНГА WB ===

def parse_price_wb(driver, url):
    """Парсит цены WB (работает с авторизованной версией страницы)"""
    try:
        driver.get(url)
        time.sleep(1)
        
        # ВАРИАНТ 1: Авторизованная страница с МОДАЛЬНЫМ ОКНОМ (2024)
        try:
            # ВАРИАНТ 1.1: Кнопка с цветом danger (красная)
            try:
                # Ищем кнопку с ценой кошелька (актуальная разметка WB - красная цена)
                wallet_button = WebDriverWait(driver, PAGE_TIMEOUT_WB).until(
                    EC.presence_of_element_located((
                        By.CSS_SELECTOR,
                        "h2.mo-typography_color_danger, span.priceBlockWalletPricePointer--WG1OK, span.priceBlockWalletPrice--RJGuT, span[class*='priceBlockWalletPrice']"
                    ))
                )
                
                # Кликаем на кнопку
                wallet_button.click()
                time.sleep(0.7)
                
                # Парсим цены из модального окна
                try:
                    price_spp_el = WebDriverWait(driver, 6).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "h2.mo-typography_color_primary"))
                    )
                    price_spp_text = price_spp_el.text.replace("\xa0", "").replace(" ", "").replace("₽", "").replace(",", ".")
                    price_spp = float(price_spp_text) if price_spp_text else None
                    
                    try:
                        price_wallet_el = driver.find_element(By.CSS_SELECTOR, "h2.mo-typography_color_danger")
                        price_wallet_text = price_wallet_el.text.replace("\xa0", "").replace(" ", "").replace("₽", "").replace(",", ".")
                        price_wallet = float(price_wallet_text) if price_wallet_text else None
                    except:
                        price_wallet = None
                    
                    if price_spp:
                        return price_spp, price_wallet
                except:
                    pass
            except:
                pass
            
            # ВАРИАНТ 1.2: Кнопка с цветом accent (синяя/фиолетовая) - если первый вариант не сработал
            try:
                # Ищем кнопку с ценой кошелька (актуальная разметка WB - цвет accent)
                wallet_button = WebDriverWait(driver, PAGE_TIMEOUT_WB).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "h2.mo-typography_color_accent"))
                )
                
                # Кликаем на кнопку
                wallet_button.click()
                time.sleep(0.7)
                
                # Парсим цены из модального окна
                try:
                    price_spp_el = WebDriverWait(driver, 6).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "h2.mo-typography_color_primary"))
                    )
                    price_spp_text = price_spp_el.text.replace("\xa0", "").replace(" ", "").replace("₽", "").replace(",", ".")
                    price_spp = float(price_spp_text) if price_spp_text else None
                    
                    try:
                        price_wallet_el = driver.find_element(By.CSS_SELECTOR, "h2.mo-typography_color_accent")
                        price_wallet_text = price_wallet_el.text.replace("\xa0", "").replace(" ", "").replace("₽", "").replace(",", ".")
                        price_wallet = float(price_wallet_text) if price_wallet_text else None
                    except:
                        price_wallet = None
                    
                    if price_spp:
                        return price_spp, price_wallet
                except:
                    pass
            except:
                pass
            
            # ВАРИАНТ 1.3: Пробуем старые селекторы (fallback)
            try:
                price_spp_el = WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "h2.finalPrice--Q5ltH, h2[class*='finalPrice']"))
                )
                price_spp_text = price_spp_el.text.replace("\xa0", "").replace(" ", "").replace("₽", "").replace(",", ".")
                price_spp = float(price_spp_text) if price_spp_text else None
                
                price_wallet_el = driver.find_element(By.CSS_SELECTOR, "h2.walletPrice--RPzXi, h2[class*='walletPrice']")
                price_wallet_text = price_wallet_el.text.replace("\xa0", "").replace(" ", "").replace("₽", "").replace(",", ".")
                price_wallet = float(price_wallet_text) if price_wallet_text else None
                
                return price_spp, price_wallet
            except:
                return None, None
        except:
            # ВАРИАНТ 2: Неавторизованная страница
            try:
                price_spp_el = WebDriverWait(driver, PAGE_TIMEOUT_WB).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "ins[class*='priceBlockFinalPrice']"))
                )
                
                price_spp_text = price_spp_el.text.replace("\xa0", "").replace(" ", "").replace("₽", "").replace(",", ".")
                price_spp = float(price_spp_text) if price_spp_text else None
                
                try:
                    price_wallet_el = driver.find_element(By.CSS_SELECTOR, "span[class*='priceBlockWalletPrice']")
                    price_wallet_text = price_wallet_el.text.replace("\xa0", "").replace(" ", "").replace("₽", "").replace(",", ".")
                    price_wallet = float(price_wallet_text) if price_wallet_text else None
                except:
                    price_wallet = None
                
                return price_spp, price_wallet
            except:
                return None, None
    except:
        return None, None


# === ОСНОВНЫЕ ФУНКЦИИ ПАРСИНГА ===

def parse_wb_with_auth(wb, api_keys):
    """Парсинг WB с авторизацией"""
    print("\n" + "="*70)
    print("ПАРСИНГ WB С АВТОРИЗАЦИЕЙ")
    print("="*70)
    
    # Загрузка артикулов
    ws_in = wb[SHEET_INPUT_WB]
    ws_out = wb[SHEET_OUTPUT_WB]
    
    articles = []
    for row in ws_in.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            articles.append(str(row[0]).strip())
    
    total = len(articles)
    print(f"\n[1/5] Найдено артикулов: {total}")
    
    if total == 0:
        print("[!] Нет артикулов!")
        return
    
    # Получаем цены до СПП через API
    prices_before_spp_dict = get_wb_prices_api(articles, api_keys)
    
    # Запуск браузера
    print("\n[2/5] Запуск браузера...")
    driver = start_browser_wb()
    
    try:
        # Авторизация
        print("\n[3/5] Открытие wildberries.ru...")
        driver.get("https://www.wildberries.ru/")
        
        print("\n" + "="*70)
        print("АВТОРИЗАЦИЯ")
        print("="*70)
        print("\nВ ОТКРЫВШЕМСЯ БРАУЗЕРЕ:")
        print("  1. АВТОРИЗУЙТЕСЬ на WB")
        print("  2. ВЫБЕРИТЕ ваш ПВЗ")
        print("="*70)
        
        response = input("\nВведите 'да' когда авторизуетесь: ").strip().lower()
        if response != 'да':
            input("Нажмите Enter когда будете готовы: ")
        
        # Проверка
        print("\n[4/5] Проверка авторизации...")
        cookies_dict = {c['name']: c['value'] for c in driver.get_cookies()}
        
        if 'x_wbaas_token' in cookies_dict:
            print("    [OK] Токен найден - АВТОРИЗОВАН!")
        else:
            print("    [WARNING] Токен не найден, продолжаем как гость...")
        
        # ПАРСИНГ
        print(f"\n[5/5] Парсинг {total} артикулов...")
        print("="*70)
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        start_time = time.time()
        success = 0
        failed = 0
        
        for i, article in enumerate(articles, 1):
            url = WB_URL_TEMPLATE.format(article)
            
            print(f"\n[{i}/{total}] {article}", end=" ")
            
            price_spp, price_wallet = parse_price_wb(driver, url)
            price_before_spp = prices_before_spp_dict.get(article)
            
            # Расчет процентов
            percent_spp = None
            percent_wallet = None
            
            if price_before_spp and price_spp:
                percent_spp = (1 - (price_spp / price_before_spp)) * 100
            
            if price_spp and price_wallet:
                percent_wallet = (1 - (price_wallet / price_spp)) * 100
            
            if price_spp:
                spp_str = f"{percent_spp:.1f}%" if percent_spp is not None else "-"
                wallet_str = f"{percent_wallet:.1f}%" if percent_wallet is not None else "-"
                print(f"- До СПП:{price_before_spp} SPP:{price_spp} ({spp_str} скидка) Kosh:{price_wallet} ({wallet_str} доп.)")
                
                # Сохраняем: Дата | Артикул | Цена До СПП | % СПП | Цена с СПП | % кошелька | Цена с кошельком
                new_row = [timestamp, article, price_before_spp, percent_spp, price_spp, percent_wallet, price_wallet]
                ws_out.append(new_row)
                success += 1
            else:
                print("- ERROR")
                failed += 1
            
            if i < total:
                time.sleep(PAUSE_BETWEEN)
        
        # Итоги
        elapsed = time.time() - start_time
        print(f"\n{'='*70}")
        print("ГОТОВО!")
        print(f"{'='*70}")
        print(f"Всего: {total} | Успешно: {success} | Ошибок: {failed}")
        print(f"Время: {elapsed:.1f} сек ({elapsed/60:.1f} мин)")
        print(f"{'='*70}")
        
        wb.save(EXCEL_FILE)
        print(f"\n[SAVE] Результаты сохранены в '{EXCEL_FILE}'")
        
    finally:
        driver.quit()

def parse_wb_no_auth(wb, api_keys):
    """Парсинг WB без авторизации"""
    print("\n" + "="*70)
    print("ПАРСИНГ WB БЕЗ АВТОРИЗАЦИИ")
    print("="*70)
    
    # Загрузка артикулов
    ws_in = wb[SHEET_INPUT_WB]
    ws_out = wb[SHEET_OUTPUT_WB]
    
    articles = []
    for row in ws_in.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            articles.append(str(row[0]).strip())
    
    total = len(articles)
    print(f"\n[1/4] Найдено артикулов: {total}")
    
    if total == 0:
        print("[!] Нет артикулов!")
        return
    
    # Получаем цены до СПП через API
    prices_before_spp_dict = get_wb_prices_api(articles, api_keys)
    
    # Запуск браузера
    print("\n[2/4] Запуск браузера (headless режим)...")
    driver = start_browser_wb(headless=True)
    
    try:
        # ПАРСИНГ
        print(f"\n[3/4] Парсинг {total} артикулов...")
        print("="*70)
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        start_time = time.time()
        success = 0
        failed = 0
        
        for i, article in enumerate(articles, 1):
            url = WB_URL_TEMPLATE.format(article)
            
            print(f"[{i}/{total}] {article}", end=" ")
            
            price_spp, price_wallet = parse_price_wb(driver, url)
            price_before_spp = prices_before_spp_dict.get(article)
            
            # Расчет процентов
            percent_spp = None
            percent_wallet = None
            
            if price_before_spp and price_spp:
                percent_spp = (1 - (price_spp / price_before_spp)) * 100
            
            if price_spp and price_wallet:
                percent_wallet = (1 - (price_wallet / price_spp)) * 100
            
            if price_spp:
                print(f"- OK (До:{price_before_spp} СПП:{price_spp} Кош:{price_wallet})")
                
                new_row = [timestamp, article, price_before_spp, percent_spp, price_spp, percent_wallet, price_wallet]
                ws_out.append(new_row)
                success += 1
            else:
                print("- ERROR")
                failed += 1
            
            if i < total:
                time.sleep(PAUSE_BETWEEN)
            
            # Прогресс
            if i % 10 == 0:
                elapsed = time.time() - start_time
                speed = i / elapsed
                remaining = (total - i) / speed if speed > 0 else 0
                print(f"\n>>> {i}/{total} | OK:{success} ERR:{failed} | {speed:.1f} art/s | ~{remaining/60:.1f} min")
        
        # Итоги
        elapsed = time.time() - start_time
        print(f"\n{'='*70}")
        print("ГОТОВО!")
        print(f"{'='*70}")
        print(f"Всего: {total} | Успешно: {success} | Ошибок: {failed}")
        print(f"Время: {elapsed:.1f} сек ({elapsed/60:.1f} мин)")
        print(f"{'='*70}")
        
        wb.save(EXCEL_FILE)
        print(f"\n[SAVE] Результаты сохранены в '{EXCEL_FILE}'")
        
    finally:
        driver.quit()


# === ГЛАВНОЕ МЕНЮ ===

def show_menu():
    """Показывает меню выбора режима"""
    print("\n" + "="*70)
    print("ПАРСЕР ЦЕН WILDBERRIES")
    print("="*70)
    
    print("\nВыберите режим работы:")
    print("  1) С авторизацией (медленнее, но точнее - видны цены кошелька)")
    print("  2) Без авторизации (быстрее)")
    print("  0) Выход")
    
    auth_choice = input("\nВведите номер (1-2): ").strip()
    
    if auth_choice == '0':
        print("\nВыход из программы...")
        return None
    
    if auth_choice not in ['1', '2']:
        print("\n[!] Неверный выбор!")
        return None
    
    return auth_choice

def main():
    print("\n" + "!"*70)
    print("ВАЖНО:")
    print("  1. ОТКЛЮЧИТЕ VPN перед запуском!")
    print("  2. ЗАКРОЙТЕ Excel файл")
    print("  3. Убедитесь что API ключи WB указаны в листе 'Настройка' (столбец D)")
    print("  4. Артикулы WB должны быть в листе 'Данные для парсера ВБ'")
    print("!"*70)
    
    input("\nНажмите Enter чтобы начать...")
    
    # Показываем меню
    auth_choice = show_menu()
    
    if not auth_choice:
        return
    
    # Загружаем Excel
    try:
        wb = load_workbook(EXCEL_FILE)
    except Exception as e:
        print(f"\n[!] Ошибка открытия файла: {e}")
        return
    
    # Загружаем API ключи
    api_keys = load_api_keys(wb)
    
    try:
        with_auth = (auth_choice == '1')
        
        # WB
        if with_auth:
            parse_wb_with_auth(wb, api_keys)
        else:
            parse_wb_no_auth(wb, api_keys)
        
        print("\n" + "="*70)
        print("ВСЕ ЗАДАЧИ ВЫПОЛНЕНЫ!")
        print("="*70)
        
    except Exception as e:
        print(f"\n[!] Ошибка: {e}")
        import traceback
        traceback.print_exc()
    finally:
        wb.close()
        print("\n[DONE] Завершено!")

if __name__ == "__main__":
    main()

