# -*- coding: utf-8 -*-
"""
ПАРСЕР ВСЕХ ТОВАРОВ WB ИЗ ВСЕХ КАБИНЕТОВ
Автоматически загружает ВСЕ товары из 6 кабинетов и получает цены
"""

import os
import json
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
import time

# === КОНФИГУРАЦИЯ ===
EXCEL_FILE = "Парсер цен.xlsx"
SHEET_OUTPUT_WB = "Парсер ВБ"

# API ENDPOINTS
WB_PRICES_API_URL = "https://discounts-prices-api.wildberries.ru/api/v2/list/goods/filter"
WB_CONTENT_API_URL = "https://content-api.wildberries.ru/content/v2/get/cards/list"

# Названия кабинетов
CABINET_NAMES = ["COSMO", "MMA", "MAB", "MAU", "DREAMLAB", "BEAUTYLAB"]

# === ФУНКЦИИ ===

def load_api_keys_from_env():
    """Загружает API ключи из .env"""
    load_dotenv()
    
    api_keys = []
    cabinet_info = []
    
    for cabinet_name in CABINET_NAMES:
        api_key = os.getenv(cabinet_name, "").strip()
        if api_key:
            api_keys.append(api_key)
            cabinet_info.append(cabinet_name)
    
    print(f"\n[API] Загружено из .env файла:")
    print(f"    Кабинетов: {len(api_keys)}")
    for name in cabinet_info:
        print(f"      ✓ {name}")
    
    return api_keys, cabinet_info


def get_all_products_from_cabinet(api_key, cabinet_name):
    """
    Получает ВСЕ товары из одного кабинета
    Возвращает список товаров {nmID, title, vendorCode}
    """
    print(f"\n[{cabinet_name}] Загрузка всех товаров из кабинета...")
    
    products = []
    headers = {
        "Authorization": api_key,
        "Content-Type": "application/json"
    }
    
    cursor_updatedAt = ""
    cursor_nmID = 0
    page = 0
    
    try:
        while True:
            page += 1
            
            payload = {
                "settings": {
                    "cursor": {
                        "limit": 100
                    },
                    "filter": {
                        "withPhoto": -1
                    }
                }
            }
            
            # Добавляем курсор для пагинации
            if cursor_updatedAt and cursor_nmID:
                payload["settings"]["cursor"]["updatedAt"] = cursor_updatedAt
                payload["settings"]["cursor"]["nmID"] = cursor_nmID
            
            response = requests.post(WB_CONTENT_API_URL, headers=headers, json=payload, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                
                cards = data.get("cards", [])
                if not cards and "data" in data:
                    cards = data.get("data", {}).get("cards", [])
                
                if not cards:
                    break
                
                # Добавляем товары
                for card in cards:
                    nm_id = str(card.get("nmID", ""))
                    title = card.get("title") or card.get("object") or f"Товар {nm_id}"
                    vendor_code = str(card.get("vendorCode", ""))
                    
                    if nm_id:
                        products.append({
                            "nmID": nm_id,
                            "title": title,
                            "vendorCode": vendor_code,
                            "cabinet": cabinet_name
                        })
                
                print(f"    Страница {page}: +{len(cards)} товаров (всего: {len(products)})")
                
                # Курсор для следующей страницы
                cursor_data = data.get("cursor", {})
                cursor_updatedAt = cursor_data.get("updatedAt", "")
                cursor_nmID = cursor_data.get("nmID", 0)
                
                if not cursor_updatedAt or not cursor_nmID:
                    break
                
                time.sleep(0.2)
            
            elif response.status_code == 401:
                print(f"    [!] Ошибка 401: Неверный API ключ")
                break
            else:
                print(f"    [!] Ошибка {response.status_code}: {response.text[:200]}")
                break
        
        print(f"    ✓ Загружено {len(products)} товаров из {cabinet_name}")
        
    except Exception as e:
        print(f"    [!] Ошибка при загрузке товаров: {e}")
    
    return products


def get_prices_for_products(products, api_key, cabinet_name):
    """
    Получает цены для списка товаров из одного кабинета
    Возвращает словарь {nmID: {price_before_spp, price_after_spp, spp, discount}}
    """
    print(f"\n[{cabinet_name}] Загрузка цен для {len(products)} товаров...")
    
    prices_dict = {}
    headers = {
        "Authorization": api_key,
        "Content-Type": "application/json"
    }
    
    # Получаем список nmID
    nm_ids = [int(p["nmID"]) for p in products if p["nmID"].isdigit()]
    
    if not nm_ids:
        print(f"    [!] Нет валидных артикулов")
        return prices_dict
    
    try:
        # Обрабатываем батчами по 1000
        batch_size = 1000
        
        for i in range(0, len(nm_ids), batch_size):
            batch = nm_ids[i:i + batch_size]
            
            payload = {
                "limit": 1000,
                "offset": 0,
                "nmList": batch
            }
            
            response = requests.post(WB_PRICES_API_URL, headers=headers, json=payload, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                
                goods_list = []
                if "data" in data and "listGoods" in data["data"]:
                    goods_list = data["data"]["listGoods"]
                elif "listGoods" in data:
                    goods_list = data["listGoods"]
                
                for item in goods_list:
                    nm_id = str(item.get("nmID", ""))
                    
                    sizes = item.get("sizes", [])
                    if sizes and len(sizes) > 0:
                        size_data = sizes[0]
                        
                        price_original = size_data.get("price", 0)
                        price_discounted = size_data.get("discountedPrice", 0)
                        price_club = size_data.get("clubDiscountedPrice", 0)
                        
                        discount_percent = item.get("discount", 0)
                        club_discount_percent = item.get("clubDiscount", 0)
                        
                        if not price_discounted and price_original:
                            price_discounted = price_original
                        
                        if not price_club and price_discounted:
                            price_club = price_discounted
                        
                        if nm_id:
                            prices_dict[nm_id] = {
                                "price_original": float(price_original) if price_original else 0,
                                "price_before_spp": float(price_discounted) if price_discounted else 0,
                                "price_after_spp": float(price_club) if price_club else 0,
                                "discount": float(discount_percent) if discount_percent else 0,
                                "spp": float(club_discount_percent) if club_discount_percent else 0
                            }
                
                print(f"    Батч {i//batch_size + 1}: получено цен для {len(goods_list)} товаров")
            
            else:
                print(f"    [!] Ошибка {response.status_code}: {response.text[:200]}")
            
            time.sleep(0.3)
        
        print(f"    ✓ Загружено цен для {len(prices_dict)} товаров")
    
    except Exception as e:
        print(f"    [!] Ошибка при загрузке цен: {e}")
    
    return prices_dict


def main():
    print("\n" + "="*80)
    print("ПАРСЕР ВСЕХ ТОВАРОВ WB ИЗ ВСЕХ КАБИНЕТОВ")
    print("="*80)
    print("\n⚡ Этот скрипт автоматически:")
    print("  1. Загрузит ВСЕ товары из всех 6 кабинетов")
    print("  2. Получит цены ДО и ПОСЛЕ СПП для каждого товара")
    print("  3. Сохранит результаты в Excel")
    print("\n⏱️  Примерное время: 5-10 минут для ~1000 товаров")
    print("="*80)
    
    input("\n💡 Нажмите Enter чтобы начать...")
    
    # Загружаем API ключи
    api_keys, cabinet_names = load_api_keys_from_env()
    
    if not api_keys:
        print("\n[!] ОШИБКА: Не найдено API ключей в .env!")
        return
    
    start_time = time.time()
    
    # ШАГ 1: Загружаем все товары из всех кабинетов
    print("\n" + "="*80)
    print("[ШАГ 1/3] ЗАГРУЗКА ВСЕХ ТОВАРОВ ИЗ КАБИНЕТОВ")
    print("="*80)
    
    all_products = []
    
    for api_key, cabinet_name in zip(api_keys, cabinet_names):
        products = get_all_products_from_cabinet(api_key, cabinet_name)
        all_products.extend(products)
    
    print(f"\n✓ ИТОГО загружено товаров из всех кабинетов: {len(all_products)}")
    
    if not all_products:
        print("\n[!] Не найдено ни одного товара!")
        return
    
    # ШАГ 2: Загружаем цены для всех товаров
    print("\n" + "="*80)
    print("[ШАГ 2/3] ЗАГРУЗКА ЦЕН ДЛЯ ВСЕХ ТОВАРОВ")
    print("="*80)
    
    # Группируем товары по кабинетам
    products_by_cabinet = {}
    for product in all_products:
        cabinet = product["cabinet"]
        if cabinet not in products_by_cabinet:
            products_by_cabinet[cabinet] = []
        products_by_cabinet[cabinet].append(product)
    
    all_prices = {}
    
    for cabinet_name in cabinet_names:
        if cabinet_name in products_by_cabinet:
            # Находим API ключ для этого кабинета
            idx = cabinet_names.index(cabinet_name)
            api_key = api_keys[idx]
            
            products = products_by_cabinet[cabinet_name]
            prices = get_prices_for_products(products, api_key, cabinet_name)
            all_prices.update(prices)
    
    print(f"\n✓ ИТОГО загружено цен: {len(all_prices)}")
    
    # ШАГ 3: Сохраняем результаты в Excel
    print("\n" + "="*80)
    print("[ШАГ 3/3] СОХРАНЕНИЕ РЕЗУЛЬТАТОВ В EXCEL")
    print("="*80)
    
    try:
        # Открываем или создаем Excel файл
        try:
            wb = load_workbook(EXCEL_FILE)
            if SHEET_OUTPUT_WB not in wb.sheetnames:
                ws = wb.create_sheet(SHEET_OUTPUT_WB)
            else:
                ws = wb[SHEET_OUTPUT_WB]
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = SHEET_OUTPUT_WB
        
        # Заголовки (если лист пустой)
        if ws.max_row == 1:
            ws.append(["Дата", "Кабинет", "Артикул", "Название", "Цена До СПП", "Цена После СПП", "СПП %", "Скидка %"])
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        saved_count = 0
        
        for product in all_products:
            nm_id = product["nmID"]
            title = product["title"]
            cabinet = product["cabinet"]
            
            prices = all_prices.get(nm_id, {})
            
            price_before = prices.get("price_before_spp", 0)
            price_after = prices.get("price_after_spp", 0)
            discount = prices.get("discount", 0)
            spp = prices.get("spp", 0)
            
            # Считаем процент СПП
            spp_percent_calc = None
            if price_before and price_after and price_before > 0:
                spp_percent_calc = ((price_before - price_after) / price_before) * 100
            
            new_row = [
                timestamp,
                cabinet,
                nm_id,
                title,
                price_before if price_before else None,
                price_after if price_after else None,
                spp_percent_calc if spp_percent_calc else spp,
                discount if discount else None
            ]
            ws.append(new_row)
            saved_count += 1
            
            if saved_count % 100 == 0:
                print(f"    Сохранено: {saved_count}/{len(all_products)}")
        
        wb.save(EXCEL_FILE)
        print(f"\n✓ Сохранено {saved_count} товаров в '{EXCEL_FILE}'")
        
    except Exception as e:
        print(f"\n[!] Ошибка при сохранении: {e}")
        import traceback
        traceback.print_exc()
    
    # Итоговая статистика
    elapsed = time.time() - start_time
    
    print("\n" + "="*80)
    print("ГОТОВО!")
    print("="*80)
    print(f"Всего товаров загружено: {len(all_products)}")
    print(f"Цены получены для: {len(all_prices)} товаров")
    print(f"Время выполнения: {elapsed:.1f} сек ({elapsed/60:.1f} мин)")
    print(f"Средняя скорость: {len(all_products)/elapsed:.1f} товаров/сек")
    print("="*80)
    
    # Статистика по кабинетам
    print("\n📊 Статистика по кабинетам:")
    for cabinet_name in cabinet_names:
        if cabinet_name in products_by_cabinet:
            count = len(products_by_cabinet[cabinet_name])
            print(f"  {cabinet_name}: {count} товаров")
    
    print("\n[DONE] Завершено!")


if __name__ == "__main__":
    main()



